"""
core.py — Журнал ВСГУТУ.

Архитектура:
  - Все операции чтения/записи идут СНАЧАЛА в локальный SQLite (мгновенно).
  - Фоновый поток синхронизирует SQLite → PostgreSQL (Railway) без блокировки UI.
  - При запуске данные загружаются из PostgreSQL в SQLite (если подключён).
  - Оффлайн: работаем с SQLite, при восстановлении сети — автосинхронизация.
"""
import sqlite3
import os
import uuid
import re
import threading
import queue
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

APP_VERSION = "Release 2.4.2"

import os as _os
import sys as _sys
import shutil as _shutil
import socket as _socket


# ─────────────────────────────────────────────────────────────
#  Где лежит локальная база.
#  ВАЖНО: SQLite ВСЕГДА на локальном диске машины — никогда на сетевой шаре
#  (иначе блокировки и порча файла при работе нескольких ПК). Общее состояние
#  между ПК — только через PostgreSQL. Папка — в профиле пользователя.
# ─────────────────────────────────────────────────────────────
def _local_data_dir() -> str:
    if _sys.platform == "win32":
        base = _os.environ.get("LOCALAPPDATA") or _os.environ.get("APPDATA") \
            or _os.path.expanduser("~")
    elif _sys.platform == "darwin":
        base = _os.path.join(_os.path.expanduser("~"), "Library", "Application Support")
    else:
        base = _os.environ.get("XDG_DATA_HOME") \
            or _os.path.join(_os.path.expanduser("~"), ".local", "share")
    d = _os.path.join(base, "GradeBookAI")
    try:
        _os.makedirs(d, exist_ok=True)
    except Exception:
        d = _os.getcwd()
    return d


def _is_network_path(path: str) -> bool:
    """Эвристика: UNC-путь (\\\\server\\share) или сетевой диск Windows."""
    p = _os.path.abspath(path)
    if p.startswith("\\\\") or p.startswith("//"):
        return True
    if _sys.platform == "win32":
        try:
            import ctypes
            drive = _os.path.splitdrive(p)[0] + "\\"
            return ctypes.windll.kernel32.GetDriveTypeW(drive) == 4  # DRIVE_REMOTE
        except Exception:
            return False
    return False


DATA_DIR = _local_data_dir()
LOCAL_DB = _os.path.join(DATA_DIR, "vsgutu_grades.db")
BACKUP_DIR = _os.path.join(DATA_DIR, "backups")
DEVICE_ID = (_socket.gethostname() or "pc")[:64]   # имя ПК — для детекта конфликтов
MAX_BACKUPS = 30                                    # сколько бэкапов держим

if _is_network_path(LOCAL_DB):
    print("[DBManager] ВНИМАНИЕ: локальная база на сетевом пути — это вызывает "
          "блокировки и порчу. Используйте локальный диск.")


# ─────────────────────────────────────────────────────────────
#  Фоновый синхронизатор SQLite → PostgreSQL
# ─────────────────────────────────────────────────────────────
class PGSyncer:
    """
    Фоновый поток который берёт задачи из очереди и выполняет их в PostgreSQL.
    UI никогда не ждёт — задачи добавляются в очередь и выполняются асинхронно.
    """
    def __init__(self):
        self._queue   = queue.Queue()
        self._thread  = None
        self._running = False
        self._pg_conn = None

    def start(self):
        if self._running:
            return
        self._running = True
        self._thread  = threading.Thread(target=self._worker, daemon=True)
        self._thread.start()

    def stop(self):
        self._running = False
        self._queue.put(None)  # сигнал остановки

    def push(self, sql: str, params: tuple = ()):
        """Добавляет SQL-задачу в очередь для выполнения в PostgreSQL."""
        self._queue.put((sql, params))

    def push_many(self, tasks: list):
        """Добавляет список (sql, params) задач."""
        for task in tasks:
            self._queue.put(task)

    def _get_conn(self):
        """Возвращает живое соединение с PostgreSQL."""
        if self._pg_conn is not None:
            try:
                self._pg_conn.cursor().execute("SELECT 1")
                return self._pg_conn
            except Exception:
                self._pg_conn = None
        try:
            from db_config import get_pg_connection
            self._pg_conn = get_pg_connection()
            return self._pg_conn
        except Exception:
            return None

    def _worker(self):
        """Фоновый рабочий поток."""
        while self._running:
            try:
                task = self._queue.get(timeout=1)
                if task is None:
                    break
                sql, params = task
                conn = self._get_conn()
                if conn is None:
                    # Нет соединения — откладываем задачу обратно
                    self._queue.put(task)
                    time.sleep(5)
                    continue
                try:
                    cur = conn.cursor()
                    cur.execute(sql, params)
                    conn.commit()
                except Exception as e:
                    try:
                        conn.rollback()
                    except Exception:
                        self._pg_conn = None
                    print(f"[PGSyncer] Ошибка: {e}")
            except queue.Empty:
                continue
            except Exception as e:
                print(f"[PGSyncer] Критическая ошибка: {e}")


# Глобальный синхронизатор
_syncer = PGSyncer()


# ─────────────────────────────────────────────────────────────
#  Менеджер подключений
# ─────────────────────────────────────────────────────────────
class DBManager:
    _use_pg = False

    @classmethod
    def init(cls):
        """Вызывается при старте. Определяет режим и загружает данные из PG в SQLite."""
        cls._init_sqlite_tables()
        try:
            from db_config import is_pg_configured, is_key_activated
            if is_pg_configured() and is_key_activated():
                cls._use_pg = True
                cls._ensure_pg_tables()
                cls._pull_from_pg()   # загружаем данные из PG в локальный SQLite
                _syncer.start()
                return True
        except Exception as e:
            print(f"[DBManager] PG недоступен: {e}")
        cls._use_pg = False
        return False

    @classmethod
    def use_pg(cls) -> bool:
        return cls._use_pg

    @classmethod
    def get_conn(cls):
        """Всегда локальное SQLite соединение. WAL + busy_timeout снижают
        риск блокировок и порчи, если файл всё же оказался общим."""
        conn = sqlite3.connect(LOCAL_DB, timeout=10)
        try:
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA busy_timeout=5000")
            conn.execute("PRAGMA synchronous=NORMAL")
        except Exception:
            pass
        return conn

    # ─────────────────────────────────────────────────────────
    #  Авто-бэкапы локальной базы
    # ─────────────────────────────────────────────────────────
    @classmethod
    def backup(cls, reason: str = "") -> str:
        """
        Копирует файл базы в DATA_DIR/backups/ с меткой времени.
        Вызывается перед синхронизацией/закрытием. Возвращает путь к бэкапу
        ('' — если бэкапить нечего). Старые бэкапы (свыше MAX_BACKUPS) удаляются.
        """
        try:
            if not _os.path.exists(LOCAL_DB):
                return ""
            _os.makedirs(BACKUP_DIR, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            tag = ("_" + "".join(ch for ch in reason if ch.isalnum())[:16]) if reason else ""
            dst = _os.path.join(BACKUP_DIR, f"vsgutu_grades_{ts}{tag}.db")
            # WAL: гарантируем, что данные на диске, затем копируем
            try:
                c = sqlite3.connect(LOCAL_DB)
                c.execute("PRAGMA wal_checkpoint(FULL)")
                c.close()
            except Exception:
                pass
            _shutil.copy2(LOCAL_DB, dst)
            cls._prune_backups()
            return dst
        except Exception as e:
            print(f"[DBManager] бэкап не удался: {e}")
            return ""

    @classmethod
    def _prune_backups(cls):
        try:
            files = sorted(
                (f for f in _os.listdir(BACKUP_DIR) if f.endswith(".db")),
                reverse=True,
            )
            for old in files[MAX_BACKUPS:]:
                try:
                    _os.remove(_os.path.join(BACKUP_DIR, old))
                except Exception:
                    pass
        except Exception:
            pass

    @classmethod
    def list_backups(cls) -> list:
        """[(имя, полный_путь, размер_байт, mtime), ...] от свежих к старым."""
        out = []
        try:
            for f in _os.listdir(BACKUP_DIR):
                if not f.endswith(".db"):
                    continue
                p = _os.path.join(BACKUP_DIR, f)
                st = _os.stat(p)
                out.append((f, p, st.st_size, st.st_mtime))
        except Exception:
            pass
        out.sort(key=lambda x: x[3], reverse=True)
        return out

    @classmethod
    def restore(cls, backup_path: str) -> bool:
        """Восстанавливает базу из бэкапа. Перед этим делает бэкап текущей."""
        try:
            if not _os.path.exists(backup_path):
                return False
            cls.backup(reason="before_restore")
            _shutil.copy2(backup_path, LOCAL_DB)
            return True
        except Exception as e:
            print(f"[DBManager] восстановление не удалось: {e}")
            return False

    # ─────────────────────────────────────────────────────────
    #  Конфликты синхронизации (детект вместо тихой перезаписи)
    # ─────────────────────────────────────────────────────────
    @classmethod
    def list_conflicts(cls, unresolved_only: bool = True) -> list:
        """Список конфликтов оценок: [dict(...), ...]."""
        out = []
        try:
            conn = cls.get_conn(); cur = conn.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS sync_conflicts ("
                        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
                        "student_f TEXT, student_n TEXT, lesson_id TEXT,"
                        "local_grade TEXT, remote_grade TEXT, remote_device TEXT,"
                        "remote_at TEXT, detected_at TEXT, resolved INTEGER DEFAULT 0)")
            q = ("SELECT id,student_f,student_n,lesson_id,local_grade,remote_grade,"
                 "remote_device,remote_at,detected_at,resolved FROM sync_conflicts")
            if unresolved_only:
                q += " WHERE resolved=0"
            q += " ORDER BY detected_at DESC"
            cur.execute(q)
            cols = ["id", "student_f", "student_n", "lesson_id", "local_grade",
                    "remote_grade", "remote_device", "remote_at", "detected_at", "resolved"]
            out = [dict(zip(cols, r)) for r in cur.fetchall()]
            conn.close()
        except Exception as e:
            print(f"[DBManager] чтение конфликтов: {e}")
        return out

    @classmethod
    def resolve_conflict(cls, conflict_id: int, chosen_grade: str) -> bool:
        """
        Применяет выбранное преподавателем значение и закрывает конфликт.
        chosen_grade записывается как победитель (в SQLite + очередь в PG).
        """
        try:
            conn = cls.get_conn(); cur = conn.cursor()
            cur.execute("SELECT student_f,student_n,lesson_id FROM sync_conflicts "
                        "WHERE id=?", (conflict_id,))
            row = cur.fetchone()
            if not row:
                conn.close(); return False
            f, n, lid = row
            now = datetime.now().isoformat(timespec="seconds")
            cur.execute("INSERT OR REPLACE INTO grades "
                        "(student_f,student_n,lesson_id,grade,updated_at,device) "
                        "VALUES (?,?,?,?,?,?)", (f, n, lid, chosen_grade, now, DEVICE_ID))
            cur.execute("UPDATE sync_conflicts SET resolved=1 WHERE id=?", (conflict_id,))
            conn.commit(); conn.close()
            if cls._use_pg:
                _syncer.push(
                    "INSERT INTO grades (student_f,student_n,lesson_id,grade,updated_at,device)"
                    " VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (student_f,student_n,lesson_id)"
                    " DO UPDATE SET grade=EXCLUDED.grade, updated_at=EXCLUDED.updated_at,"
                    " device=EXCLUDED.device",
                    (f, n, lid, chosen_grade, now, DEVICE_ID))
            return True
        except Exception as e:
            print(f"[DBManager] resolve_conflict: {e}")
            return False

    @classmethod
    def placeholder(cls) -> str:
        return "?"  # всегда SQLite локально

    @classmethod
    def _init_sqlite_tables(cls):
        conn = sqlite3.connect(LOCAL_DB)
        cur  = conn.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS lessons
            (id TEXT PRIMARY KEY, group_name TEXT, subject TEXT,
             type TEXT, number INTEGER, topic TEXT, date TEXT,
             retake_date TEXT DEFAULT '', hour INTEGER DEFAULT 0)""")
        for col, default in [("retake_date", "TEXT DEFAULT ''"), ("hour", "INTEGER DEFAULT 0")]:
            try:
                cur.execute(f"ALTER TABLE lessons ADD COLUMN {col} {default}")
            except Exception:
                pass
        cur.execute("""CREATE TABLE IF NOT EXISTS students
            (f TEXT, n TEXT, group_name TEXT, PRIMARY KEY(f, n, group_name))""")
        cur.execute("""CREATE TABLE IF NOT EXISTS grades
            (student_f TEXT, student_n TEXT, lesson_id TEXT, grade TEXT,
             updated_at TEXT DEFAULT '', device TEXT DEFAULT '',
             PRIMARY KEY(student_f, student_n, lesson_id))""")
        # миграция старых баз без новых колонок
        for col in ("updated_at", "device"):
            try:
                cur.execute(f"ALTER TABLE grades ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass
        cur.execute("""CREATE TABLE IF NOT EXISTS sync_conflicts
            (id INTEGER PRIMARY KEY AUTOINCREMENT,
             student_f TEXT, student_n TEXT, lesson_id TEXT,
             local_grade TEXT, remote_grade TEXT, remote_device TEXT,
             remote_at TEXT, detected_at TEXT, resolved INTEGER DEFAULT 0)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS kv_store
            (key TEXT PRIMARY KEY, value TEXT NOT NULL)""")
        conn.commit()
        conn.close()

    @classmethod
    def _ensure_pg_tables(cls):
        """Создаёт таблицы в PostgreSQL если их нет."""
        try:
            from db_config import get_pg_connection
            conn = get_pg_connection()
            cur  = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lessons (
                    id TEXT PRIMARY KEY, group_name TEXT NOT NULL,
                    subject TEXT NOT NULL, type TEXT NOT NULL,
                    number INTEGER NOT NULL, topic TEXT DEFAULT '',
                    date TEXT NOT NULL, retake_date TEXT DEFAULT '',
                    hour INTEGER DEFAULT 0,
                    retake_date_2 TEXT DEFAULT '', retake_date_3 TEXT DEFAULT '',
                    retake_date_4 TEXT DEFAULT '', retake_date_5 TEXT DEFAULT ''
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS students (
                    f TEXT NOT NULL, n TEXT NOT NULL, group_name TEXT NOT NULL,
                    PRIMARY KEY (f, n, group_name)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS grades (
                    student_f TEXT NOT NULL, student_n TEXT NOT NULL,
                    lesson_id TEXT NOT NULL, grade TEXT NOT NULL,
                    updated_at TEXT DEFAULT '', device TEXT DEFAULT '',
                    PRIMARY KEY (student_f, student_n, lesson_id)
                )
            """)
            for col in ("updated_at", "device"):
                try:
                    cur.execute(f"ALTER TABLE grades ADD COLUMN {col} TEXT DEFAULT ''")
                except Exception:
                    pass
            cur.execute("""
                CREATE TABLE IF NOT EXISTS secure_store (
                    key_name TEXT PRIMARY KEY, value_enc TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("CREATE TABLE IF NOT EXISTS subjects (name TEXT PRIMARY KEY, created_at TIMESTAMP DEFAULT NOW())")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS kv_store (
                    key TEXT PRIMARY KEY, value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS pc_keys (
                    id SERIAL PRIMARY KEY, key_value TEXT NOT NULL,
                    key_hash TEXT NOT NULL UNIQUE, pc_name TEXT DEFAULT '',
                    active BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS install_keys (
                    id SERIAL PRIMARY KEY, key_value TEXT NOT NULL,
                    key_hash TEXT NOT NULL UNIQUE, pc_name TEXT DEFAULT '',
                    active BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[DBManager] Ошибка создания PG таблиц: {e}")

    @classmethod
    def _pull_from_pg(cls):
        """Загружает данные из PostgreSQL в локальный SQLite при старте.
        Перед загрузкой делает бэкап. Для grades — НЕ слепая перезапись:
        расхождения локального и серверного значений фиксируются как конфликты."""
        cls.backup(reason="before_pull")
        try:
            from db_config import get_pg_connection
            pg   = get_pg_connection()
            pgc  = pg.cursor()
            loc  = sqlite3.connect(LOCAL_DB)
            lc   = loc.cursor()

            # Добавляем недостающие колонки в SQLite
            pgc.execute("SELECT column_name FROM information_schema.columns WHERE table_name='lessons' ORDER BY ordinal_position")
            pg_cols = [r[0] for r in pgc.fetchall()]
            for col in pg_cols:
                if col not in ("id","group_name","subject","type","number","topic","date","retake_date","hour"):
                    try:
                        lc.execute(f"ALTER TABLE lessons ADD COLUMN {col} TEXT DEFAULT ''")
                    except Exception:
                        pass

            # Lessons
            col_str  = ", ".join(pg_cols) if pg_cols else "id,group_name,subject,type,number,topic,date,retake_date,hour"
            ph_str   = ", ".join(["?"] * len(pg_cols)) if pg_cols else "?,?,?,?,?,?,?,?,?"
            pgc.execute(f"SELECT {col_str} FROM lessons")
            rows = pgc.fetchall()
            for r in rows:
                lc.execute(f"INSERT OR REPLACE INTO lessons ({col_str}) VALUES ({ph_str})", r)

            # Students
            pgc.execute("SELECT f, n, group_name FROM students")
            for r in pgc.fetchall():
                lc.execute("INSERT OR IGNORE INTO students (f,n,group_name) VALUES (?,?,?)", r)

            # Grades — БЕЗ слепой перезаписи. Берём serverное значение и
            # сравниваем с локальным. Если локально ДРУГОЕ значение и оно
            # изменено позднее серверного (или сервер не знает времени) —
            # это конфликт: фиксируем и НЕ затираем работу преподавателя.
            try:
                pgc.execute("SELECT student_f, student_n, lesson_id, grade, "
                            "COALESCE(updated_at,''), COALESCE(device,'') FROM grades")
                remote_rows = pgc.fetchall()
            except Exception:
                pgc.execute("SELECT student_f, student_n, lesson_id, grade FROM grades")
                remote_rows = [(a, b, c, d, "", "") for (a, b, c, d) in pgc.fetchall()]

            now_iso = datetime.now().isoformat(timespec="seconds")
            for rf, rn, rlid, rgrade, rat, rdev in remote_rows:
                lc.execute("SELECT grade, COALESCE(updated_at,'') FROM grades "
                           "WHERE student_f=? AND student_n=? AND lesson_id=?",
                           (rf, rn, rlid))
                local = lc.fetchone()
                if local is None:
                    lc.execute("INSERT OR REPLACE INTO grades "
                               "(student_f,student_n,lesson_id,grade,updated_at,device) "
                               "VALUES (?,?,?,?,?,?)", (rf, rn, rlid, rgrade, rat, rdev))
                    continue
                lgrade, lat = local
                if (lgrade or "") == (rgrade or ""):
                    continue
                if lat and rat and lat > rat:
                    continue   # локальное новее — оставляем, само уйдёт в PG
                if lat and rat and lat < rat:
                    lc.execute("INSERT OR REPLACE INTO grades "
                               "(student_f,student_n,lesson_id,grade,updated_at,device) "
                               "VALUES (?,?,?,?,?,?)", (rf, rn, rlid, rgrade, rat, rdev))
                    continue
                # времени нет/равны, значения разные → НАСТОЯЩИЙ конфликт
                lc.execute(
                    "INSERT INTO sync_conflicts "
                    "(student_f,student_n,lesson_id,local_grade,remote_grade,"
                    "remote_device,remote_at,detected_at,resolved) "
                    "VALUES (?,?,?,?,?,?,?,?,0)",
                    (rf, rn, rlid, lgrade, rgrade, rdev, rat, now_iso))
                # локальное значение НЕ трогаем — препод решит вручную

            # KV-хранилище (студенты/учителя/группы/конфиг) — PG источник правды
            try:
                pgc.execute("SELECT key, value FROM kv_store")
                for k, v in pgc.fetchall():
                    lc.execute("INSERT OR REPLACE INTO kv_store (key, value) VALUES (?, ?)", (k, v))
            except Exception as e:
                print(f"[DBManager] kv_store pull: {e}")

            loc.commit()
            loc.close()
            pg.close()
            print(f"[DBManager] Синхронизировано из PostgreSQL: {len(rows)} занятий")
        except Exception as e:
            print(f"[DBManager] Ошибка загрузки из PG: {e}")

    @classmethod
    def sync_to_pg(cls, sql_sqlite: str, params: tuple, sql_pg: str = None, params_pg: tuple = None):
        """
        Выполняет SQL в SQLite немедленно.
        Добавляет задачу в очередь для асинхронного выполнения в PostgreSQL.
        """
        if sql_pg is None:
            # Конвертируем ? → %s для PostgreSQL
            sql_pg = sql_sqlite.replace("?", "%s")
        if params_pg is None:
            params_pg = params
        if cls._use_pg:
            _syncer.push(sql_pg, params_pg)

    @classmethod
    def upsert_lesson(cls, cur, vals: tuple):
        """INSERT OR REPLACE для занятия в SQLite. Асинхронно в PG."""
        cur.execute("INSERT OR REPLACE INTO lessons (id,group_name,subject,type,number,topic,date,retake_date,hour) VALUES (?,?,?,?,?,?,?,?,?)", vals[:9])
        if cls._use_pg:
            _syncer.push("""
                INSERT INTO lessons (id,group_name,subject,type,number,topic,date,retake_date,hour)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                    group_name=EXCLUDED.group_name, subject=EXCLUDED.subject,
                    type=EXCLUDED.type, number=EXCLUDED.number, topic=EXCLUDED.topic,
                    date=EXCLUDED.date, retake_date=EXCLUDED.retake_date, hour=EXCLUDED.hour
            """, vals[:9])

    @classmethod
    def upsert_student(cls, cur, vals: tuple):
        cur.execute("INSERT OR IGNORE INTO students (f,n,group_name) VALUES (?,?,?)", vals)
        if cls._use_pg:
            _syncer.push("INSERT INTO students (f,n,group_name) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING", vals)

    @classmethod
    def upsert_grade(cls, cur, vals: tuple):
        """
        vals = (student_f, student_n, lesson_id, grade).
        Проставляем updated_at и device — это делает синхронизацию
        детерминированной (newest-wins по времени, а не «как повезёт»).
        """
        f, n, lid, grade = vals[:4]
        now = datetime.now().isoformat(timespec="seconds")
        cur.execute(
            "INSERT OR REPLACE INTO grades "
            "(student_f,student_n,lesson_id,grade,updated_at,device) "
            "VALUES (?,?,?,?,?,?)", (f, n, lid, grade, now, DEVICE_ID))
        if cls._use_pg:
            _syncer.push("""
                INSERT INTO grades (student_f,student_n,lesson_id,grade,updated_at,device)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT (student_f,student_n,lesson_id) DO UPDATE SET
                    grade=EXCLUDED.grade, updated_at=EXCLUDED.updated_at,
                    device=EXCLUDED.device
                WHERE grades.updated_at IS NULL OR grades.updated_at=''
                   OR grades.updated_at <= EXCLUDED.updated_at
            """, (f, n, lid, grade, now, DEVICE_ID))


# ─────────────────────────────────────────────────────────────
#  Датаклассы
# ─────────────────────────────────────────────────────────────
@dataclass
class Lesson:
    id: str
    type: str
    number: int
    topic: str
    date: str
    retake_date: str = ""
    hour: int = 0


@dataclass
class Student:
    n: str
    f: str
    group: str
    records: Dict[str, str] = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────
#  Журнал
# ─────────────────────────────────────────────────────────────
class GradeBook:
    def __init__(self, group: str, subject: str):
        self.group   = group
        self.subject = subject
        if not DBManager.use_pg():
            DBManager._init_sqlite_tables()
        self.lessons: List[Lesson] = []
        self.spisok_stud: List[Student] = []
        self.load_from_db()

    def add_student(self, st: Student):
        self.spisok_stud.append(st)
        self.save_to_db()

    def delete_student(self, surname: str, name: str):
        surname = surname.strip()
        name    = name.strip()
        self.spisok_stud = [
            s for s in self.spisok_stud
            if not (s.f.strip().lower() == surname.lower()
                    and s.n.strip().lower() == name.lower())
        ]
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        cur.execute("DELETE FROM grades WHERE student_f=? AND student_n=?", (surname, name))
        cur.execute("DELETE FROM students WHERE f=? AND n=?", (surname, name))
        conn.commit()
        conn.close()
        if DBManager.use_pg():
            _syncer.push("DELETE FROM grades WHERE student_f=%s AND student_n=%s", (surname, name))
            _syncer.push("DELETE FROM students WHERE f=%s AND n=%s", (surname, name))

    def add_lesson(self, lesson_type: str, topic: str = "", date: str = "", hour: int = 0) -> "Lesson":
        nums     = [l.number for l in self.lessons if l.type == lesson_type and getattr(l, 'hour', 0) in (0, 1)]
        next_num = max(nums) + 1 if nums else 1
        dt = date or datetime.now().strftime('%d.%m.%Y')

        if lesson_type == "Лекция" and hour == 0:
            l1 = Lesson(id=str(uuid.uuid4()), type="Лекция", number=next_num, topic=topic, date=dt, retake_date="", hour=1)
            l2 = Lesson(id=str(uuid.uuid4()), type="Лекция", number=next_num, topic=topic, date=dt, retake_date="", hour=2)
            self.lessons.extend([l1, l2])
            self.save_to_db()
            return l1
        else:
            l = Lesson(id=str(uuid.uuid4()), type=lesson_type, number=next_num, topic=topic, date=dt, retake_date="", hour=hour)
            self.lessons.append(l)
            self.save_to_db()
            return l

    def set_retake_date(self, lesson_id: str, retake_date: str, retake_n: int = 1):
        attr = 'retake_date' if retake_n == 1 else f'retake_date_{retake_n}'
        for l in self.lessons:
            if l.id == lesson_id:
                setattr(l, attr, retake_date)
                break
        col = 'retake_date' if retake_n == 1 else f'retake_date_{retake_n}'
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        try:
            cur.execute(f"ALTER TABLE lessons ADD COLUMN {col} TEXT DEFAULT ''")
        except Exception:
            pass
        cur.execute(f"UPDATE lessons SET {col}=? WHERE id=?", (retake_date, lesson_id))
        conn.commit()
        conn.close()
        if DBManager.use_pg():
            _syncer.push(f"ALTER TABLE lessons ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''", ())
            _syncer.push(f"UPDATE lessons SET {col}=%s WHERE id=%s", (retake_date, lesson_id))

    def save_to_db(self):
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        for l in self.lessons:
            DBManager.upsert_lesson(cur, (
                l.id, self.group, self.subject, l.type,
                l.number, l.topic, l.date,
                getattr(l, 'retake_date', ''),
                getattr(l, 'hour', 0)
            ))
        for s in self.spisok_stud:
            DBManager.upsert_student(cur, (s.f, s.n, s.group))
            for lesson_id, grade in s.records.items():
                DBManager.upsert_grade(cur, (s.f, s.n, lesson_id, grade))
        conn.commit()
        conn.close()

    def load_from_db(self):
        conn = DBManager.get_conn()
        cur  = conn.cursor()

        cur.execute(
            "SELECT id, type, number, topic, date, retake_date, hour "
            "FROM lessons WHERE group_name=? AND subject=? ORDER BY type, number, hour",
            (self.group, self.subject)
        )
        self.lessons = []
        for row in cur.fetchall():
            l = Lesson(
                id=row[0], type=row[1], number=row[2],
                topic=row[3], date=row[4],
                retake_date=row[5] if row[5] else "",
                hour=row[6] if row[6] else 0
            )
            self.lessons.append(l)

        # ── Синхронизация студентов из GitHub ─────────────────────────────
        # Студенты управляются через GitHub (GradeBookGitHub.get_students).
        # TeacherDashboard._sync_students_from_gh() добавляет их в SQLite.
        # ВАЖНО: НЕ удаляем студентов из SQLite здесь — это приводило к тому,
        # что студенты, добавленные администратором через GitHub, исчезали
        # при каждой перезагрузке журнала.
        try:
            from data_store import get_store as get_gh_store
            gh = get_gh_store()
            if gh:
                gh_students = gh.get_students()
                existing = set()
                cur.execute("SELECT f, n FROM students WHERE group_name=?", (self.group,))
                for row in cur.fetchall():
                    existing.add((row[0].strip().lower(), row[1].strip().lower()))
                for s in gh_students:
                    if s.get("group", "") != self.group:
                        continue
                    surname = s.get("surname", "").strip()
                    name    = s.get("name", "").strip()
                    if surname and name and (surname.lower(), name.lower()) not in existing:
                        DBManager.upsert_student(cur, (surname, name, self.group))
                        existing.add((surname.lower(), name.lower()))
                conn.commit()
        except Exception as e:
            print(f"[GradeBook] GitHub student sync в load_from_db: {e}")

        cur.execute("SELECT f, n, group_name FROM students WHERE group_name=?", (self.group,))
        self.spisok_stud = []
        for f, n, g in cur.fetchall():
            student = Student(n, f, g)
            cur.execute("SELECT lesson_id, grade FROM grades WHERE student_f=? AND student_n=?", (f, n))
            student.records = {row[0]: row[1] for row in cur.fetchall()}
            self.spisok_stud.append(student)
        conn.close()

    # Обратная совместимость
    def init_db(self): pass
    def save_to_sqlite(self): self.save_to_db()
    def load_from_sqlite(self): self.load_from_db()

    def calculate_average(self, student: Student) -> float:
        """Средний балл через единый модуль grading (та же формула, что у Вектора).
        Методика (Н=вес, учитывать ли пропуск, включать ли экзамены) берётся
        из config, дефолты сохраняют прежнее поведение."""
        import grading
        try:
            from data_store import get_store
            cfg = get_store()._config()
        except Exception:
            cfg = {}
        return grading.practice_average(
            grading.pairs_from_objects(self.lessons), student.records, cfg)

    def export_to_excel(self, file_path: str):
        wb = Workbook()
        ws = wb.active
        title      = f"Успеваемость {self.group}"
        safe_title = re.sub(r'[\[\]:*?/\\]', '_', title)
        ws.title   = safe_title[:31]
        headers = ["Фамилия", "Имя"]
        for l in self.lessons:
            if l.type == "Экзамен":
                headers.append(f"Экзамен №{l.number}\n({l.date})\n{l.topic}")
                if l.retake_date:
                    headers.append(f"Пересдача\n({l.retake_date})")
            else:
                headers.append(f"{l.type} №{l.number}\n({l.date})")
        headers.append("Средний балл")
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for s in self.spisok_stud:
            row = [s.f, s.n]
            for l in self.lessons:
                base = s.records.get(l.id, "")
                row.append(base)
                if l.type == "Экзамен" and l.retake_date:
                    rv = s.records.get(l.id + "_retake", "")
                    if not rv:
                        # пересдача только у заваливших основной экзамен
                        b = (base or "").strip()
                        failed = b.startswith(("2", "Н")) or "Не зачтено" in b
                        rv = "" if failed else "—"
                    row.append(rv)
            row.append(round(self.calculate_average(s), 2))
            ws.append(row)
        wb.save(file_path)
