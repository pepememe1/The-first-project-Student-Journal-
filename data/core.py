"""
core.py — Журнал ВСГУТУ.

Архитектура (offline-first):
  - Все операции чтения/записи идут в локальный SQLite (мгновенно, без блокировки UI).
  - Обмен с общей базой колледжа идёт через REST API-сервер ВСГУТУ — отдельным
    фоновым потоком (см. sync_runner), а не прямым подключением к БД с клиента.
  - Оффлайн: работаем с SQLite, при восстановлении сети — автосинхронизация.
"""
#⚠️ Прямого `import sqlite3` здесь БОЛЬШЕ НЕТ и быть не должно: соединение к
#локальной базе открывает ТОЛЬКО `local_db.connect` (иначе файл откроется без
#ключа SQLCipher и ляжет открытым текстом). Держит
#`test_no_direct_sqlite_connect_to_the_sync_database`.
from data import local_db
import log
import uuid
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

#Строка версии живёт в ОДНОМ месте — корневом `desktop_update.py` (там же, где логика
#сравнения версий). Здесь ре-экспорт, чтобы `core.APP_VERSION` продолжал работать у всех,
#кто им пользуется: заголовок окна, автообновление, раздел «Сервер».
#⚠️ Правится ТАМ. Держать её в нескольких местах уже пробовали — расходилась трижды
#(3.5.3, 3.5.6 и 3.7, где заголовок показывал 3.6.9, а сайт — 3.6.1).
from desktop_update import APP_VERSION  # noqa: F401  (ре-экспорт, см. выше)

import os as _os
import sys as _sys
import shutil as _shutil
import socket as _socket
import app_paths


#Об одном и том же сбое сообщаем ОДИН раз за запуск. Часть мест ниже зовётся на КАЖДОЕ
#соединение с базой, на каждое открытие журнала и на каждый пересчёт среднего — писать
#по строке на вызов значит превратить лог в шум, в котором не найти ни одной настоящей
#записи, и его перестанут читать. Тот же приём и та же причина, что в data/app_settings.py.
_reported: set[str] = set()


def _report_once(tag: str, message: str, *args, level: str = "warning") -> None:
    """Записать сбой в лог один раз за запуск (ключ повтора — tag)."""
    if tag in _reported:
        return
    _reported.add(tag)
    getattr(log.get("core"), level)(message, *args)


def _alter_ignored(table: str, column: str, exc: Exception) -> None:
    """Не прошёл идемпотентный `ALTER TABLE ... ADD COLUMN` при миграции схемы.

    «duplicate column name» — ШТАТНЫЙ исход: миграция зовётся на каждом старте, и на
    уже мигрированной базе она обязана падать именно так. Такой случай не логируем
    совсем — иначе каждый запуск давал бы полтора десятка строк-пустышек.

    ЛЮБАЯ другая причина (залоченная база, нет прав на файл) оставляет таблицу БЕЗ
    колонки, и дальше запросы к ней падают «непонятно почему», уже далеко от места
    настоящей ошибки. Класс дефекта «колонка без ALTER» у нас уже случался, поэтому
    поведение не меняем (падать на старте из-за миграции нельзя), но след оставляем."""
    if "duplicate column name" in str(exc).lower():
        return
    _report_once(f"alter:{table}.{column}",
                 "[DBManager] миграция схемы: колонка «%s» не добавлена в «%s» (%s) — "
                 "запросы к этой колонке будут падать", column, table, exc, level="error")


#Где лежит локальная база.
#ВАЖНО: SQLite ВСЕГДА на локальном диске машины — никогда на сетевой шаре
#(иначе блокировки и порча файла при работе нескольких ПК). Где именно лежит
#папка — решает app_paths: рядом с .exe (портативно) или в профиле (dev).


def _is_network_path(path: str) -> bool:
    """Эвристика: UNC-путь (\\\\server\\share) или сетевой диск Windows."""
    p = _os.path.abspath(path)
    if p.startswith("\\\\") or p.startswith("//"):
        return True
    if _sys.platform == "win32":
        try:
            import ctypes
            drive = _os.path.splitdrive(p)[0] + "\\"
            return ctypes.windll.kernel32.GetDriveTypeW(drive) == 4  # DRIVE_REMOTE
        except Exception as e:
            #Тип диска не определился — считаем путь локальным (как и раньше), иначе
            #программа не запустилась бы из-за неудавшейся проверки. Но молчать нельзя:
            #предупреждение о базе на сетевой шаре, ради которого функция и написана,
            #в этом случае не сработает НИКОГДА, и блокировки/порчу файла будут искать
            #где угодно, только не здесь.
            _report_once("drive_type", "[DBManager] тип диска для «%s» не определён (%s) — "
                         "проверка «база на сетевом пути» пропущена", p, e)
            return False
    return False


def _db_key() -> str:
    """Ключ шифрования базы ('' — на этой машине шифровать нечем).

    Тот же самый ключ устройства, которым уже шифруется копия локального сервера
    (`local_app_*.enc.db`). Второй ключ на одну машину означал бы, что одна из двух
    баз однажды не откроется — см. докстринг `data/device_key.py`."""
    from data import device_key
    return device_key.db_key()


DATA_DIR = app_paths.data_dir()
LOCAL_DB = _os.path.join(DATA_DIR, "vsgutu_grades.db")
BACKUP_DIR = _os.path.join(DATA_DIR, "backups")
DEVICE_ID = (_socket.gethostname() or "pc")[:64]   #имя ПК — для детекта конфликтов
MAX_BACKUPS = 48                                    #сколько бэкапов держим (~сутки при цикле 30 мин)
AUTO_BACKUP_INTERVAL_SEC = 30 * 60                  #не чаще одного авто-бэкапа в 30 минут

if _is_network_path(LOCAL_DB):
    log.get("core").warning("[DBManager] ВНИМАНИЕ: локальная база на сетевом пути — это вызывает "
          "блокировки и порчу. Используйте локальный диск.")


def resolve_student_id(f: str, n: str, group: str = "") -> str:
    """ФИО → неизменяемый id студента (строка users), '' — если не нашли.

    Этап 1 миграции оценок с ФИО-ключей на student_id. Оценка по-прежнему ключуется по
    ФИО, но ДОПОЛНИТЕЛЬНО несёт id: ФИО меняется (замужество, исправление опечатки), id
    нет. Пустой ответ не ошибка — студента могло не быть в справочнике (ростер препода
    ведётся отдельно), и тогда связь остаётся прежней, по ФИО.

    Импорт data_store ЛЕНИВЫЙ и намеренно: владелец таблицы users — он, а он сам стоит
    на core. Импорт на уровне модуля замкнул бы кольцо.
    """
    try:
        from data.data_store import student_id_by_name
        return student_id_by_name(f, n, group)
    except Exception as e:
        log.get("core").warning(f"[student_id] не удалось определить id для {f} {n}: {e}")
        return ""


#Ключи оценок — ЭТАП 3 миграции. ЗЕРКАЛО server/app/models.py::grade_id/term_grade_id:
#формат обязан совпадать до символа, иначе десктоп и сервер будут считать одну и ту же
#оценку разными записями и наплодят дубли. Правишь здесь — правь и там.
def grade_id(student_id: str, lesson_id: str) -> str:
    """Ключ оценки за занятие: `{student_id}|{lesson_id}`."""
    return f"{student_id}|{lesson_id}"


def term_grade_id(student_id: str, subject: str, year: str, semester) -> str:
    """Ключ итоговой оценки: `{student_id}|{subject}|{year}|{semester}`."""
    return f"{student_id}|{subject}|{year}|{int(semester or 0)}"


def legacy_term_grade_id(f: str, n: str, subject: str, year: str, semester) -> str:
    """СТАРЫЙ ключ (по ФИО). Нужен только миграции — чтобы найти строки, которые ещё
    не переведены. В новом коде не использовать."""
    return f"{f}|{n}|{subject}|{year}|{semester}"


#Менеджер подключений
class DBManager:
    @classmethod
    def init(cls):
        """Вызывается при старте: создаёт локальные таблицы SQLite.

        Прямого подключения к серверной БД с клиента НЕТ — обмен с общей базой
        колледжа идёт через REST API-сервер (см. sync_runner). Offline-first
        сохраняется: приложение всегда работает на локальном SQLite, а синхронизация
        подхватывается фоном при наличии сети.

        🔒 ЗДЕСЬ ЖЕ ПРОИСХОДИТ ПЕРЕВОД БАЗЫ НА ШИФРОВАНИЕ, и место выбрано не
        случайно: это единственная точка, которая заведомо отрабатывает ДО первого
        обращения к данным и ровно один раз за запуск. Перевод идемпотентен —
        зашифрованную базу `encrypt_in_place` не трогает вовсе."""
        key = _db_key()
        if key and local_db.encrypt_in_place(LOCAL_DB, key):
            #Копии, снятые ДО перехода, — это полные снимки журнала открытым текстом.
            #Оставить их значило бы зашифровать дверь и не тронуть окно.
            local_db.purge_plaintext_backups(BACKUP_DIR)
        cls._init_sqlite_tables()
        #Разовая перекладка крупных значений в сжатую форму. На живых данных 92 %
        #файла занимал ОДИН ключ (кэш расписания), и сжимается он в 19.6 раза.
        local_db.repack_large_values(LOCAL_DB, key)
        if not key:
            #Молчать нельзя: человек вправе считать, что его данные защищены. Из .exe
            #эта ветка не срабатывает (драйвер вшит) — она про запуск из исходников.
            log.get("core").warning(
                "[DBManager] драйвера sqlcipher3 нет — локальная база НЕ ЗАШИФРОВАНА, "
                "ФИО и оценки в файле читаются любым просмотрщиком")
        print("ℹ️  Локальный SQLite (синхронизация с сервером — через API)")

    @classmethod
    def get_conn(cls):
        """Всегда локальное соединение — ЗАШИФРОВАННОЕ, если на машине есть чем.

        ⚠️ Единственная точка, где эта база открывается. Раньше `sqlite3.connect`
        стоял в пяти местах, и добавить шифрование «в основном пути», забыв про
        резервное копирование или восстановление, было бы очень легко — а забытое
        место как раз и оставляло бы файл с ФИО открытым.
        WAL, busy_timeout и secure_delete ставит сам `local_db.connect`."""
        return local_db.connect(LOCAL_DB, _db_key(), timeout=10)

    #Авто-бэкапы локальной базы
    @classmethod
    def backup(cls, reason: str = "") -> str:
        """
        Копирует файл базы в DATA_DIR/backups/ с меткой времени.
        Вызывается перед синхронизацией/закрытием. Возвращает путь к бэкапу
        ('' — если бэкапить нечего). Старые бэкапы (свыше MAX_BACKUPS) удаляются.
        """
        try:
            if not _os.path.exists(LOCAL_DB):
                return ""
            _os.makedirs(BACKUP_DIR, exist_ok=True)
            #Микросекунды (%f) в имени — чтобы два бэкапа в одну и ту же секунду
            #(например, авто-бэкап и «перед восстановлением») не затирали друг друга.
            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            tag = ("_" + "".join(ch for ch in reason if ch.isalnum())[:16]) if reason else ""
            dst = _os.path.join(BACKUP_DIR, f"vsgutu_grades_{ts}{tag}.db")
            #На Windows datetime.now() имеет разрешение ~15 мс — два бэкапа подряд могут
            #получить ОДНУ микросекунду и одно имя, и второй затёр бы первый (потеря копии).
            #Гарантируем уникальность суффиксом, если файл с таким именем уже есть.
            if _os.path.exists(dst):
                base = _os.path.join(BACKUP_DIR, f"vsgutu_grades_{ts}{tag}")
                n = 1
                while _os.path.exists(f"{base}_{n}.db"):
                    n += 1
                dst = f"{base}_{n}.db"
            #WAL: гарантируем, что данные на диске, затем копируем
            try:
                c = local_db.connect(LOCAL_DB, _db_key())
                c.execute("PRAGMA wal_checkpoint(FULL)")
                c.close()
            except Exception as e:
                #Копию всё равно делаем (неполная лучше, чем никакой), но WAL не сброшен —
                #значит самые свежие правки могли остаться в -wal и в копию не попасть.
                #Без этой строки «после восстановления пропали последние оценки» нечем
                #объяснить: сам файл копии выглядит совершенно нормальным.
                log.get("core").warning("[DBManager] WAL не сброшен перед копией (%s) — копия "
                                        "может не содержать последних правок", e)
            _shutil.copy2(LOCAL_DB, dst)
            cls._prune_backups()
            return dst
        except Exception as e:
            log.get("core").warning(f"[DBManager] бэкап не удался: {e}")
            return ""

    @classmethod
    def backup_if_due(cls, min_interval_sec: int = AUTO_BACKUP_INTERVAL_SEC,
                      reason: str = "auto") -> str:
        """Бэкап ПО РАСПИСАНИЮ: делает копию, только если с последнего бэкапа прошло
        не меньше min_interval_sec (или бэкапов ещё нет). Вызывается по таймеру из UI,
        поэтому таймер может тикать чаще — лишних копий не наплодим. Защищает от
        потери данных при аварийном завершении (не дождались бэкапа «на выходе»)."""
        try:
            import time as _time
            backups = cls.list_backups()      #свежие первыми, элемент = (имя,путь,размер,mtime)
            if backups:
                elapsed = _time.time() - backups[0][3]
                #⚠️ elapsed может быть ОТРИЦАТЕЛЬНЫМ — метка файла оказывается «в будущем».
                #Это не экзотика: на Windows NTFS хранит время с точностью 100 нс, а
                #time.time() читает системный таймер с шагом ~15 мс, и только что созданный
                #файл регулярно выглядит новее текущего момента (воспроизводится примерно в
                #14% случаев). Наивное `elapsed < min_interval` считало такой бэкап «слишком
                #свежим» и молча пропускало копию. Тот же эффект даёт шаг часов назад после
                #синхронизации по NTP — тогда бэкапы переставали делаться на часы вперёд.
                #Правило то же, что у лага дельта-push: лучше лишняя копия, чем пропущенная.
                if 0 <= elapsed < min_interval_sec:
                    return ""                 #ещё рано — последний бэкап действительно свежий
            return cls.backup(reason=reason)
        except Exception as e:
            log.get("core").warning(f"[DBManager] backup_if_due: {e}")
            return ""

    @classmethod
    def _prune_backups(cls):
        failed = 0
        try:
            files = sorted(
                (f for f in _os.listdir(BACKUP_DIR) if f.endswith(".db")),
                reverse=True,
            )
            for old in files[MAX_BACKUPS:]:
                try:
                    _os.remove(_os.path.join(BACKUP_DIR, old))
                except Exception:
                    #Копий в папке до полусотни, и файл может быть занят антивирусом или
                    #проводником. Пишем ОДНУ строку с количеством после цикла, а не по
                    #строке на файл — иначе одна занятая папка даёт полсотни записей.
                    failed += 1
        except Exception as e:
            log.get("core").warning("[DBManager] уборка старых копий не выполнена: %s — папка "
                                    "бэкапов будет расти", e)
            return
        if failed:
            log.get("core").warning("[DBManager] не удалось удалить старых копий: %d — папка "
                                    "бэкапов будет расти", failed)

    @classmethod
    def list_backups(cls) -> list:
        """[(имя, полный_путь, размер_байт, mtime), ...] от свежих к старым."""
        out = []
        try:
            for f in _os.listdir(BACKUP_DIR):
                if not f.endswith(".db"):
                    continue
                p = _os.path.join(BACKUP_DIR, f)
                st = _os.stat(p)
                out.append((f, p, st.st_size, st.st_mtime))
        except Exception as e:
            #Пустой список неотличим от «копий ещё нет»: раздел «Резервные копии» покажет
            #пусто при живых файлах на диске, а backup_if_due решит, что копий нет вовсе.
            #Второе безвредно (сделает лишнюю копию), первое — прямой повод к неверному
            #выводу «нас не бэкапят».
            log.get("core").warning("[DBManager] список резервных копий не прочитан: %s", e)
        out.sort(key=lambda x: x[3], reverse=True)
        return out

    @classmethod
    def restore(cls, backup_path: str) -> bool:
        """Восстанавливает базу из бэкапа. Перед этим делает бэкап текущей.

        🔒 ПУСТОЙ ФАЙЛ КОПИИ ОТВЕРГАЕТСЯ (18.08.2026). Раньше проверялось только
        СУЩЕСТВОВАНИЕ пути: копия, оборвавшаяся на закачке или созданная при кончившемся
        месте, копировалась поверх рабочей базы совершенно успешно — и человек оставался
        без данных вовсе, придя сюда именно за спасением данных. Держит
        `tests/test_restore_wal.py::test_an_empty_backup_file_is_refused`.

        ⚠️ Уборка `-wal`/`-shm` ниже — ЗАЩИТА ПО ТРЕБОВАНИЮ SQLite, а НЕ починка
        наблюдавшегося дефекта, и честнее сказать это прямо. SQLite требует убирать журнал
        при подмене файла базы, иначе кадры прежней базы накатываются на восстановленную.
        Я счёл это дырой, написал починку и два теста — и ОТКАТ показал, что оба зелёные и
        на исходном коде: `backup(reason="before_restore")` строкой выше делает
        `wal_checkpoint(FULL)`, то есть незакреплённых кадров к моменту подмены уже нет.
        Защита была, просто не названной. Код оставлен (он верен и стоит копейки), но
        воспроизводимого отказа за ним НЕТ — не считайте его прикрытым тестом.
        """
        try:
            if not _os.path.exists(backup_path):
                return False
            src_size = _os.path.getsize(backup_path)
            if src_size <= 0:
                #Пустой файл копии — это не «восстановление в пустоту», а потеря рабочей
                #базы: `copy2` отработал бы успешно и оставил человека без данных вовсе.
                log.get("core").warning("[DBManager] восстановление отклонено: файл копии "
                                        "%s пуст", backup_path)
                return False
            cls.backup(reason="before_restore")
            #Сбрасываем WAL текущей базы в её же файл и отпускаем соединение. TRUNCATE (а
            #не FULL) — чтобы журнал обнулился, а не остался лежать с кадрами.
            try:
                c = local_db.connect(LOCAL_DB, _db_key(), timeout=10)
                c.execute("PRAGMA wal_checkpoint(TRUNCATE)")
                c.close()
            except Exception as e:
                #Не прерываемся: хвосты всё равно удаляются ниже поимённо. Но записать надо —
                #иначе «почему после восстановления база повела себя странно» нечем объяснить.
                log.get("core").warning("[DBManager] журнал текущей базы не сброшен перед "
                                        "восстановлением (%s)", e)
            _shutil.copy2(backup_path, LOCAL_DB)
            for suffix in ("-wal", "-shm"):
                try:
                    _os.remove(LOCAL_DB + suffix)
                except FileNotFoundError:
                    pass                      #норма: журнала могло и не быть
                except OSError as e:
                    #А вот это уже опасно: файл занят, и SQLite накатит его на чужую базу.
                    #Честно кричим — молчаливое продолжение и есть исходный дефект.
                    log.get("core").error("[DBManager] не удалось убрать %s%s (%s) — база "
                                          "может быть повреждена при следующем открытии",
                                          LOCAL_DB, suffix, e)
                    return False
            if _os.path.getsize(LOCAL_DB) != src_size:
                log.get("core").error("[DBManager] размер восстановленной базы (%d) не совпал "
                                      "с копией (%d) — вероятно, кончилось место",
                                      _os.path.getsize(LOCAL_DB), src_size)
                return False
            return True
        except Exception as e:
            log.get("core").warning(f"[DBManager] восстановление не удалось: {e}")
            return False

    @classmethod
    def wipe_all_local_data(cls, remove_backups: bool = True) -> dict:
        """ПОЛНЫЙ сброс локальных данных ЭТОГО ПК до состояния «первый запуск».

        Удаляет файл базы (вместе со студентами, преподавателями, оценками, группами,
        конфигом, ХЕШЕМ пароля администратора, адресом сервера, сохранённым токеном и
        кэшем темы — всё это лежит в kv_store одной базы), её WAL/SHM-хвосты и, по
        умолчанию, все резервные копии (иначе восстановление вернуло бы старые данные).

        Что НЕ трогаем сознательно:
          • общую базу колледжа на сервере — это ЛОКАЛЬНЫЙ сброс одного ПК;
          • ключ шифрования data.key — новой пустой базе он не мешает;
          • audit.log и login_throttle.json — журнал безопасности по 152-ФЗ сохраняем.

        Таблицы заново НЕ создаём: при следующем старте их пересоздаст DBManager.init().
        Если файл занят другим процессом и удалить его не вышло — откатываемся на
        очистку содержимого через SQL (DROP всех таблиц), чтобы данные всё равно ушли.

        Возвращает {'removed': [...], 'errors': [...]} — что удалили и где споткнулись.
        """
        removed, errors = [], []

        #Гасим фоновую синхронизацию: иначе её поток может в этот момент держать
        #соединение (файл не удалить) или тут же подтянуть данные обратно с сервера.
        try:
            from sync import sync_runner
            sync_runner.stop()
        except Exception as e:
            #Синк не остановлен — а это ровно та причина, по которой сброс может «не
            #сработать»: живой поток держит соединение (файл не удалить) или сразу после
            #очистки тянет данные обратно с сервера. В `errors` не кладём намеренно: этот
            #список показывается человеку как итог операции, и его состав — поведение,
            #которое мы здесь не меняем. Лога достаточно, чтобы объяснить «данные вернулись».
            log.get("core").warning("[wipe] фоновая синхронизация не остановлена: %s", e)

        #Сбрасываем WAL в основной файл и отпускаем соединение — чтобы -wal/-shm не
        #держали данные и файлы освободились для удаления.
        try:
            c = local_db.connect(LOCAL_DB, _db_key())
            c.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            c.close()
        except Exception as e:
            #Продолжаем (файлы -wal/-shm всё равно удаляются ниже поимённо), но знать
            #надо: незакрытый WAL — самая частая причина, по которой файл базы занят и
            #сброс уходит в запасной путь «очистка через SQL».
            log.get("core").warning("[wipe] WAL не сброшен перед удалением базы: %s", e)

        file_ok = True
        for suffix in ("-wal", "-shm", "-journal", ""):
            p = LOCAL_DB + suffix
            try:
                if _os.path.exists(p):
                    _os.remove(p)
                    removed.append(_os.path.basename(p))
            except Exception as e:
                file_ok = False
                errors.append(f"{_os.path.basename(p)}: {e}")

        #Файл не удалился (занят) — чистим содержимое через SQL как запасной путь.
        if not file_ok and _os.path.exists(LOCAL_DB):
            try:
                conn = cls.get_conn(); cur = conn.cursor()
                cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
                for (t,) in cur.fetchall():
                    if t.startswith("sqlite_"):
                        continue
                    cur.execute(f"DROP TABLE IF EXISTS {t}")
                conn.commit(); conn.close()
                cls._init_sqlite_tables()   #оставляем валидную пустую базу
                removed.append("(содержимое базы очищено через SQL)")
            except Exception as e:
                errors.append(f"sql-wipe: {e}")

        if remove_backups and _os.path.isdir(BACKUP_DIR):
            for f in list(_os.listdir(BACKUP_DIR)):
                if not f.endswith(".db"):
                    continue
                try:
                    _os.remove(_os.path.join(BACKUP_DIR, f))
                    removed.append("backups/" + f)
                except Exception as e:
                    errors.append(f"backups/{f}: {e}")

        #Сбрасываем singleton хранилища, чтобы в памяти не осталось ссылок на старое.
        try:
            from data import data_store
            data_store.reset_store()
        except Exception as e:
            #Singleton остался жить со ссылками на стёртые данные: до перезапуска
            #программа может показывать студентов и группы, которых на диске уже нет.
            #Выглядит как «сброс не сработал», хотя база пуста.
            log.get("core").warning("[wipe] singleton хранилища не сброшен: %s — "
                                    "до перезапуска возможны данные из памяти", e)

        return {"removed": removed, "errors": errors}

    @classmethod
    def clear_synced_tables(cls) -> None:
        """Очищает ТОЛЬКО синхронизируемые таблицы (занятия, оценки, конфликты),
        оставляя саму базу и kv_store с локальными настройками (адрес сервера, токен,
        device_id) на месте. Нужна для «реконсиляции с сервером»: на границе сессии
        стираем локальный кэш данных, чтобы затем полным pull привести его в точное
        соответствие серверу — иначе аддитивное слияние оставляло бы «осиротевшие»
        локальные записи (баг с фантомным студентом). Файл НЕ удаляем (в отличие от
        wipe_all_local_data), поэтому WAL-checkpoint не нужен."""
        conn = cls.get_conn(); cur = conn.cursor()
        try:
            for t in ("grades", "lessons", "term_grades", "groups", "users", "sync_conflicts"):
                #{t} — элемент литерального кортежа строкой выше, а не значение
                #извне. Имя таблицы параметром не передать, поэтому f-строка.
                cur.execute(f"DELETE FROM {t}")  # nosec B608
            conn.commit()
        finally:
            conn.close()

    #Конфликты синхронизации (детект вместо тихой перезаписи)
    #
    #⚠️ ЧЕСТНО О СОСТОЯНИИ: ЭКРАНА РАЗБОРА КОНФЛИКТОВ СЕЙЧАС НЕТ. Он жил в Qt-оболочке
    #(`conflict_dialog.py`) и удалён вместе с ней, а нового в SPA не появилось — то есть
    #`list_conflicts`/`resolve_conflict` ниже НЕ ЗОВЁТ никто. Оставлены осознанно (это
    #готовое API для будущего экрана, и detect-часть в sync_engine продолжает работать),
    #но пока конфликт виден только двумя способами: WARNING в `gradebook.log` и число в
    #`sync_runner.status()['conflicts']`. Не считать §4.4 закрытым инвариантом: он
    #выполняется наполовину — локальное значение действительно не затирается, а вот
    #«преподаватель решит вручную» сегодня неправда.
    @classmethod
    def count_unresolved_conflicts(cls) -> int:
        """Сколько расхождений по оценкам ждут решения (0 — всё чисто).

        Отдельный COUNT, а не `len(list_conflicts())`: значение читает индикатор синка на
        каждом обновлении статуса, и тянуть ради счётчика все строки с ПДн незачем.
        Сбой чтения — это 0 плюс запись в лог: уронить индикатор из-за залоченной базы
        нельзя, но и молчать о сбое тоже."""
        try:
            conn = cls.get_conn(); cur = conn.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS sync_conflicts ("
                        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
                        "student_f TEXT, student_n TEXT, lesson_id TEXT,"
                        "local_grade TEXT, remote_grade TEXT, remote_device TEXT,"
                        "remote_at TEXT, detected_at TEXT, resolved INTEGER DEFAULT 0)")
            cur.execute("SELECT COUNT(*) FROM sync_conflicts WHERE resolved=0")
            n = int((cur.fetchone() or [0])[0])
            conn.close()
            return n
        except Exception as e:
            log.get("core").warning(f"[DBManager] счётчик конфликтов: {e}")
            return 0

    @classmethod
    def list_conflicts(cls, unresolved_only: bool = True) -> list:
        """Список конфликтов оценок: [dict(...), ...]."""
        out = []
        try:
            conn = cls.get_conn(); cur = conn.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS sync_conflicts ("
                        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
                        "student_f TEXT, student_n TEXT, lesson_id TEXT,"
                        "local_grade TEXT, remote_grade TEXT, remote_device TEXT,"
                        "remote_at TEXT, detected_at TEXT, resolved INTEGER DEFAULT 0)")
            q = ("SELECT id,student_f,student_n,lesson_id,local_grade,remote_grade,"
                 "remote_device,remote_at,detected_at,resolved FROM sync_conflicts")
            if unresolved_only:
                q += " WHERE resolved=0"
            q += " ORDER BY detected_at DESC"
            cur.execute(q)
            cols = ["id", "student_f", "student_n", "lesson_id", "local_grade",
                    "remote_grade", "remote_device", "remote_at", "detected_at", "resolved"]
            out = [dict(zip(cols, r, strict=False)) for r in cur.fetchall()]
            conn.close()
        except Exception as e:
            log.get("core").warning(f"[DBManager] чтение конфликтов: {e}")
        return out

    @classmethod
    def resolve_conflict(cls, conflict_id: int, chosen_grade: str) -> bool:
        """
        Применяет выбранное преподавателем значение и закрывает конфликт.
        chosen_grade записывается как победитель в SQLite; updated_at обновляется,
        поэтому решение само уедет на сервер при следующей синхронизации (API).

        ⚠️ ВЫЗЫВАЮЩЕГО У ФУНКЦИИ СЕЙЧАС НЕТ (см. комментарий над list_conflicts).
        🔥 И прежняя версия молча ТЕРЯЛА `student_id`: она перезаписывала строку оценки,
        не перенося неизменяемую привязку к студенту (§4.10, этап 2 миграции). Разбор
        конфликта — единственное место, где преподаватель трогает уже уехавшую оценку;
        обнулить ей id значило бы отвязать историю от человека ровно там, где он
        уверен, что всё исправил. Забирать id надо из САМОЙ строки оценки, а не из
        конфликта: конфликт мог быть заведён старым клиентом, который о колонке не знал.
        """
        try:
            conn = cls.get_conn(); cur = conn.cursor()
            cur.execute("SELECT student_f,student_n,lesson_id FROM sync_conflicts "
                        "WHERE id=?", (conflict_id,))
            row = cur.fetchone()
            if not row:
                conn.close(); return False
            f, n, lid = row
            cur.execute("SELECT COALESCE(student_id,'') FROM grades "
                        "WHERE student_f=? AND student_n=? AND lesson_id=?", (f, n, lid))
            sid_row = cur.fetchone()
            sid = (sid_row[0] if sid_row else "") or ""
            now = datetime.now(timezone.utc).isoformat()
            cur.execute("INSERT OR REPLACE INTO grades "
                        "(student_f,student_n,lesson_id,grade,updated_at,device,deleted,"
                        "student_id) VALUES (?,?,?,?,?,?,0,?)",
                        (f, n, lid, chosen_grade, now, DEVICE_ID, sid))
            cur.execute("UPDATE sync_conflicts SET resolved=1 WHERE id=?", (conflict_id,))
            conn.commit(); conn.close()
            return True
        except Exception as e:
            log.get("core").warning(f"[DBManager] resolve_conflict: {e}")
            return False

    @classmethod
    def placeholder(cls) -> str:
        return "?"  #всегда SQLite локально

    @classmethod
    def _init_sqlite_tables(cls):
        conn = local_db.connect(LOCAL_DB, _db_key())
        cur  = conn.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS lessons
            (id TEXT PRIMARY KEY, group_name TEXT, subject TEXT,
             type TEXT, number INTEGER, topic TEXT, date TEXT,
             retake_date TEXT DEFAULT '', hour INTEGER DEFAULT 0)""")
        #deleted — «надгробие»: удалённое занятие НЕ стираем, а помечаем, чтобы
        #удаление доехало до других ПК (иначе занятие воскресало бы при pull).
        #year/semester — учебный период занятия (как на сервере/вебе): год «YYYY/YYYY+1»,
        #семестр 1|2. Нужны для журнала по семестрам и архива прошлых периодов. Старые
        #базы получают колонки через ALTER; сами занятия «усыновляются» в текущий термин
        #при первой загрузке журнала (GradeBook.load_from_db — бэкфилл пустого периода).
        for col, default in [("retake_date", "TEXT DEFAULT ''"), ("hour", "INTEGER DEFAULT 0"),
                             ("updated_at", "TEXT DEFAULT ''"), ("deleted", "INTEGER DEFAULT 0"),
                             ("year", "TEXT DEFAULT ''"), ("semester", "INTEGER DEFAULT 0")]:
            try:
                cur.execute(f"ALTER TABLE lessons ADD COLUMN {col} {default}")
            except Exception as e:
                _alter_ignored("lessons", col, e)
        #Плановые учебные часы предмета на семестр. Задаёт админ (на сайте/в админке ПК),
        #сюда приезжают синком — журнал показывает «пройдено X из Y ч». Ключ
        #детерминированный (hrs:группа|предмет|год|семестр), как у остальных синкуемых
        #сущностей. teacher_id (§ролей препод↔предмет↔группа, 3.3.1) — та же строка несёт
        #и назначение препода на эту пару (группа,предмет): единственный источник правды
        #«какие группы видит преподаватель», см. webdata.teacher_assignments на сервере.
        cur.execute("""CREATE TABLE IF NOT EXISTS subject_hours
            (id TEXT PRIMARY KEY, group_name TEXT, subject TEXT, year TEXT,
             semester INTEGER DEFAULT 0, hours_total INTEGER DEFAULT 0,
             updated_at TEXT DEFAULT '', deleted INTEGER DEFAULT 0, teacher_id TEXT DEFAULT '',
             zet REAL)""")
        #Старая база могла завести таблицу ДО teacher_id/zet — CREATE TABLE IF NOT EXISTS
        #колонку не досоздаёт, нужен ALTER (тот же паттерн, что уже применяем к lessons).
        try:
            cur.execute("ALTER TABLE subject_hours ADD COLUMN teacher_id TEXT DEFAULT ''")
        except Exception as e:
            _alter_ignored("subject_hours", "teacher_id", e)
        #zet — ЗЕТ предмета (docs/PLAN-ZET.md), NULL = администратор не задавал.
        try:
            cur.execute("ALTER TABLE subject_hours ADD COLUMN zet REAL")
        except Exception as e:
            _alter_ignored("subject_hours", "zet", e)
        cur.execute("""CREATE TABLE IF NOT EXISTS students
            (f TEXT, n TEXT, group_name TEXT, PRIMARY KEY(f, n, group_name))""")
        cur.execute("""CREATE TABLE IF NOT EXISTS grades
            (student_f TEXT, student_n TEXT, lesson_id TEXT, grade TEXT,
             updated_at TEXT DEFAULT '', device TEXT DEFAULT '',
             PRIMARY KEY(student_f, student_n, lesson_id))""")
        #миграция старых баз без новых колонок
        for col in ("updated_at", "device"):
            try:
                cur.execute(f"ALTER TABLE grades ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception as e:
                _alter_ignored("grades", col, e)
        #deleted у оценок — то же надгробие (удалённая оценка не должна воскресать).
        try:
            cur.execute("ALTER TABLE grades ADD COLUMN deleted INTEGER DEFAULT 0")
        except Exception as e:
            _alter_ignored("grades", "deleted", e)
        cur.execute("""CREATE TABLE IF NOT EXISTS sync_conflicts
            (id INTEGER PRIMARY KEY AUTOINCREMENT,
             student_f TEXT, student_n TEXT, lesson_id TEXT,
             local_grade TEXT, remote_grade TEXT, remote_device TEXT,
             remote_at TEXT, detected_at TEXT, resolved INTEGER DEFAULT 0)""")
        cur.execute("""CREATE TABLE IF NOT EXISTS kv_store
            (key TEXT PRIMARY KEY, value TEXT NOT NULL)""")
        #Итоговые оценки за семестр (промежуточная аттестация) — отдельно от оценок за
        #занятия. Ключ f|n|subject|year|semester (совпадает с сервером). Синхронизируется
        #как «term_grades» (SYNC_MODELS): ведомости на ПК и на сайте — из одних данных.
        cur.execute("""CREATE TABLE IF NOT EXISTS term_grades
            (id TEXT PRIMARY KEY, student_f TEXT, student_n TEXT, subject TEXT,
             year TEXT DEFAULT '', semester INTEGER DEFAULT 0, grade TEXT DEFAULT '',
             form TEXT DEFAULT '', updated_at TEXT DEFAULT '', deleted INTEGER DEFAULT 0)""")
        #student_id — НЕИЗМЕНЯЕМАЯ привязка оценки к студенту (id строки users).
        #Этап 1 миграции с ФИО-ключей: колонка ДОБАВОЧНАЯ, PRIMARY KEY пока прежний
        #(student_f, student_n, lesson_id). Пустое значение допустимо — старые строки
        #и старые клиенты, которые о колонке не знают. Смысл: ФИО меняется (замужество,
        #опечатка), id — нет, поэтому история перестаёт зависеть от написания фамилии.
        for _tbl in ("grades", "term_grades", "sync_conflicts"):
            try:
                cur.execute(f"ALTER TABLE {_tbl} ADD COLUMN student_id TEXT DEFAULT ''")
            except Exception as e:
                _alter_ignored(_tbl, "student_id", e)
        #Правки расписания админом (overlay поверх портала) — синкуемая сущность.
        #Схему создаём ЗДЕСЬ, при инициализации БД, а не на каждое чтение/запись в
        #schedule/overrides.py: DDL на каждый вызов — лишний парсинг запроса, а флаг
        #«уже создано» в модуле врал бы при ПЕРЕСОЗДАНИИ базы (тесты, сброс кэша).
        cur.execute("""CREATE TABLE IF NOT EXISTS schedule_overrides
            (id TEXT PRIMARY KEY, group_name TEXT, week INTEGER DEFAULT 1, day TEXT,
             pair_no INTEGER DEFAULT 0, action TEXT DEFAULT 'set', subject TEXT,
             time TEXT, room TEXT, teacher TEXT, kind TEXT,
             updated_at TEXT DEFAULT '', deleted INTEGER DEFAULT 0)""")
        #Группы — в СЕРВЕРНОЙ форме (как lessons/grades), а не JSON-блобом в kv_store:
        #синк ходит прямым upsert'ом без переводчика (план техдолга №2, пилот). id=grp:{name}.
        cur.execute("""CREATE TABLE IF NOT EXISTS groups
            (id TEXT PRIMARY KEY, name TEXT, subjects TEXT DEFAULT '[]',
             updated_at TEXT DEFAULT '', deleted INTEGER DEFAULT 0,
             specialty_code TEXT, enrollment_year INTEGER, category TEXT)""")
        #Старая база могла завести таблицу ДО specialty_code/enrollment_year/category
        #(импорт учебного плана ВСГУТУ и категория расписания портала — обе фичи
        #серверные) — тот же паттерн ALTER, что уже применён к subject_hours выше.
        #Десктоп эти поля не показывает, но приезжают синком вместе с остальной
        #группой (Group уже в SYNC_MODELS).
        for _col, _decl in (("specialty_code", "TEXT"), ("enrollment_year", "INTEGER"),
                            ("category", "TEXT")):
            try:
                cur.execute(f"ALTER TABLE groups ADD COLUMN {_col} {_decl}")
            except Exception as e:
                _alter_ignored("groups", _col, e)
        #Пользователи (студенты/преподаватели) — в таблице для прямого синка (LWW/надгробия/
        #дельта, как lessons), НО payload лежит ЗАШИФРОВАННЫМ blob'ом (Fernet+DPAPI): хеши
        #паролей и ПДн НЕ оголяются на диске (152-ФЗ; та же защита, что была у kv_store).
        #id/role/updated_at/deleted — открытые служебные колонки для синка; blob — секрет.
        #Админ здесь НЕ хранится (его хеш живёт в config). План техдолга №2, Стадия 2.
        cur.execute("""CREATE TABLE IF NOT EXISTS users
            (id TEXT PRIMARY KEY, role TEXT, updated_at TEXT DEFAULT '',
             deleted INTEGER DEFAULT 0, blob TEXT DEFAULT '')""")
        conn.commit()
        conn.close()

    @classmethod
    def upsert_lesson(cls, cur, vals: tuple):
        """INSERT OR REPLACE для занятия в SQLite.
        vals = (id, group, subject, type, number, topic, date, retake_date, hour[, year, semester]).
        Проставляем updated_at — нужно для синхронизации через API (LWW). year/semester —
        учебный период (по умолчанию пусто/0; журнал усыновляет пустые в текущий термин)."""
        now = datetime.now(timezone.utc).isoformat()
        base = tuple(vals[:9])
        year = vals[9] if len(vals) > 9 else ""
        semester = int(vals[10]) if len(vals) > 10 else 0
        #deleted=0 — это сохранение АКТИВНОГО занятия (в т.ч. «воскрешает» ранее
        #удалённое, если занятие создали заново с тем же id — на практике id новый).
        cur.execute("INSERT OR REPLACE INTO lessons "
                    "(id,group_name,subject,type,number,topic,date,retake_date,hour,"
                    "year,semester,updated_at,deleted) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,0)",
                    base + (year, semester, now))

    @classmethod
    def upsert_student(cls, cur, vals: tuple):
        cur.execute("INSERT OR IGNORE INTO students (f,n,group_name) VALUES (?,?,?)", vals)

    @classmethod
    def upsert_grade(cls, cur, vals: tuple):
        """
        vals = (student_f, student_n, lesson_id, grade).
        Проставляем updated_at и device — это делает синхронизацию
        детерминированной (newest-wins по времени, а не «как повезёт»).
        """
        f, n, lid, grade = vals[:4]
        now = datetime.now(timezone.utc).isoformat()
        #Пятый элемент (student_id) НЕОБЯЗАТЕЛЕН: старые вызывающие места шлют кортеж из
        #четырёх, и ломать их ради этапа 1 незачем — тогда id доставит бэкофилл.
        sid = vals[4] if len(vals) > 4 else resolve_student_id(f, n)
        #deleted=0 — выставление оценки делает запись активной (снимает прежнее
        #надгробие, если оценку ставят заново после удаления).
        cur.execute(
            "INSERT OR REPLACE INTO grades "
            "(student_f,student_n,lesson_id,grade,updated_at,device,deleted,student_id) "
            "VALUES (?,?,?,?,?,?,0,?)", (f, n, lid, grade, now, DEVICE_ID, sid or ""))

    #Итоговые оценки за семестр (аттестация) ────────────────────────────────────────
    @staticmethod
    def _term_grade_id(f: str, n: str, subject: str, year: str, semester: int,
                       student_id: str = "") -> str:
        """Ключ итоговой оценки — тот же формат, что на сервере.

        ЭТАП 3: ключуем по неизменяемому student_id. Если студент не опознан (нет в
        справочнике, неразличимые тёзки) — откатываемся на СТАРЫЙ ФИО-ключ. Это не
        недоделка, а осознанный компромисс: ростер преподавателя ведётся отдельно, и
        такая оценка законна. Сервер умеет работать со смешанной базой и нормализует
        ключ на приёме, а доклейка проставит id позже (data/student_link.py)."""
        if student_id:
            return term_grade_id(student_id, subject, year, semester)
        return legacy_term_grade_id(f, n, subject, year, semester)

    @classmethod
    def set_term_grade(cls, f: str, n: str, subject: str, year: str, semester: int,
                       grade: str, form: str = "") -> None:
        """Выставить/снять итоговую оценку за семестр (offline-first: сразу в SQLite).
        Пустая оценка = надгробие (deleted=1) — снятие распространится синком. Метку
        updated_at ставим локально (UTC); на сервере при push её перештампует сервер."""
        sid = resolve_student_id(f, n)
        gid = cls._term_grade_id(f, n, subject, year, semester, sid)
        now = datetime.now(timezone.utc).isoformat()
        deleted = 1 if not (grade or "").strip() else 0
        conn = cls.get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT OR REPLACE INTO term_grades "
            "(id,student_f,student_n,subject,year,semester,grade,form,updated_at,deleted,"
            "student_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (gid, f, n, subject, year, int(semester or 0), grade, form, now, deleted, sid))
        conn.commit()
        conn.close()

    @classmethod
    def list_terms(cls) -> list:
        """Учебные периоды, по которым есть занятия: [(year, semester)], новые сверху.
        Для селектора семестра в журнале и архива прошлых периодов (как list_terms в вебе)."""
        conn = cls.get_conn()
        cur = conn.cursor()
        try:
            cur.execute("SELECT DISTINCT year, COALESCE(semester,0) FROM lessons "
                        "WHERE COALESCE(deleted,0)=0 AND COALESCE(year,'')<>''")
            terms = sorted({(y, int(s or 0)) for y, s in cur.fetchall() if y},
                           key=lambda t: (t[0], t[1]), reverse=True)
        except Exception as e:
            #Пустой список = селектор семестра пуст, и человек читает это как «занятий не
            #было ни в одном периоде». Пустоту оставляем (журнал важнее селектора), но
            #отличить сбой чтения от честного отсутствия периодов теперь можно по логу.
            log.get("core").warning("[DBManager] список учебных периодов не прочитан: %s", e)
            terms = []
        conn.close()
        return terms

    @classmethod
    def group_subjects_with_lessons(cls, group: str) -> list:
        """Предметы, по которым в группе РЕАЛЬНО есть занятия (активные). Нужен, чтобы
        студент видел предмет с выставленными оценками, даже если этого предмета нет в
        портальном расписании / списке предметов группы (иначе оценки «пропадают»)."""
        conn = cls.get_conn(); cur = conn.cursor()
        try:
            cur.execute("SELECT DISTINCT subject FROM lessons WHERE group_name=? "
                        "AND COALESCE(deleted,0)=0 AND COALESCE(subject,'')<>''", (group,))
            subs = [r[0] for r in cur.fetchall()]
        except Exception as e:
            #Функция существует ровно затем, чтобы студент увидел предмет с оценками,
            #которого нет в портальном расписании. Пустой список молча возвращает тот
            #самый баг, который она закрывает: «оценки пропали». Поведение сохраняем.
            log.get("core").warning("[DBManager] предметы с занятиями группы «%s» не "
                                    "прочитаны: %s", group, e)
            subs = []
        conn.close()
        return subs

    @classmethod
    def get_term_grades(cls, subject: str, year: str, semester: int) -> dict:
        """Итоговые оценки по предмету за термин: {«f|n»: {'grade','form'}} (без надгробий)."""
        conn = cls.get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT student_f,student_n,grade,form FROM term_grades "
            "WHERE subject=? AND year=? AND semester=? AND COALESCE(deleted,0)=0",
            (subject, year, int(semester or 0)))
        out = {f"{r[0]}|{r[1]}": {"grade": r[2] or "", "form": r[3] or ""}
               for r in cur.fetchall()}
        conn.close()
        return out


#Датаклассы
@dataclass
class Lesson:
    id: str
    type: str
    number: int
    topic: str
    date: str
    retake_date: str = ""
    hour: int = 0
    year: str = ""          #учебный год «YYYY/YYYY+1» (период занятия)
    semester: int = 0       #семестр 1 (осень) | 2 (весна)


@dataclass
class Student:
    n: str
    f: str
    group: str
    records: dict[str, str] = field(default_factory=dict)


#Журнал
class GradeBook:
    def __init__(self, group: str, subject: str, year: str = "", semester: int = 0):
        self.group   = group
        self.subject = subject
        #Учебный период журнала. Пусто → БЕЗ фильтра (все периоды — прежнее поведение,
        #обратная совместимость для вызовов без термина). Задан → журнал показывает
        #только этот семестр; прошлые периоды в UI открываются read-only (архив).
        self.year    = year or ""
        self.semester = int(semester or 0)
        DBManager._init_sqlite_tables()
        self.lessons: list[Lesson] = []
        self.spisok_stud: list[Student] = []
        self.load_from_db()

    def add_student(self, st: Student):
        self.spisok_stud.append(st)
        self.save_to_db()

    def delete_student(self, surname: str, name: str):
        surname = surname.strip()
        name    = name.strip()
        self.spisok_stud = [
            s for s in self.spisok_stud
            if not (s.f.strip().lower() == surname.lower()
                    and s.n.strip().lower() == name.lower())
        ]
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        #Оценки удаляем НЕ физически, а надгробием (deleted=1 + свежий updated_at):
        #так удаление доедет до других ПК через синхронизацию, а не «воскреснет» при
        #следующем pull. Физическое удаление оставило бы серверную копию живой.
        now = datetime.now(timezone.utc).isoformat()
        cur.execute("UPDATE grades SET deleted=1, updated_at=?, device=? "
                    "WHERE student_f=? AND student_n=?", (now, DEVICE_ID, surname, name))
        cur.execute("DELETE FROM students WHERE f=? AND n=?", (surname, name))
        conn.commit()
        conn.close()

    def delete_lesson(self, lesson_id: str):
        """Удаляет занятие НАДГРОБИЕМ (deleted=1), а не физически — иначе удаление не
        доезжало бы до других ПК и занятие воскресало бы при следующем pull. Раньше
        занятие лишь убиралось из списка в памяти, а строка в SQLite оставалась —
        и при перезагрузке журнала столбец возвращался. Заодно надгробим оценки
        этого занятия, чтобы они не висели «осиротевшими» и тоже удалились везде."""
        self.lessons = [l for l in self.lessons if l.id != lesson_id]
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        now = datetime.now(timezone.utc).isoformat()
        cur.execute("UPDATE lessons SET deleted=1, updated_at=? WHERE id=?", (now, lesson_id))
        cur.execute("UPDATE grades SET deleted=1, updated_at=?, device=? WHERE lesson_id=?",
                    (now, DEVICE_ID, lesson_id))
        conn.commit()
        conn.close()

    def add_lesson(self, lesson_type: str, topic: str = "", date: str = "", hour: int = 0) -> "Lesson":
        nums     = [l.number for l in self.lessons if l.type == lesson_type and getattr(l, 'hour', 0) in (0, 1)]
        next_num = max(nums) + 1 if nums else 1
        dt = date or datetime.now().strftime('%d.%m.%Y')

        #Новое занятие штампуем текущим периодом журнала (self.year/semester) — чтобы
        #оно попало в правильный семестр. Если период не задан (легаси-вызов) — пусто,
        #сервер проставит термин при push, а бэкфилл усыновит его в текущий период.
        yr, sem = self.year, self.semester
        if lesson_type == "Лекция" and hour == 0:
            l1 = Lesson(id=str(uuid.uuid4()), type="Лекция", number=next_num, topic=topic, date=dt, retake_date="", hour=1, year=yr, semester=sem)
            l2 = Lesson(id=str(uuid.uuid4()), type="Лекция", number=next_num, topic=topic, date=dt, retake_date="", hour=2, year=yr, semester=sem)
            self.lessons.extend([l1, l2])
            self.save_to_db()
            return l1
        else:
            l = Lesson(id=str(uuid.uuid4()), type=lesson_type, number=next_num, topic=topic, date=dt, retake_date="", hour=hour, year=yr, semester=sem)
            self.lessons.append(l)
            self.save_to_db()
            return l

    def set_retake_date(self, lesson_id: str, retake_date: str, retake_n: int = 1):
        attr = 'retake_date' if retake_n == 1 else f'retake_date_{retake_n}'
        for l in self.lessons:
            if l.id == lesson_id:
                setattr(l, attr, retake_date)
                break
        col = 'retake_date' if retake_n == 1 else f'retake_date_{retake_n}'
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        try:
            cur.execute(f"ALTER TABLE lessons ADD COLUMN {col} TEXT DEFAULT ''")
        except Exception as e:
            _alter_ignored("lessons", col, e)
        #SAST B608: {col} — имя колонки из нашего же списка миграции; данные
        #идут параметрами.
        cur.execute(f"UPDATE lessons SET {col}=? WHERE id=?",  # nosec B608
                    (retake_date, lesson_id))
        conn.commit()
        conn.close()

    def save_to_db(self):
        conn = DBManager.get_conn()
        cur  = conn.cursor()
        for l in self.lessons:
            DBManager.upsert_lesson(cur, (
                l.id, self.group, self.subject, l.type,
                l.number, l.topic, l.date,
                getattr(l, 'retake_date', ''),
                getattr(l, 'hour', 0),
                #период занятия сохраняем как есть (архивные не переносим в текущий);
                #если у занятия периода нет — берём период журнала.
                getattr(l, 'year', '') or self.year,
                getattr(l, 'semester', 0) or self.semester,
            ))
        for s in self.spisok_stud:
            DBManager.upsert_student(cur, (s.f, s.n, s.group))
            for lesson_id, grade in s.records.items():
                DBManager.upsert_grade(cur, (s.f, s.n, lesson_id, grade))
        conn.commit()
        conn.close()

    def load_from_db(self):
        conn = DBManager.get_conn()
        cur  = conn.cursor()

        #Бэкфилл периода: занятия без учебного периода (старые базы, до семестров)
        #«усыновляем» в ТЕКУЩИЙ термин — иначе при фильтрации по семестру они бы не
        #показались ни в одном. Только когда журнал открыт с термином (self.year задан).
        if self.year:
            try:
                from data import terms
                cy, cs = terms.current_term()
                cur.execute(
                    "UPDATE lessons SET year=?, semester=? "
                    "WHERE group_name=? AND subject=? AND COALESCE(year,'')='' "
                    "AND COALESCE(deleted,0)=0",
                    (cy, int(cs or 0), self.group, self.subject))
                conn.commit()
            except Exception as e:
                log.get("core").warning(f"[GradeBook] бэкфилл периода пропущен: {e}")

        #deleted=0 — удалённые (надгробия) занятия в журнал не показываем. Если задан
        #период журнала (self.year) — показываем только его семестр (архив/текущий).
        base_sql = ("SELECT id, type, number, topic, date, retake_date, hour, "
                    "COALESCE(year,''), COALESCE(semester,0) "
                    "FROM lessons WHERE group_name=? AND subject=? AND COALESCE(deleted,0)=0")
        params = [self.group, self.subject]
        if self.year:
            base_sql += " AND COALESCE(year,'')=? AND COALESCE(semester,0)=?"
            params += [self.year, self.semester]
        base_sql += " ORDER BY type, number, hour"
        cur.execute(base_sql, tuple(params))
        self.lessons = []
        for row in cur.fetchall():
            l = Lesson(
                id=row[0], type=row[1], number=row[2],
                topic=row[3], date=row[4],
                retake_date=row[5] if row[5] else "",
                hour=row[6] if row[6] else 0,
                year=row[7] or "", semester=row[8] or 0,
            )
            self.lessons.append(l)

        #Синхронизация студентов из общего хранилища (data_store)
        #Студентами управляет администратор через data_store (kv_store в SQLite,
        #а общая база — через API-сервер). Здесь мы лишь дозаполняем локальную
        #таблицу students теми, кого ещё нет. ВАЖНО: НЕ удаляем студентов из SQLite —
        #иначе добавленные администратором студенты исчезали бы при каждой
        #перезагрузке журнала.
        try:
            from data.data_store import get_store
            store = get_store()
            if store:
                known_students = store.get_students()
                existing = set()
                cur.execute("SELECT f, n FROM students WHERE group_name=?", (self.group,))
                for row in cur.fetchall():
                    existing.add((row[0].strip().lower(), row[1].strip().lower()))
                for s in known_students:
                    if s.get("group", "") != self.group:
                        continue
                    surname = s.get("surname", "").strip()
                    name    = s.get("name", "").strip()
                    if surname and name and (surname.lower(), name.lower()) not in existing:
                        DBManager.upsert_student(cur, (surname, name, self.group))
                        existing.add((surname.lower(), name.lower()))
                conn.commit()
        except Exception as e:
            log.get("core").warning(f"[GradeBook] синхронизация студентов в load_from_db: {e}")

        cur.execute("SELECT f, n, group_name FROM students WHERE group_name=?", (self.group,))
        self.spisok_stud = []
        for f, n, g in cur.fetchall():
            student = Student(n, f, g)
            #deleted=0 — удалённые (надгробия) оценки в журнал/расчёт не берём.
            cur.execute("SELECT lesson_id, grade FROM grades "
                        "WHERE student_f=? AND student_n=? AND COALESCE(deleted,0)=0", (f, n))
            student.records = {row[0]: row[1] for row in cur.fetchall()}
            self.spisok_stud.append(student)
        conn.close()

    def _hours_plan_keys(self) -> list:
        """Ключи subject_hours, по которым ищем план — в порядке убывания точности.

        Журнал открыт с ЯВНЫМ периодом (архив прошлого семестра) — берём только его:
        подставить туда план текущего семестра значило бы показать чужую цифру.

        Периода нет (так журнал открывают экраны студента и часть экранов препода) —
        сначала пробуем ТЕКУЩИЙ термин, и только потом бестерминный ключ. Это и есть
        причина, по которой часы не появлялись на ПК: админ сохраняет их с сайта, а там
        период подставляется всегда (`hrs:Группа|Предмет|2025/2026|1`), тогда как журнал
        без периода искал `hrs:Группа|Предмет||0` — ключи не совпадали никогда.
        Бестерминный вариант оставляем запасным: с ним лежат строки, заведённые до
        появления учебных периодов."""
        def _key(year, semester):
            return f"hrs:{self.group}|{self.subject}|{year}|{int(semester or 0)}"
        if self.year and self.semester:
            return [_key(self.year, self.semester)]
        keys = []
        try:
            from data import terms
            y, s = terms.current_term()
            keys.append(_key(y, s))
        except Exception as e:
            #Обойдёмся бестерминным ключом ниже — но это ровно тот сценарий, из-за
            #которого часы «не появлялись на ПК»: админ сохраняет их С периодом, а мы
            #ищем `hrs:Группа|Предмет||0`, и ключи не совпадают никогда. Раз в запуск:
            #зовётся на каждое открытие журнала.
            _report_once("hours_term", "[GradeBook] текущий учебный период не определён (%s) — "
                         "план часов ищу по бестерминному ключу, он может не найтись", e)
        keys.append(_key("", 0))
        return keys

    def hours_progress(self) -> tuple:
        """(пройдено, план) академических часов по этому журналу.

        Пройденное считает общий с сервером study_hours (лекция = пара из двух строк,
        практика = одна строка, обе дают по 2 часа). План приходит от админа синком; его
        нет — возвращаем 0, и интерфейс просто не показывает строку. Придумывать план
        нельзя: «пройдено 24 из 0» выглядит как поломка данных."""
        import study_hours
        from contextlib import closing
        done = study_hours.hours_done(self.lessons)
        total = 0
        try:
            with closing(DBManager.get_conn()) as conn:
                cur = conn.cursor()
                for hid in self._hours_plan_keys():
                    cur.execute("SELECT COALESCE(hours_total,0) FROM subject_hours "
                                "WHERE id=? AND COALESCE(deleted,0)=0", (hid,))
                    row = cur.fetchone()
                    if row:
                        total = int(row[0])
                        break
        except Exception:
            #Нет таблицы (старая база до первого синка) — молча считаем «план не задан».
            total = 0
        return done, total

    def calculate_average(self, student: Student, cfg=None, scale=None) -> float:
        """Средний балл через единый модуль grading (та же формула, что у Вектора).
        Методика (Н=вес, учитывать ли пропуск, включать ли экзамены) берётся из config,
        дефолты сохраняют прежнее поведение. cfg можно передать явно — например, чтобы
        ВРЕМЕННО (визуально) отключить учёт пропусков «Н=2» тумблером в интерфейсе, не
        меняя общий config. scale — шкала преподавателя (§ролей, 3.3.1), ведущего эти
        занятия; без него — "5" (сегодняшнее поведение). Разрешение «чья шкала» — забота
        вызывающего: шкалу знает тот, кто знает преподавателя (на сервере — `webdata.
        lesson_scale_map` по назначению препод↔предмет↔группа), а этот модуль ролей не
        видит и решать за них не должен. Прежняя ссылка на `teacher_dashboard` устарела —
        нативных экранов нет с удаления Qt."""
        import grading
        if cfg is None:
            try:
                from data.data_store import get_store
                cfg = get_store()._config()
            except Exception:
                cfg = {}
        if scale is None:
            scale = grading.DEFAULT_SCALE
        return grading.practice_average(
            grading.pairs_from_objects(self.lessons), student.records, cfg, scale=scale)

    def export_to_excel(self, file_path: str, scale=None):
        """Экспорт журнала в аккуратный xlsx: титульная шапка (группа/предмет/дата),
        фирменные цвета, рамки, цвет оценок по уровню, закреплённые области и строка
        «средний по группе». Весь текст — Times New Roman 14 (требование заказчика).
        scale — шкала преподавателя (§ролей, 3.3.1); журнал — один предмет/один препод,
        поэтому одна строка на весь экспорт (как в calculate_average)."""
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        #ЕДИНЫЙ стиль десктопа и веба: Times New Roman 14, БЕЗ цветов (ч/б), рамки,
        #адаптивная ширина. Совпадает с server/app/xlsx_export.py.
        FNT = "Times New Roman"
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        safe_title = re.sub(r'[\[\]:*?/\\]', '_', f"Успеваемость {self.group}")
        ws.title = safe_title[:31]

        headers = ["Фамилия", "Имя"]
        for l in self.lessons:
            if l.type == "Экзамен":
                headers.append(f"Экзамен №{l.number}\n({l.date})\n{l.topic}")
                if l.retake_date:
                    headers.append(f"Пересдача\n({l.retake_date})")
            else:
                headers.append(f"{l.type} №{l.number}\n({l.date})")
        headers.append("Средний балл")
        ncols = len(headers)
        last_col = get_column_letter(ncols)

        #Титульная шапка: что это, чья группа, когда выгружено.
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = "Журнал успеваемости"
        ws["A1"].font = Font(name=FNT, size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = f"Группа {self.group}  ·  {self.subject}"
        ws["A2"].font = Font(name=FNT, size=14, bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A3:{last_col}3")
        ws["A3"] = (f"Выгружено {datetime.now().strftime('%d.%m.%Y %H:%M')}  ·  "
                    f"GradeBookAI · Технологический колледж ВСГУТУ")
        ws["A3"].font = Font(name=FNT, size=12, italic=True)
        ws["A3"].alignment = Alignment(horizontal="center")

        HDR = 5                                             #строка заголовков таблицы
        ws.append([])                                       #строка 4 — воздух
        ws.append(headers)
        for cell in ws[HDR]:
            cell.font = Font(name=FNT, size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        averages = []
        for i, s in enumerate(self.spisok_stud):
            row = [s.f, s.n]
            for l in self.lessons:
                base = s.records.get(l.id, "")
                row.append(base)
                if l.type == "Экзамен" and l.retake_date:
                    rv = s.records.get(l.id + "_retake", "")
                    if not rv:
                        #пересдача только у заваливших основной экзамен
                        b = (base or "").strip()
                        failed = b.startswith(("2", "Н")) or "Не зачтено" in b
                        rv = "" if failed else "—"
                    row.append(rv)
            avg = round(self.calculate_average(s, scale=scale), 2)
            averages.append(avg)
            row.append(avg if avg > 0 else "")
            ws.append(row)
            r = HDR + 1 + i
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.font = Font(name=FNT, size=14, bold=(c == ncols))
                cell.alignment = Alignment(
                    horizontal="left" if c <= 2 else "center", vertical="center")

        #Итог: средний по группе (только по студентам с оценками).
        vals = [a for a in averages if a > 0]
        total_row = HDR + len(self.spisok_stud) + 1
        ws.cell(row=total_row, column=1, value="Средний по группе:")
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row,
                       end_column=ncols - 1)
        ws.cell(row=total_row, column=1).font = Font(name=FNT, size=14, bold=True)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
        tc = ws.cell(row=total_row, column=ncols,
                     value=round(sum(vals) / len(vals), 2) if vals else "—")
        tc.font = Font(name=FNT, size=14, bold=True)
        tc.alignment = Alignment(horizontal="center")

        #Адаптивная ширина по содержимому (шапка + строки студентов + «средний»), чтобы
        #текст не обрезался. Титульные merge-строки не мерим. TNR 14 шире — коэф. 1.45.
        for c in range(1, ncols + 1):
            best = 0
            for rr in [HDR] + list(range(HDR + 1, total_row + 1)):
                v = ws.cell(row=rr, column=c).value
                if v is None:
                    continue
                best = max(best, max((len(s) for s in str(v).split("\n")), default=0))
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(44, best * 1.45 + 2))
        ws.row_dimensions[HDR].height = 62
        ws.freeze_panes = f"C{HDR + 1}"
        wb.save(file_path)
