"""
test_curator.py — Роль «Куратор» (teacher с непустым curated_groups).

Ключевой инвариант: куратор видит ВСЕ предметы своей курируемой группы (в т.ч. которые
сам НЕ ведёт), но ТОЛЬКО НА ЧТЕНИЕ, и строго по row-level (group ∈ curated_groups).
"""
from conftest import make_admin, make_teacher
from app.security import hash_password


def _push(client, h, **ent):
    return client.post("/sync/push", json={"changes": ent}, headers=h)


def _login(client, login, pw):
    r = client.post("/auth/login", json={"login": login, "password": pw})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def test_curator_sees_all_group_subjects_readonly(client):
    admin = make_admin(client)
    L_mat = {"id": "Lm", "group_name": "G1", "subject": "Мат", "type": "Практика", "number": 1}
    L_fiz = {"id": "Lf", "group_name": "G1", "subject": "Физ", "type": "Практика", "number": 1}
    stud = {"id": "stud:s1", "role": "student", "login": "s1", "surname": "Пуп",
            "name": "Пётр", "group_name": "G1", "password_hash": hash_password("p")}
    g_fiz = {"id": "Пуп|Пётр|Lf", "student_f": "Пуп", "student_n": "Пётр",
             "lesson_id": "Lf", "grade": "4"}
    assert _push(client, admin, lessons=[L_mat, L_fiz], users=[stud], grades=[g_fiz]).status_code == 200

    #Куратор ведёт только «Мат», но курирует всю группу G1
    r = client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор", "login": "t1", "password": "pass1234",
        "subjects": ["Мат"], "curated_groups": ["G1"]}, headers=admin)
    assert r.status_code == 200, r.text
    th = _login(client, "t1", "pass1234")

    assert client.get("/web/curator/groups", headers=th).json()["groups"] == ["G1"]
    #видит и «Физ» — чужой для него предмет
    subs = client.get("/web/curator/subjects", params={"group":"G1"}, headers=th).json()["subjects"]
    assert "Мат" in subs and "Физ" in subs, subs
    #видит оценки по «Физ» (не свой предмет!) — чтение
    view = client.get("/web/curator/group-subject", params={"group":"G1","subject":"Физ"}, headers=th).json()
    assert any(st["grades"].get("Lf") == "4" for st in view["students"]), view
    #некурируемая группа → 403
    assert client.get("/web/curator/subjects", params={"group":"G2"}, headers=th).status_code == 403


def test_curator_sees_freshly_imported_subjects_before_any_lesson_exists(client):
    """§3.5.5: импорт учебного плана/расписания (import-esstu/import-schedule-category)
    пишет ТОЛЬКО Group.subjects, ни одной строки Lesson не создаёт. Куратор раньше видел
    список предметов ИСКЛЮЧИТЕЛЬНО через Lesson (`group_lessons`) — спарсили в группу с
    0 предметов дюжину новых, а у куратора всё ещё «0», пока хоть один препод не откроет
    журнал по новому предмету. Живой баг, поймано на реальном импорте."""
    admin = make_admin(client)
    r = client.post("/web/admin/groups", json={"name": "G3", "subjects": []}, headers=admin)
    assert r.status_code == 200, r.text
    r = client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор", "login": "t2", "password": "pass1234",
        "subjects": [], "curated_groups": ["G3"]}, headers=admin)
    assert r.status_code == 200, r.text
    th = _login(client, "t2", "pass1234")

    #Ровно сценарий из отзыва: изначально 0 предметов.
    subs = client.get("/web/curator/subjects", params={"group": "G3"}, headers=th).json()["subjects"]
    assert subs == []

    #Админ "спарсил" предметы в Group.subjects — ни одной Lesson-строки при этом нет.
    r = client.put("/web/admin/groups/G3", json={"subjects": ["Физика", "Химия", "Мат"]}, headers=admin)
    assert r.status_code == 200, r.text

    subs = client.get("/web/curator/subjects", params={"group": "G3"}, headers=th).json()["subjects"]
    assert set(subs) == {"Физика", "Химия", "Мат"}


def test_curator_subjects_from_archived_term_unaffected_by_current_group_subjects(client):
    """Регрессия: архивный (явно выбранный year/semester, не текущий термин) список
    ОБЯЗАН остаться чисто Lesson-based — Group.subjects отражает СЕГОДНЯШНИЙ учебный
    план, который импорт мог заменить, и подмешивать его в прошлый семестр значило бы
    приписать группе предметы, которых тогда могло не быть."""
    admin = make_admin(client)
    L = {"id": "La", "group_name": "G4", "subject": "История", "type": "Практика",
        "number": 1, "year": "2020/2021", "semester": 1}
    assert client.post("/sync/push", json={"changes": {"lessons": [L]}}, headers=admin).status_code == 200
    r = client.post("/web/admin/groups", json={"name": "G4", "subjects": ["Новый предмет"]}, headers=admin)
    assert r.status_code == 200, r.text
    r = client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор2", "login": "t3", "password": "pass1234",
        "subjects": [], "curated_groups": ["G4"]}, headers=admin)
    assert r.status_code == 200, r.text
    th = _login(client, "t3", "pass1234")

    r = client.get("/web/curator/subjects",
                   params={"group": "G4", "year": "2020/2021", "semester": 1}, headers=th)
    assert r.status_code == 200, r.text
    assert r.json()["subjects"] == ["История"]     #НЕ "Новый предмет" — тот из ТЕКУЩЕГО плана


def test_non_curator_teacher_has_no_curator_access(client):
    admin = make_admin(client)
    th = make_teacher(client, admin, subjects=["Мат"])
    assert client.get("/web/curator/groups", headers=th).json()["groups"] == []
    #даже зная имя группы — 403 (не курирует)
    assert client.get("/web/curator/subjects", params={"group":"G1"}, headers=th).status_code == 403


#  Сводный отчёт успеваемости (Excel / Word) + скоуп по группам

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def _setup_curator_two_groups(client):
    admin = make_admin(client)
    _push(client, admin,
          lessons=[{"id": "Lm", "group_name": "G1", "subject": "Мат", "type": "Практика",
                    "number": 1},
                   {"id": "Lb", "group_name": "G2", "subject": "Био", "type": "Практика",
                    "number": 1}],
          users=[{"id": "stud:s1", "role": "student", "login": "s1", "surname": "Пуп",
                  "name": "Пётр", "group_name": "G1", "password_hash": hash_password("p")},
                 {"id": "stud:s2", "role": "student", "login": "s2", "surname": "Кот",
                  "name": "Анна", "group_name": "G2", "password_hash": hash_password("p")}],
          grades=[{"id": "Пуп|Пётр|Lm", "student_f": "Пуп", "student_n": "Пётр",
                   "lesson_id": "Lm", "grade": "5"}])
    #Куратор ДВУХ групп
    client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор", "login": "t1", "password": "pass1234",
        "subjects": ["Мат"], "curated_groups": ["G1", "G2"]}, headers=admin)
    return _login(client, "t1", "pass1234")


def test_curator_report_xlsx_and_docx(client):
    th = _setup_curator_two_groups(client)
    rx = client.get("/web/curator/report", params={"groups": "all", "fmt": "xlsx"}, headers=th)
    assert rx.status_code == 200, rx.text
    assert rx.headers["content-type"].startswith(_XLSX)
    assert len(rx.content) > 500          #непустой файл

    rd = client.get("/web/curator/report", params={"groups": "all", "fmt": "docx"}, headers=th)
    assert rd.status_code == 200, rd.text
    assert rd.headers["content-type"].startswith(_DOCX)


def test_curator_report_single_group(client):
    th = _setup_curator_two_groups(client)
    r = client.get("/web/curator/report", params={"groups": "G1", "fmt": "xlsx"}, headers=th)
    assert r.status_code == 200, r.text
    #лист ровно один — G1
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["G1"], wb.sheetnames


def test_curator_report_rejects_foreign_group(client):
    """Скоуп: чужую группу в параметре подставить нельзя — она отфильтруется, а если
    в запросе ТОЛЬКО чужая, то 403 (нечего отдавать)."""
    th = _setup_curator_two_groups(client)
    #G9 не курируется → в пересечении пусто → 403
    assert client.get("/web/curator/report", params={"groups": "G9", "fmt": "xlsx"},
                      headers=th).status_code == 403
    #G1 (своя) + G9 (чужая) → выгружается ТОЛЬКО G1
    import io
    from openpyxl import load_workbook
    r = client.get("/web/curator/report", params={"groups": "G1,G9", "fmt": "xlsx"}, headers=th)
    assert r.status_code == 200
    assert load_workbook(io.BytesIO(r.content)).sheetnames == ["G1"]


def test_non_curator_cannot_get_report(client):
    admin = make_admin(client)
    th = make_teacher(client, admin, subjects=["Мат"])
    assert client.get("/web/curator/report", params={"groups": "all", "fmt": "xlsx"},
                      headers=th).status_code == 403


def test_curator_group_with_slash_in_name(client):
    """Группы колледжа содержат слэш («К75/1»). Раньше group шёл в PATH, и encodeURIComponent
    давал «%2F», который ломал роутинг → эндпоинт молча 404-ил и вкладка показывала «нет
    предметов». Теперь group — query-параметр, слэш проходит как значение."""
    admin = make_admin(client)
    _push(client, admin,
          lessons=[{"id": "L1", "group_name": "К75/1", "subject": "Биология",
                    "type": "Практика", "number": 1, "year": "2026/2027", "semester": 1}],
          users=[{"id": "stud:s1", "role": "student", "login": "s1", "surname": "Пуп",
                  "name": "Пётр", "group_name": "К75/1", "password_hash": hash_password("p")}])
    client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор", "login": "t1", "password": "pass1234",
        "subjects": ["X"], "curated_groups": ["К75/1"]}, headers=admin)
    th = _login(client, "t1", "pass1234")

    r = client.get("/web/curator/subjects",
                   params={"group": "К75/1", "year": "2026/2027", "semester": 1}, headers=th)
    assert r.status_code == 200, r.text
    assert r.json()["subjects"] == ["Биология"], r.json()

    v = client.get("/web/curator/group-subject",
                   params={"group": "К75/1", "subject": "Биология",
                           "year": "2026/2027", "semester": 1}, headers=th)
    assert v.status_code == 200, v.text
    assert v.json()["group"] == "К75/1"


def test_curator_group_subject_shows_hours_plan(client):
    """Живой баг 3.6.1: у преподавателя в ЕГО журнале часы «пройдено X из Y» есть, а у
    куратора на ТОМ ЖЕ предмете/группе — не было вовсе (эндпоинт их просто не отдавал)."""
    admin = make_admin(client)
    r = client.post("/web/admin/groups", json={"name": "G5", "subjects": ["Физ"]}, headers=admin)
    assert r.status_code == 200, r.text
    r = client.post("/web/admin/group-hours",
                    json={"group": "G5", "hours": {"Физ": 72}}, headers=admin)
    assert r.status_code == 200, r.text
    r = client.post("/web/admin/teachers", json={
        "full_name": "Кур Атор", "login": "t4", "password": "pass1234",
        "subjects": [], "curated_groups": ["G5"]}, headers=admin)
    assert r.status_code == 200, r.text
    th = _login(client, "t4", "pass1234")

    v = client.get("/web/curator/group-subject",
                   params={"group": "G5", "subject": "Физ"}, headers=th)
    assert v.status_code == 200, v.text
    assert v.json()["hours"] == {"done": 0, "total": 72}, v.json()
