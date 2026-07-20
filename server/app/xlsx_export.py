"""
xlsx_export.py — сборка журнала и ведомости в xlsx (веб).

ЕДИНЫЙ стиль десктопа и веба (по требованию заказчика): весь текст Times New Roman 14,
БЕЗ цветов (чёрный текст на белом), тонкие рамки, заголовки жирным, адаптивная ширина
столбцов (текст не обрезается). Официальный чёрно-белый вид, пригодный для печати.
Данные приходят уже выбранными по роли (routers/web.py): (group, subject, lessons, rows).
"""
import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FNT = "Times New Roman"
SZ = 14                    #единый размер шрифта везде
_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _autofit(ws, ncols: int, rows: list, min_w=8, max_w=48):
    """Адаптивная ширина столбцов по содержимому (openpyxl сам не умеет auto-fit).
    Меряем ТОЛЬКО переданные строки (шапка+данные), без объединённых титульных —
    иначе длинный заголовок раздул бы все колонки. Для многострочных ячеек берём
    самую длинную строку. Times New Roman 14 шире — коэффициент 1.45."""
    for c in range(1, ncols + 1):
        best = 0
        for r in rows:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            best = max(best, max((len(s) for s in str(v).split("\n")), default=0))
        ws.column_dimensions[get_column_letter(c)].width = max(min_w, min(max_w, best * 1.45 + 2))


def _title(ws, row, last_col, text, size, bold=False, italic=False):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(name=FNT, size=size, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal="center", vertical="center")


def build_journal_xlsx(group: str, subject: str, lessons, rows) -> bytes:
    """lessons — ORM-строки Lesson; rows — [{surname, name, records: {lid→оценка}, average}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\[\]:*?/\\]', '_', f"Успеваемость {group}")[:31]

    headers = ["Фамилия", "Имя"]
    keys = []
    for l in lessons:
        if l.type == "Экзамен":
            headers.append(f"Экзамен №{l.number}\n({l.date})\n{l.topic or ''}".strip())
            keys.append(l.id)
            if l.retake_date:
                headers.append(f"Пересдача\n({l.retake_date})")
                keys.append(l.id + "_retake")
        else:
            headers.append(f"{l.type} №{l.number}\n({l.date})")
            keys.append(l.id)
    headers.append("Средний балл")
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    _title(ws, 1, last_col, "Журнал успеваемости", 16, bold=True)
    _title(ws, 2, last_col, f"Группа {group}  ·  {subject}", SZ, bold=True)
    _title(ws, 3, last_col, f"Выгружено {datetime.now().strftime('%d.%m.%Y %H:%M')}  ·  "
                            f"Технологический колледж ВСГУТУ", 12, italic=True)

    HDR = 5
    ws.append([])
    ws.append(headers)
    for cell in ws[HDR]:
        cell.font = Font(name=FNT, size=SZ, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    averages = []
    for i, s in enumerate(rows):
        recs = s.get("records") or {}
        row = [s.get("surname", ""), s.get("name", "")]
        for k in keys:
            val = recs.get(k, "")
            if not val and k.endswith("_retake"):
                base = (recs.get(k[:-len("_retake")], "") or "").strip()
                failed = base.startswith(("2", "Н")) or "Не зачтено" in base
                val = "" if failed else "—"
            row.append(val)
        avg = round(float(s.get("average") or 0), 2)
        averages.append(avg)
        row.append(avg if avg > 0 else "")
        ws.append(row)
        r = HDR + 1 + i
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name=FNT, size=SZ, bold=(c == ncols))
            cell.alignment = Alignment(horizontal="left" if c <= 2 else "center", vertical="center")

    vals = [a for a in averages if a > 0]
    total_row = HDR + len(rows) + 1
    ws.cell(row=total_row, column=1, value="Средний по группе:")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=ncols - 1)
    ws.cell(row=total_row, column=1).font = Font(name=FNT, size=SZ, bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    tc = ws.cell(row=total_row, column=ncols, value=round(sum(vals) / len(vals), 2) if vals else "—")
    tc.font = Font(name=FNT, size=SZ, bold=True)
    tc.alignment = Alignment(horizontal="center")

    _autofit(ws, ncols, [HDR] + list(range(HDR + 1, total_row + 1)), min_w=10, max_w=44)
    ws.row_dimensions[HDR].height = 62
    ws.freeze_panes = f"C{HDR + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_vedomost_xlsx(group: str, subject: str, term: dict, form: str, rows,
                        teacher: str = "") -> bytes:
    """Экзаменационно-зачётная (аттестационная) ведомость: группа+предмет+семестр,
    список студентов с ИТОГОВОЙ оценкой, форма контроля, дата и строка подписи.
    rows — [{surname, name, patronymic, grade}]. term — {year, semester}."""
    year = (term or {}).get("year", "")
    sem = (term or {}).get("semester", "")
    sem_txt = "осенний" if sem == 1 else ("весенний" if sem == 2 else str(sem))

    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r'[\[\]:*?/\\]', '_', f"Ведомость {group}")[:31]
    headers = ["№", "Фамилия Имя Отчество", "Итоговая оценка", "Подпись"]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    _title(ws, 1, last_col, "Технологический колледж ВСГУТУ", 12)
    _title(ws, 2, last_col, "Экзаменационная ведомость" if (form or "").lower().startswith("экз")
           else "Зачётно-экзаменационная ведомость", 16, bold=True)
    _title(ws, 3, last_col, f"Группа {group}  ·  {subject}", SZ, bold=True)
    _title(ws, 4, last_col, f"{year}, {sem_txt} семестр"
           + (f"  ·  форма контроля: {form}" if form else ""), 12, italic=True)

    HDR = 6
    for _ in range(5):
        ws.append([])
    ws.append(headers)
    for cell in ws[HDR]:
        cell.font = Font(name=FNT, size=SZ, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, s in enumerate(rows):
        fio = " ".join(x for x in (s.get("surname", ""), s.get("name", ""),
                                   s.get("patronymic", "")) if x).strip()
        if s.get("patronymic") and s.get("name", "").endswith(s.get("patronymic", "")):
            fio = f"{s.get('surname','')} {s.get('name','')}".strip()
        grade = (s.get("grade") or "").strip()
        ws.append([i + 1, fio, grade, ""])
        r = HDR + 1 + i
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(name=FNT, size=SZ)
            cell.alignment = Alignment(horizontal="left" if c == 2 else "center", vertical="center")

    foot = HDR + len(rows) + 2
    ws.cell(row=foot, column=1, value=f"Дата: {datetime.now().strftime('%d.%m.%Y')}").font = Font(name=FNT, size=SZ)
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=2)
    sign = ws.cell(row=foot, column=3, value=f"Преподаватель: {teacher or '____________'}")
    sign.font = Font(name=FNT, size=SZ)
    ws.merge_cells(start_row=foot, start_column=3, end_row=foot, end_column=ncols)

    _autofit(ws, ncols, [HDR] + list(range(HDR + 1, HDR + len(rows) + 1)), min_w=6, max_w=55)
    ws.column_dimensions["A"].width = 5
    ws.row_dimensions[HDR].height = 30
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


#Расписание группы: сетка «пары × дни», по таблице на каждую неделю.
#Данные приходят уже слитыми (портал + правки админа) — тем же кодом, что видит студент,
#иначе выгрузка разошлась бы с сайтом.
_DAYS = ["Пнд", "Втр", "Срд", "Чтв", "Птн", "Сбт"]
_WEEK_NAME = {1: "I неделя (нечётная)", 2: "II неделя (чётная)"}


def schedule_cell_text(lesson: dict) -> str:
    """Текст клетки: предмет (тип), преподаватель, аудитория — каждый со своей строки.

    Общая функция для xlsx и docx: расхождение форматов между двумя выгрузками одного
    и того же расписания выглядит как ошибка данных, хотя это была бы ошибка вёрстки."""
    subject = (lesson.get("subject") or "").strip()
    kind = (lesson.get("kind") or "").strip()
    head = f"{subject} ({kind})" if subject and kind else (subject or kind)
    parts = [head, (lesson.get("teacher") or "").strip(), (lesson.get("room") or "").strip()]
    return "\n".join(p for p in parts if p)


def _week_pairs(days: dict) -> list:
    """Номера пар, которые в этой неделе реально есть (пустые строки не рисуем)."""
    nums = set()
    for lessons in (days or {}).values():
        for ls in (lessons or []):
            try:
                nums.add(int(ls.get("pair_no") or 0))
            except (TypeError, ValueError):
                continue
    return sorted(n for n in nums if n)


def build_schedule_xlsx(group: str, weeks: dict, pair_times=None) -> bytes:
    """Расписание группы в xlsx. weeks — {"1": {"Пнд": [пары]}, ...} как отдаёт сервер."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    ncols = 2 + len(_DAYS)                     #«Пара» + «Время» + дни
    last_col = get_column_letter(ncols)

    row = 1
    _title(ws, row, last_col, f"Расписание занятий — {group}", SZ + 2, bold=True)
    row += 1
    _title(ws, row, last_col, f"Сформировано {datetime.now().strftime('%d.%m.%Y')}",
           SZ - 2, italic=True)
    row += 2

    measured = []                              #строки для расчёта ширины столбцов
    for wk in sorted((weeks or {}).keys(), key=lambda x: str(x)):
        days = (weeks or {}).get(wk) or {}
        pairs = _week_pairs(days)
        if not pairs:
            continue                           #неделя без пар — таблицу не рисуем
        try:
            wk_int = int(wk)
        except (TypeError, ValueError):
            wk_int = 0
        _title(ws, row, last_col, _WEEK_NAME.get(wk_int, f"Неделя {wk}"), SZ, bold=True)
        row += 1

        hdr = row
        for col, name in enumerate(["Пара", "Время"] + _DAYS, start=1):
            c = ws.cell(row=hdr, column=col, value=name)
            c.font = Font(name=FNT, size=SZ, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        measured.append(hdr)
        row += 1

        for pair_no in pairs:
            #Время берём из первой найденной пары с этим номером: у портала оно одно
            #на весь поток, но в ручных правках может быть не заполнено.
            time_text = ""
            for day in _DAYS:
                for ls in (days.get(day) or []):
                    if int(ls.get("pair_no") or 0) == pair_no and (ls.get("time") or "").strip():
                        time_text = ls["time"].strip()
                        break
                if time_text:
                    break
            if not time_text and pair_times:
                idx = pair_no - 1
                if 0 <= idx < len(pair_times):
                    time_text = str(pair_times[idx])

            values = [str(pair_no), time_text]
            for day in _DAYS:
                found = [ls for ls in (days.get(day) or [])
                         if int(ls.get("pair_no") or 0) == pair_no]
                values.append("\n\n".join(schedule_cell_text(ls) for ls in found))
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name=FNT, size=SZ)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = BORDER
            measured.append(row)
            row += 1
        row += 1                               #пустая строка между неделями

    _autofit(ws, ncols, measured, min_w=10, max_w=32)
    ws.column_dimensions["A"].width = 6
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
