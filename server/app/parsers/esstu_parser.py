"""
esstu_parser.py — парсер данных с сайта ВСГУТУ (esstu.ru) о специальностях и
учебных планах Технологического колледжа. Используется для автоматического
импорта справочника специальностей и часов/ЗЕТ по предметам при создании
группы в GradeBookAI (админка → «Группы»).

Разведано вручную (реальными запросами, не по документации сайта — её нет):

━━ СПИСОК СПЕЦИАЛЬНОСТЕЙ ━━
Страница `/uportal/departments/college.htm` — обычная (не JS-рендерная) HTML-
страница, кодировка сайт сам отдаёт как utf-8 (Content-Type в ответе), в
отличие от portal.esstu.ru у расписания (windows-1251, см. schedule/parser.py
— НЕ путать эти два разных сайта ВСГУТУ). Таблица специальностей — ЕДИНСТВЕННАЯ
на странице, идёт сразу после заголовка `<h4>Специальности</h4>` (этот текст
на странице встречается только там — проверено). Строка = 6 ячеек: код+
название+квалификация, кафедра, срок на базе 9 кл., срок на базе 11 кл., ссылка
«ФГОС» (общий гос.стандарт — НЕ то, что нам нужно), ссылка «ППССЗ».

━━ УЧЕБНЫЙ ПЛАН (часы/ЗЕТ) ━━
Ссылка «ППССЗ» ведёт НЕ на файл, а на страницу
`/uportal/test/directions2.htm?code=<код>&lvlCode=SPO9` со таблицей «Год
набора» (`itemprop="eduAccred"`, одна строка на год) — то, что реально нужно
для `enrollment_year`, лежит в колонке `itemprop="educationPlan"` этой
страницы. Ссылка «ФГОС» с самой college.htm — это ДРУГОЙ документ (сам гос.
стандарт целиком, без разбивки по дисциплинам ВСГУТУ), поэтому она нам не
подходит и в get_study_plan не используется вовсе, хотя в специальностях мы
её тоже отдаём (fgos_url) — вдруг понадобится показать её как есть.

Файл учебного плана (`educationPlan`) на практике ВСЕГДА оказывается PDF
(проверено на нескольких специальностях) — вопреки ожиданию из ТЗ, что
xlsx может быть основным форматом. xlsx-ветка оставлена в коде и рабочая
(если её файл встретится), но проверить её было не на чем, и она заведомо
слабее протестирована PDF-ветки — см. предупреждение в _parse_xlsx_plan.

━━ ФОРМАТ PDF ━━
Таблица «Учебный план» внутри PDF (обычно раздел с номером «3.») — со
сложной «повёрнутой» многоярусной шапкой (заголовки колонок нарисованы по
одной букве в столбик), из-за которой заголовок у pdfplumber ПОЛЗЁТ, а вот
СТРОКИ ДАННЫХ (дисциплина → числа) выравниваются в обычную сетку. Колонка
«Всего часов» — ПОДТВЕРЖДЁННО НЕ фиксирована по индексу: у одной специальности
(09.02.07) это колонка 9, у другой (09.02.06) — колонка 8 (разное число
столбцов «форм промежуточной аттестации» перед ней). Единственный надёжный
способ, который получилось найти без верстальщика под рукой — САМОПРОВЕРКА:
план всегда содержит циклы-агрегаты («ОГСЭ», «ЕН», …) и их дисциплины
(«ОГСЭ.01», «ОГСЭ.02», …), а часы цикла ВСЕГДА равны (или чуть больше — из-за
отдельных «безындексных» строк вроде «Индивидуального проекта») сумме часов
его дисциплин. Перебираем колонки, ищем ту, где это совпадает для наибольшего
числа циклов — см. _detect_hours_column. Заодно эта же проверка САМА отсеивает
посторонние таблицы дальше в документе («Матрица компетенций» и т.п., где
структура похожая по виду, но числа — не часы), потому что там совпадений
не находится вовсе.

Явной колонки «ЗЕТ» в этих PDF НЕТ — колледжи (СПО) считают часами и
неделями, зачётные единицы — вузовское понятие. zet в возвращаемых записях
поэтому считается ТЕМ ЖЕ способом, что и весь остальной продукт
(`study_hours.ZET_HOURS` = 36 ч./ЗЕТ), а не берётся из документа.

✅ **Шаблон 2024+ поддержан (было ограничением, теперь починено).** Изначально
план 2024/2025 читался в 0 строк — казалось, что у ВСГУТУ появился ВТОРОЙ,
совсем другой шаблон PDF. Живая доразведка (после того, как заказчик прогнал
фичу на реальной группе с годом 2024) показала, что дело куда проще: в
шаблоне 2024+ перед колонкой индекса дисциплины появилась ОДНА пустая
служебная колонка (индекс уехал с `row[0]` на `row[1]`, название — с `row[1]`
на `row[2]`), а `_detect_hours_column`/`_pdf_table_to_rows` жёстко читали
`row[0]`/`row[1]` — самопроверка агрегатов смотрела в пустую колонку и не
находила НИЧЕГО, отсюда и походило на «структура совсем другая». Сама сетка
часов и колонка «Всего» устроены ТАК ЖЕ (та же самопроверка их находит) —
просто индекс/название сдвинуты. Колонка индекса теперь тоже определяется
самопроверкой, а не хардкодом — см. `_detect_index_column`: считаем, на какой
из первых нескольких колонок чаще всего встречаются значения вида «ОУП.01»/
«ОУП», а не предполагаем «это всегда колонка 0». Проверено на ОБЕИХ
специальностях (09.02.07, 09.02.06) за все года, для которых на сайте вообще
есть ссылка на план (2022-2026) — везде читается штатно.

━━ ЧАСЫ ПО СЕМЕСТРАМ ━━
Правее «Всего» в тех же строках лежит сетка «Семестр 1» … «Семестр 8» — те же
подколонки (Лекции/Пр.занятия/Лаб.занятия/…), просто повторённые 8 раз с
ФИКСИРОВАННЫМ ШАГОМ между семестрами. Шаг (period) и позиция первого семестра
(base) — ТАК ЖЕ НЕ фиксированы между документами, как и «Всего» (проверено: у
09.02.07 — база 16/шаг 7, у 09.02.06 — база 21/шаг 13), и определяются той же
самопроверкой: сумма часов дисциплины по всем 8 найденным семестровым ячейкам
обязана совпасть с её же «Всего». ⚠️ Калибровку ведём ТОЛЬКО по дисциплинам,
где часы разбросаны МИНИМУМ по двум семестрам (`_MIN_SEMESTER_GRID_HITS`) —
однa непустая ячейка «угадывает» практически любую сетку одной точкой (первая
попытка без этого ограничения нашла шаг 4 вместо истинных 13, потому что
односеместровых дисциплин в плане большинство, и они не различают сетки).
См. `_detect_semester_grid`. Не нашлась сетка (структура ещё сильнее
отличается) — `hours_by_semester` пустой у ВСЕХ дисциплин этого плана,
`hours`/`zet` (общий итог) при этом остаются — деградация, а не падение.

━━ ЗАВИСИМОСТИ ━━
requests (уже в проекте), openpyxl (уже в проекте, план в xlsx). pdfplumber —
НОВАЯ, добавлена в server/requirements.txt; импортируется ЛЕНИВО (осознанный
приём проекта для опциональных пакетов, см. CLAUDE.md §7) — без неё PDF-планы
просто не читаются (пустой список + предупреждение в лог), список
специальностей читается всегда (requests + html.parser, без сторонних
HTML-парсеров — тот же приём, что и у schedule/parser.py).
"""
from __future__ import annotations

import logging
import re
from html import unescape
from html.parser import HTMLParser
from io import BytesIO

import requests

logger = logging.getLogger("esstu_parser")

BASE_URL = "https://esstu.ru"
COLLEGE_URL = BASE_URL + "/uportal/departments/college.htm"
DIRECTIONS_URL = BASE_URL + "/uportal/test/directions2.htm"
_TIMEOUT_PAGE = 20
_TIMEOUT_FILE = 40
_HEADERS = {"User-Agent": "GradeBookAI/3.5 esstu-parser (+esstu-gradebook.ru)"}

_ZET_HOURS = 36.0  #то же правило, что study_hours.ZET_HOURS — не дублируем импортом
                    #модуля целиком (у server/app своя зона зависимостей), но ЧИСЛО то же.


def _fetch(url: str, timeout: int = _TIMEOUT_PAGE):
    """GET с таймаутом и понятным логом. None — сайт недоступен/ошибка HTTP —
    вызывающий код обязан признать это НЕ падением, а пустым результатом."""
    try:
        resp = requests.get(url, timeout=timeout, verify=True, headers=_HEADERS)
        resp.raise_for_status()
        return resp
    except requests.RequestException as e:
        logger.warning("[esstu] запрос %s не удался: %s", url, e)
        return None


def _absolute(url: str) -> str:
    if not url:
        return ""
    if url.startswith("http://") or url.startswith("https://"):
        return url
    if url.startswith("/"):
        return BASE_URL + url
    return BASE_URL + "/" + url


def _to_number(v):
    """Строка/число из ячейки таблицы → float либо None (пусто/не число)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".").replace("\xa0", "")
    if not s:
        return None
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else None


def _zet_from_hours(hours):
    return round(hours / _ZET_HOURS, 2) if hours else None


# ══════════════════════════════════════════════════════════════════════════
# 1. Список специальностей (college.htm)
# ══════════════════════════════════════════════════════════════════════════

_SPECIALTY_CODE_RE = re.compile(r"^(\d{2}\.\d{2}\.\d{2})\s*(.*)$")
_QUALIFICATION_RE = re.compile(r"^квалификаци[ия]\s*[-–—]\s*(.+)$", re.IGNORECASE)


class _SpecialtyTableParser(HTMLParser):
    """Строит таблицу «Специальности» построчно (тот же приём, что
    schedule/parser.py::_TableRowParser — html.parser потоком, без сторонних
    HTML-парсеров), но с ОДНИМ отличием: внутри первой ячейки строки два
    абзаца («код + название» и «квалификация – …»), и их нельзя схлопывать
    в одну строку текста — граница абзаца (<p>/<br>) отмечается отдельным
    маркером и разбирается позже, при сборке словаря специальности."""

    _SEP = "\x00"

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows: list[list[tuple[list[str], str]]] = []
        self._seen_heading = False
        self._in_table = False
        self._table_depth = 0
        self._row: list | None = None
        self._cell: list[str] | None = None
        self._cell_href = ""

    def handle_starttag(self, tag, attrs):
        t = tag.lower()
        if not self._seen_heading:
            return
        if not self._in_table:
            if t == "table":
                self._in_table = True
                self._table_depth = 1
            return
        if t == "table":
            self._table_depth += 1
        elif t == "tr":
            if self._row is not None:
                self._flush_cell()
                self.rows.append(self._row)
            self._row = []
            self._cell = None
        elif t == "td":
            if self._row is None:
                self._row = []
            self._flush_cell()
            self._cell = []
            self._cell_href = ""
        elif t in ("p", "br") and self._cell:
            self._cell.append(self._SEP)
        elif t == "a" and self._cell is not None and not self._cell_href:
            href = dict(attrs).get("href") or ""
            if href:
                self._cell_href = href

    def handle_endtag(self, tag):
        t = tag.lower()
        if not self._in_table:
            return
        if t == "table":
            self._table_depth -= 1
            if self._table_depth == 0:
                if self._row is not None:
                    self._flush_cell()
                    self.rows.append(self._row)
                    self._row = None
                self._in_table = False
        elif t == "tr" and self._row is not None:
            self._flush_cell()
            self.rows.append(self._row)
            self._row = None

    def handle_data(self, data):
        if not self._seen_heading:
            if "Специальности" in data:
                self._seen_heading = True
            return
        if self._cell is not None:
            self._cell.append(data)

    def _flush_cell(self):
        if self._cell is not None and self._row is not None:
            raw = unescape("".join(self._cell)).replace("\xa0", " ")
            paragraphs = [re.sub(r"\s+", " ", p).strip() for p in raw.split(self._SEP)]
            paragraphs = [p for p in paragraphs if p]
            self._row.append((paragraphs, self._cell_href))
        self._cell = None
        self._cell_href = ""


def get_all_specialties() -> list[dict]:
    """Список специальностей колледжа с college.htm.

    Возвращает list[dict] с ключами: code, name, qualification, department,
    duration_9, duration_11, fgos_url, directions_url. [] — сайт недоступен
    ИЛИ таблица не нашлась (вёрстка страницы изменилась) — не падаем, вызывающий
    код должен показать это пользователю как «справочник сейчас недоступен»,
    а не как ошибку 500."""
    resp = _fetch(COLLEGE_URL)
    if resp is None:
        return []
    if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
        resp.encoding = resp.apparent_encoding or "utf-8"

    parser = _SpecialtyTableParser()
    try:
        parser.feed(resp.text)
        parser.close()
    except Exception as e:
        logger.warning("[esstu] не удалось разобрать HTML страницы специальностей: %s", e)
        return []

    out: list[dict] = []
    for row in parser.rows:
        if len(row) < 6:
            continue  # служебная/шапочная строка — не 6 колонок данных
        name_paras, _ = row[0]
        if not name_paras:
            continue
        m = _SPECIALTY_CODE_RE.match(name_paras[0])
        if not m:
            continue  # первый абзац не начинается с кода — не строка специальности
        code, name = m.group(1), m.group(2).strip()
        qualification = ""
        for p in name_paras[1:]:
            qm = _QUALIFICATION_RE.match(p)
            if qm:
                qualification = qm.group(1).strip()
                break
        dept_paras, _ = row[1]
        dur9_paras, _ = row[2]
        dur11_paras, _ = row[3]
        fgos_paras, fgos_href = row[4]
        out.append({
            "code": code,
            "name": name,
            "qualification": qualification,
            "department": "; ".join(dept_paras),
            "duration_9": "; ".join(dur9_paras),
            "duration_11": "; ".join(dur11_paras),
            "fgos_url": _absolute(fgos_href),
            #Ссылка на «ППССЗ» с самой college.htm иногда вообще без href (см.
            #заголовок модуля — на боевой странице такое реально встречается,
            #строка «40.02.04 Юриспруденция»), поэтому строим URL из кода сами —
            #он предсказуем и не зависит от того, приклеили ли ссылку на сайте.
            "directions_url": f"{DIRECTIONS_URL}?code={code}&lvlCode=SPO9",
        })

    if not out:
        logger.warning("[esstu] таблица специальностей найдена, но ни одной строки с кодом "
                       "не распознано — похоже, вёрстка страницы изменилась")
    return out


# ══════════════════════════════════════════════════════════════════════════
# 2. Подбор специальности по имени группы
# ══════════════════════════════════════════════════════════════════════════

#Расширяемый маппинг «буквенный префикс имени группы → код специальности».
#Заполняется ПРОВЕРЕННЫМИ соответствиями по мере необходимости, а НЕ выводится
#автоматически из справочника специальностей: префикс группы — внутреннее
#обозначение колледжа и не обязан совпадать с буквами из названия специальности
#(частый случай — кафедральная нумерация вида «К74/1», где «74» ничего не
#говорит о специальности без отдельного, никем сейчас не опубликованного
#списка соответствий кафедра→код; такие префиксы добавлять в словарь только
#когда соответствие подтверждено, а не угадано).
GROUP_PREFIX_TO_SPECIALTY: dict[str, str] = {
    "ИС": "09.02.07",  # Информационные системы и программирование
    "КС": "09.02.06",  # Сетевое и системное администрирование
}

_GROUP_PREFIX_RE = re.compile(r"^[А-ЯЁ]+")


def match_specialty(group_name: str, specialties: list[dict] | None = None) -> dict | None:
    """Подобрать специальность по имени группы («ИС-21» → 09.02.07). None —
    префикс неизвестен GROUP_PREFIX_TO_SPECIALTY, справочник специальностей
    пуст/недоступен, либо код из маппинга не нашёлся в справочнике.

    specialties — свой уже загруженный список (если он под рукой), иначе
    будет запрошен заново через get_all_specialties()."""
    prefix_m = _GROUP_PREFIX_RE.match((group_name or "").strip().upper())
    if not prefix_m:
        return None
    code = GROUP_PREFIX_TO_SPECIALTY.get(prefix_m.group(0))
    if not code:
        return None
    if specialties is None:
        specialties = get_all_specialties()
    for s in specialties:
        if s.get("code") == code:
            return s
    return None


# ══════════════════════════════════════════════════════════════════════════
# 3. Учебный план конкретного года набора (directions2.htm → файл плана)
# ══════════════════════════════════════════════════════════════════════════

_YEAR_RE = re.compile(r"^(19|20)\d{2}$")


class _DirectionsYearParser(HTMLParser):
    """Строка `<tr itemprop="eduAccred">` на directions2.htm = один год набора;
    внутри нужна ссылка из `<td itemprop="educationPlan">`. Год в строке идёт
    ПЕРВОЙ ячейкой (без itemprop) и по документу всегда раньше ссылки на план —
    порядок не отслеживаем отдельно, полагаемся на естественный порядок тегов
    (то же допущение, что уже сделано у html.parser-разборов в проекте)."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.years: dict[int, str] = {}
        self._in_row = False
        self._year = None
        self._cur_prop = None

    def handle_starttag(self, tag, attrs):
        adict = dict(attrs)
        t = tag.lower()
        if t == "tr":
            self._in_row = adict.get("itemprop") == "eduAccred"
            self._year = None
            self._cur_prop = None
            return
        if not self._in_row:
            return
        if t == "td":
            self._cur_prop = adict.get("itemprop")
        elif t == "a" and self._cur_prop == "educationPlan":
            href = adict.get("href") or ""
            if href and self._year is not None:
                self.years.setdefault(self._year, href)

    def handle_endtag(self, tag):
        if tag.lower() == "tr":
            self._in_row = False

    def handle_data(self, data):
        if self._in_row and self._year is None:
            text = data.strip()
            if _YEAR_RE.match(text):
                self._year = int(text)


def _fetch_plan_years(specialty_code: str) -> dict[int, str]:
    url = f"{DIRECTIONS_URL}?code={specialty_code}&lvlCode=SPO9"
    resp = _fetch(url)
    if resp is None:
        return {}
    if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
        resp.encoding = resp.apparent_encoding or "utf-8"
    parser = _DirectionsYearParser()
    try:
        parser.feed(resp.text)
        parser.close()
    except Exception as e:
        logger.warning("[esstu] не удалось разобрать страницу учебных планов %s: %s",
                       specialty_code, e)
        return {}
    return parser.years


def get_study_plan(specialty_code: str, enrollment_year: int) -> list[dict]:
    """Учебный план (часы/ЗЕТ по дисциплинам) специальности за год набора.

    Возвращает list[{subject, index, hours, zet, hours_by_semester}], где
    hours_by_semester — {номер_семестра(1-8): часы} — может быть ПУСТЫМ
    словарём у отдельной дисциплины (сноска без своего семестра в исходнике)
    или у ВСЕХ дисциплин сразу (сетка семестров не откалибровалась на этом
    документе — см. заголовок модуля, «Часы по семестрам»); `hours`/`zet`
    (общий итог по дисциплине) при этом всегда на месте, если распознан хотя
    бы столбец «Всего». [] — план для этого года не опубликован, файл не
    открылся, формат не распознан, ИЛИ структуру таблицы не удалось надёжно
    определить (см. _detect_hours_column) — в каждом случае причина уходит в
    лог, наружу — просто пустой список, никогда не выдуманные цифры."""
    years = _fetch_plan_years(specialty_code)
    href = years.get(int(enrollment_year))
    if not href:
        logger.info("[esstu] для %s набора %s учебный план не найден (доступные годы: %s)",
                    specialty_code, enrollment_year, sorted(years) or "нет ни одного")
        return []

    file_resp = _fetch(_absolute(href), timeout=_TIMEOUT_FILE)
    if file_resp is None:
        return []
    content = file_resp.content
    if content[:4] == b"%PDF":
        return _parse_pdf_plan(content)
    if content[:2] == b"PK":  # xlsx = zip-контейнер
        return _parse_xlsx_plan(content)
    logger.warning("[esstu] учебный план %s/%s — нераспознанный формат файла (не PDF, не xlsx)",
                   specialty_code, enrollment_year)
    return []


# ── PDF-вариант (реально встречающийся формат) ──────────────────────────────

#Индекс дисциплины/цикла в учебном плане: 1-6 букв (кириллица или латиница —
#в реальных файлах бывает опечаткой смешанный алфавит, напр. «ЭM.11»), затем
#точка или пробел, затем цифра. Отсекает мусорные строки-сноски вида «*»,
#«МДК*» и заголовочные строки без цифр («Индекс», «Итого час/нед …»).
_DISCIPLINE_INDEX_RE = re.compile(r"^[A-ZА-ЯЁ]{1,6}[\s.]*\d")
#«Голый» код цикла БЕЗ цифр (ОГСЭ, ЕН, ОПЦ, ПП, ОУП, СО, ОП…) — потенциальный
#агрегат, у которого есть дочерние дисциплины ОГСЭ.01, ОГСЭ.02 и т.п.
_CYCLE_AGGREGATE_RE = re.compile(r"^[A-ZА-ЯЁ]{1,6}$")
_MIN_GROUP_MATCHES = 2  # меньше — слишком велик риск случайного совпадения


def _detect_index_column(rows: list[list]) -> int:
    """Колонка с индексом дисциплины/цикла («ОУП.01», «ОГСЭ») — НЕ ВСЕГДА
    самая первая колонка таблицы. В шаблоне ВСГУТУ до 2024 г. это была
    колонка 0; в шаблоне 2024+ перед ней появилась ОДНА пустая служебная
    колонка (0 всегда пуст, индекс уехал в 1), а колонка «Всего часов»
    сдвинулась вместе с ним — это и было причиной, по которой самопроверка
    часов не находила НИЧЕГО на новых планах: она искала агрегаты в row[0],
    а там пусто. Определяем колонку так же, самопроверкой: где чаще всего
    встречаются значения вида «ОУП.01»/«ОУП» на первых нескольких колонках
    (индекс всегда у самого начала строки, дальше — трёхзначные коды
    компетенций и цифры часов, под тот же паттерн не подходящие)."""
    max_len = max((len(r) for r in rows if r), default=0)
    best_col, best_count = 0, -1
    for col in range(min(max_len, 6)):
        count = 0
        for row in rows:
            if not row or col >= len(row) or not row[col]:
                continue
            v = str(row[col]).strip()
            if _DISCIPLINE_INDEX_RE.match(v) or _CYCLE_AGGREGATE_RE.match(v):
                count += 1
        if count > best_count:
            best_count, best_col = count, col
    return best_col


def _detect_hours_column(rows: list[list], index_col: int = 0) -> int | None:
    """Определяет индекс колонки «Всего часов» САМОПРОВЕРКОЙ: часы цикла-
    агрегата (ОГСЭ) обязаны быть БЛИЗКИ к сумме часов его дисциплин (ОГСЭ.01 +
    ОГСЭ.02 + …) — «близки», а не «равны», потому что у цикла бывают ещё
    строки-сноски без собственного индекса (напр. «Индивидуальный проект»),
    которые в сумму дочерних не попадают, но это всегда ЧАСТЬ родителя, не
    вычесть в минус, только недобор. Проверено на нескольких специальностях
    и ДВУХ поколениях шаблона (09.02.07/2022 → колонка 9, 09.02.06/2022 →
    колонка 8, 09.02.07/2024 → колонка 12) — колонка НЕ ФИКСИРОВАНА между
    разными планами, перебор обязателен каждый раз заново. index_col — от
    _detect_index_column, а НЕ обязательно 0 (см. её докстринг)."""
    groups = []
    for i, row in enumerate(rows):
        if not row or index_col >= len(row) or not row[index_col]:
            continue
        idx = str(row[index_col]).strip()
        if not _CYCLE_AGGREGATE_RE.match(idx):
            continue
        children = [r for r in rows if r and index_col < len(r) and r[index_col]
                   and str(r[index_col]).strip().startswith(idx + ".")]
        if len(children) >= _MIN_GROUP_MATCHES:
            groups.append((row, children))
    if not groups:
        return None

    max_len = max(len(r) for r in rows if r)
    best_col, best_score = None, 0
    for col in range(index_col + 2, max_len):
        score = 0
        for agg, children in groups:
            agg_v = _to_number(agg[col]) if col < len(agg) else None
            if agg_v is None or agg_v <= 0:
                continue
            total = 0.0
            ok = True
            for c in children:
                v = _to_number(c[col]) if col < len(c) else None
                if v is None:
                    ok = False
                    break
                total += v
            if ok and agg_v * 0.5 <= total <= agg_v + 0.5:
                score += 1
        if score > best_score:  # первая по счёту (самая левая) колонка с лучшим счётом
            best_score, best_col = score, col
    return best_col if best_score >= _MIN_GROUP_MATCHES else None


#Сколько семестров в учебном плане СПО в принципе бывает (4 курса × 2) — сетка
#никогда не длиннее, даже если у специальности реально меньше (4-6): лишние
#позиции просто не попадут в границы строки и отсеются по len(row).
_SEMESTER_COUNT = 8
#Разумный диапазон шага между соседними семестровыми блоками колонок — нашлись
#7 (09.02.07) и 13 (09.02.06), берём с запасом в обе стороны.
_SEMESTER_PERIOD_RANGE = range(4, 16)
#Калибровочная строка обязана иметь часы МИНИМУМ в двух семестровых ячейках
#сетки-кандидата — одна непустая ячейка «угадывает» почти любую сетку одной
#точкой (см. заголовок модуля: без этого ограничения на реальном файле находился
#шаг 4 вместо истинных 13, потому что односеместровых дисциплин большинство).
_MIN_SEMESTER_GRID_HITS = 2
#Сколько дисциплин ВСЕГО должны так откалиброваться, чтобы сетке доверять —
#меньше похоже на случайное совпадение на шумных данных.
_MIN_SEMESTER_GRID_SCORE = 3


def _detect_semester_grid(rows: list[list], hours_col: int, max_len: int, index_col: int = 0):
    """(base, period) — колонка «Семестр 1» и шаг между семестрами, либо
    (None, None), если сетка не откалибровалась. Тот же приём самопроверки,
    что и у _detect_hours_column, только сумма берётся не по детям цикла, а
    по 8 колонкам-кандидатам одной и той же дисциплины — она обязана сойтись
    с уже известным «Всего» (hours_col) этой же строки."""
    disc_rows = [r for r in rows if r and index_col < len(r) and r[index_col]
                and _DISCIPLINE_INDEX_RE.match(str(r[index_col]).strip())]
    best_score, best_base, best_period = 0, None, None
    for period in _SEMESTER_PERIOD_RANGE:
        for base in range(hours_col + 1, max_len - period):
            cols = [base + i * period for i in range(_SEMESTER_COUNT) if base + i * period < max_len]
            if len(cols) < _MIN_SEMESTER_GRID_HITS:
                continue
            score = 0
            for row in disc_rows:
                total_v = _to_number(row[hours_col]) if hours_col < len(row) else None
                if not total_v:
                    continue
                total, hits = 0.0, 0
                for c in cols:
                    v = _to_number(row[c]) if c < len(row) else None
                    if v:
                        total += v
                        hits += 1
                if hits >= _MIN_SEMESTER_GRID_HITS and abs(total - total_v) < 0.5:
                    score += 1
            if score > best_score:
                best_score, best_base, best_period = score, base, period
    if best_score < _MIN_SEMESTER_GRID_SCORE:
        return None, None
    return best_base, best_period


def _row_hours_by_semester(row: list, base: int, period: int) -> dict:
    out = {}
    if base is None:
        return out
    for i in range(_SEMESTER_COUNT):
        c = base + i * period
        if c >= len(row):
            break
        v = _to_number(row[c])
        if v:
            out[i + 1] = v
    return out


def _pdf_table_to_rows(table: list[list], hours_col: int, sem_base=None, sem_period=None,
                       index_col: int = 0) -> list[dict]:
    name_col = index_col + 1
    out = []
    for row in table:
        if not row or index_col >= len(row) or not row[index_col]:
            continue
        index = str(row[index_col]).strip()
        if not _DISCIPLINE_INDEX_RE.match(index):
            continue
        name = re.sub(r"\s+", " ",
                      str((row[name_col] if name_col < len(row) else "") or "").strip())
        if not name:
            continue
        hours = _to_number(row[hours_col]) if hours_col < len(row) else None
        out.append({
            "subject": name,
            #Индекс — НЕ из ТЗ, но нужен: одинаковые названия встречаются в
            #разных циклах плана с РАЗНЫМИ часами (напр. «История» — и в общих
            #предметах, и отдельно в ОГСЭ) — без индекса такие строки в общем
            #списке неотличимы.
            "index": index,
            "hours": hours,
            "zet": _zet_from_hours(hours),
            #{номер семестра 1-8: часы} — пусто, если сетка не откалибровалась
            #на этом документе (см. _detect_semester_grid); не гадаем.
            "hours_by_semester": _row_hours_by_semester(row, sem_base, sem_period),
        })
    return out


def _parse_pdf_plan(content: bytes) -> list[dict]:
    try:
        import pdfplumber
    except Exception as e:
        logger.warning("[esstu] pdfplumber не установлен — PDF учебного плана не прочитан "
                       "(pip install pdfplumber): %s", e)
        return []

    try:
        pdf = pdfplumber.open(BytesIO(content))
    except Exception as e:
        logger.warning("[esstu] не удалось открыть PDF учебного плана: %s", e)
        return []

    anchor_col = None
    anchor_index_col = None
    anchor_width = None
    anchor_rows = None
    out: list[dict] = []
    try:
        with pdf:
            page_tables = []  # (page_index, table_rows) — на случай многостраничной таблицы
            for pi, page in enumerate(pdf.pages):
                try:
                    tables = page.extract_tables()
                except Exception as e:
                    logger.warning("[esstu] страница %s PDF-плана не разобралась: %s", pi, e)
                    continue
                for table in tables:
                    if not table:
                        continue
                    page_tables.append((pi, table))
                    if anchor_col is None:
                        #Колонка индекса — НЕ всегда 0 (см. _detect_index_column: в
                        #шаблоне 2024+ перед ней есть пустая служебная колонка) —
                        #определяем её ПЕРЕД поиском часов, иначе самопроверка
                        #агрегатов смотрит в пустую колонку и не находит ничего.
                        idx_col = _detect_index_column(table)
                        col = _detect_hours_column(table, idx_col)
                        if col is not None:
                            anchor_col = col
                            anchor_index_col = idx_col
                            anchor_width = max(len(r) for r in table if r)
                            anchor_rows = table

            if anchor_col is None:
                logger.warning("[esstu] в PDF учебного плана не нашлось таблицы с "
                               "определяемой колонкой часов (структура непривычная — "
                               "план не прочитан, а не выдуман)")
                return []

            #Сетку семестров калибруем ОДИН раз по самой первой (якорной)
            #таблице — see _detect_semester_grid; не откалибровалась — все
            #строки этого плана останутся без hours_by_semester (см. заголовок
            #модуля), но hours/zet (общий итог) это не затрагивает.
            sem_base, sem_period = _detect_semester_grid(anchor_rows, anchor_col, anchor_width,
                                                          anchor_index_col)
            if sem_base is None:
                logger.info("[esstu] сетка часов по семестрам не откалибровалась для этого "
                           "плана — hours_by_semester будет пустым у всех дисциплин")

            #Многостраничная таблица «Учебный план» не всегда повторяет колонку
            #часов на следующих страницах в распознаваемом виде — но ширина
            #строк (число ячеек) у продолжения та же самая, это и используем.
            for pi, table in page_tables:
                width = max((len(r) for r in table if r), default=0)
                if width == anchor_width:
                    out.extend(_pdf_table_to_rows(table, anchor_col, sem_base, sem_period,
                                                  anchor_index_col))
    except Exception as e:
        logger.warning("[esstu] сбой при разборе PDF учебного плана: %s", e)
        return []

    if not out:
        logger.warning("[esstu] колонка часов определена, но ни одной строки дисциплины "
                       "не извлечено — вёрстка могла отличаться сильнее ожидаемого")
    return out


# ── xlsx-вариант (в реальности не встречен, но предусмотрен по ТЗ) ─────────

def _parse_xlsx_plan(content: bytes) -> list[dict]:
    """xlsx-план — НЕ встречен вживую при разведке (все проверенные
    специальности отдают PDF), поэтому это best-effort по общей форме таких
    файлов: ищем на любом листе строку-шапку с понятными заголовками колонок
    и читаем данные под ней. Если реальный xlsx окажется устроен иначе — эту
    функцию почти наверняка придётся поправить под его фактическую структуру,
    протестировать её было не на чем."""
    try:
        import openpyxl
    except Exception as e:
        logger.warning("[esstu] openpyxl недоступен: %s", e)
        return []
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        logger.warning("[esstu] не удалось открыть xlsx учебного плана: %s", e)
        return []

    def norm(v) -> str:
        return re.sub(r"\s+", " ", str(v or "")).strip().lower()

    out: list[dict] = []
    for ws in wb.worksheets:
        col = {}
        header_row_idx = None
        for row in ws.iter_rows(max_row=30):
            texts = [norm(c.value) for c in row]
            hits = {}
            for i, text in enumerate(texts):
                if not text:
                    continue
                if "subject" not in hits and ("дисципл" in text or "наименован" in text):
                    hits["subject"] = i
                elif "hours" not in hits and "час" in text:
                    hits["hours"] = i
                elif "zet" not in hits and ("зет" in text or "зачетн" in text or "зачётн" in text):
                    hits["zet"] = i
                elif "semester" not in hits and "семестр" in text:
                    hits["semester"] = i
            if "subject" in hits and ("hours" in hits or "zet" in hits):
                header_row_idx = row[0].row
                col = hits
                break
        if header_row_idx is None:
            continue

        sheet_rows = []
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            name_cell = row[col["subject"]].value if col.get("subject") is not None else None
            name = str(name_cell).strip() if name_cell is not None else ""
            if not name:
                continue
            hours = _to_number(row[col["hours"]].value) if "hours" in col else None
            zet = _to_number(row[col["zet"]].value) if "zet" in col else None
            semester = _to_number(row[col["semester"]].value) if "semester" in col else None
            if hours is None and zet is None:
                continue
            sheet_rows.append({
                "subject": name,
                "index": "",
                "hours": hours,
                "zet": zet if zet is not None else _zet_from_hours(hours),
                #Явной сетки по всем 8 семестрам в xlsx можно и не быть — только
                #одно число в своей колонке. Кладём его как единственный ключ,
                #а не выдумываем остальные семестры.
                "hours_by_semester": {int(semester): hours} if semester and hours else {},
            })
        if sheet_rows:
            out.extend(sheet_rows)
            break  # первый подходящий лист — этого достаточно
    if not out:
        logger.warning("[esstu] xlsx учебного плана открыт, но таблица дисциплин не найдена "
                       "(этот код ни разу не проверялся на настоящем файле — см. докстринг)")
    return out


# ══════════════════════════════════════════════════════════════════════════
# Самопроверка (python -m server.app.parsers.esstu_parser)
# ══════════════════════════════════════════════════════════════════════════

def _self_test():
    print(f"GET {COLLEGE_URL}")
    resp = _fetch(COLLEGE_URL)
    if resp is None:
        print("Сайт недоступен — см. предупреждение в логе выше.")
        return
    print(f"HTTP {resp.status_code}, {len(resp.text)} байт HTML")

    specialties = get_all_specialties()
    print(f"\nСпециальностей найдено: {len(specialties)}")
    for s in specialties[:3]:
        print(f"  {s['code']} — {s['name']} (квалификация: {s['qualification'] or '?'}, "
             f"каф.: {s['department'] or '?'})")

    print("\nmatch_specialty('ИС-21') ->",
         (match_specialty("ИС-21", specialties) or {}).get("code"))

    target_code = "09.02.07"
    target = next((s for s in specialties if s["code"] == target_code), None)
    if target is None:
        print(f"\n{target_code} не найдена в списке специальностей — план не запрашиваю.")
        return
    year = 2022
    print(f"\nGET учебный план {target_code}, набор {year}…")
    plan = get_study_plan(target_code, year)
    print(f"Строк плана: {len(plan)}")
    with_semesters = sum(1 for r in plan if r["hours_by_semester"])
    print(f"С разбивкой по семестрам: {with_semesters} из {len(plan)}")
    for row in plan[:5]:
        print(f"  {row['index']:<12} {row['subject']:<55} "
             f"{row['hours']!s:>6} ч.  ЗЕТ={row['zet']!s:<5} "
             f"по семестрам={row['hours_by_semester']}")

    #study_hours.py лежит в корне репозитория (общий с десктопом модуль, как
    #grading.py) — тот же приём, что webdata.py уже применяет для grading/
    #study_hours: добавляем корень в sys.path перед импортом. Здесь это нужно
    #только самотесту (публичные функции модуля study_hours не используют).
    import os
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__)))))
    if _root not in sys.path:
        sys.path.insert(0, _root)
    import study_hours
    course, semester = study_hours.course_and_semester(year, "2025/2026", 1)
    print(f"\nГруппа набора {year}, сейчас 2025/2026 сем.1 -> курс {course}, "
         f"семестр с начала обучения {semester}")
    current_subjects = [r for r in plan if r["hours_by_semester"].get(semester)]
    print(f"Предметов ИМЕННО в этом семестре: {len(current_subjects)}")
    for row in current_subjects[:5]:
        print(f"  {row['index']:<12} {row['subject']:<55} "
             f"{row['hours_by_semester'][semester]} ч. в этом семестре")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    _self_test()
