"""
admin_panel.py — Панель администратора ВСГУТУ Журнала
"""

import json
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QListWidget, QListWidgetItem, QStackedWidget,
    QFormLayout, QScrollArea, QCheckBox, QMessageBox,
    QComboBox, QSizePolicy, QFrame, QInputDialog
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont
from security import secure_store
from subjects import load_subjects, save_subjects, rename_subject, delete_subject, add_subject

ALL_SUBJECTS = [
    "Аварийно-спасательные работы на высоте",
    "Аварийно-спасательное, газоспасательное и пожарное оборудование и инструменты",
    "Административный процесс",
    "Архитектура аппаратных средств",
    "Архитектурно-строительные конструкции",
    "Безопасность жизнедеятельности",
    "Безопасность работ при эксплуатации и ремонте оборудования электрических подстанций и сетей",
    "Гражданское право",
    "Дискретная математика",
    "Договорное право",
    "Инженерная графика",
    "Иностранный язык",
    "Иностранный язык второй",
    "Иностранный язык в профессиональной деятельности",
    "Информационно-коммуникационные технологии в туризме и гостеприимстве",
    "Информационные технологии в профессиональной деятельности",
    "Информационные технологии в юридической деятельности",
    "История России",
    "Компьютерные сети",
    "Компьютерные сети ЭВМ",
    "Координация работы по реализации заказа экскурсионных услуг",
    "Корпоративное право и юридическое сопровождение деятельности организаций и физических лиц",
    "Метрология и стандартизация",
    "Монтаж и наладка воздушных линий электропередачи",
    "Оказание первой помощи и психологическая поддержка",
    "Операционные системы и среды",
    "Организация принципы построения и функционирования компьютерных сетей",
    "Организация ремонта и наладки устройств электроснабжения",
    "Освоение должности служащего",
    "Основы алгоритмизации и программирования",
    "Основы бережливого производства",
    "Основы ведения аварийно-спасательных работ",
    "Основы геологии",
    "Основы предпринимательской деятельности",
    "Основы проектирования баз данных",
    "Основы топографии",
    "Основы финансовой грамотности",
    "Основы эксплуатации электрооборудования",
    "Оформление и обработка заказов клиентов экскурсионных услуг",
    "Охрана труда",
    "Подготовка по профессии пожарный",
    "Подготовка по профессии промышленный альпинист",
    "Подготовка по профессии электромонтёр по обслуживанию подстанций",
    "Потенциально опасные процессы",
    "Предпринимательская деятельность в сфере туризма и гостиничного бизнеса",
    "Правовое и документационное обеспечение в туризме и гостеприимстве",
    "Психология делового общения и конфликтология",
    "Психология экстремальных ситуаций",
    "Разработка объемно-планировочных и конструктивных решений объектов капитального строительства",
    "Разработка программных модулей",
    "Разработка проектной документации по организации строительства объектов капитального строительства",
    "Сервисная деятельность в туризме и гостеприимстве",
    "Сопровождение туристов при прохождении маршрута",
    "Судебная и альтернативные формы защиты прав организаций и физических лиц",
    "Судоустройство и правоохранительные органы",
    "Техническая механика",
    "Техническое обслуживание и ремонт оборудования электрических подстанций и сетей",
    "Трудовое право",
    "Физическая культура",
    "Численные методы",
    "Экономика и бухгалтерский учёт предприятий туризма и гостиничного дела",
    "Экологические основы природопользования",
    "Электроматериаловедение",
    "Электротехника и электроника",
    "Элементы высшей математики",
]

STYLE_BTN = """
QPushButton {
    background-color: #2c3e50;
    color: white;
    border-radius: 6px;
    padding: 8px 14px;
    font-size: 13px;
}
QPushButton:hover { background-color: #34495e; }
QPushButton:pressed { background-color: #1a252f; }
"""

STYLE_BTN_RED = """
QPushButton {
    background-color: #c0392b;
    color: white;
    border-radius: 6px;
    padding: 8px 14px;
}
QPushButton:hover { background-color: #e74c3c; }
"""

STYLE_BTN_GREEN = """
QPushButton {
    background-color: #27ae60;
    color: white;
    border-radius: 6px;
    padding: 8px 14px;
}
QPushButton:hover { background-color: #2ecc71; }
"""


def make_back_btn(callback) -> QPushButton:
    btn = QPushButton("← Назад")
    btn.setStyleSheet(STYLE_BTN)
    btn.clicked.connect(callback)
    return btn



def get_all_subjects() -> list:
    """Возвращает актуальный список предметов из subjects.json."""
    return load_subjects()


class SubjectSelector(QWidget):
    """Выпадающий список предметов с множественным выбором."""
    changed = Signal(list)

    def __init__(self, selected: list = None, parent=None):
        super().__init__(parent)
        self._selected = list(selected or [])
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.btn = QPushButton("Выбрать предметы ▼")
        self.btn.setStyleSheet(STYLE_BTN)
        self.btn.clicked.connect(self._toggle)
        layout.addWidget(self.btn)

        self.panel = QWidget()
        self.panel.hide()
        panel_layout = QVBoxLayout(self.panel)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(200)
        inner = QWidget()
        self.checkboxes = []
        inner_l = QVBoxLayout(inner)
        for subj in ALL_SUBJECTS:
            cb = QCheckBox(subj)
            cb.setChecked(subj in self._selected)
            cb.stateChanged.connect(self._update)
            self.checkboxes.append(cb)
            inner_l.addWidget(cb)
        scroll.setWidget(inner)
        panel_layout.addWidget(scroll)
        layout.addWidget(self.panel)
        self._update_btn_text()

    def _toggle(self):
        self.panel.setVisible(not self.panel.isVisible())

    def _update(self):
        self._selected = [cb.text() for cb in self.checkboxes if cb.isChecked()]
        self._update_btn_text()
        self.changed.emit(self._selected)

    def _update_btn_text(self):
        if self._selected:
            self.btn.setText(f"Выбрано предметов: {len(self._selected)} ▼")
        else:
            self.btn.setText("Выбрать предметы ▼")

    def get_selected(self) -> list:
        return self._selected

    def set_selected(self, subjects: list):
        self._selected = list(subjects)
        for cb in self.checkboxes:
            cb.setChecked(cb.text() in self._selected)
        self._update_btn_text()


class GroupSelector(QWidget):
    """Выпадающий список групп с опциональной фильтрацией по предмету."""
    changed = Signal(str)

    def __init__(self, selected: str = "", filter_subject: str = "", parent=None):
        super().__init__(parent)
        self._selected = selected
        self._filter_subject = filter_subject  # если задан — показывать только группы с этим предметом
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.btn = QPushButton(selected or "Выбрать группу ▼")
        self.btn.setStyleSheet(STYLE_BTN)
        self.btn.clicked.connect(self._toggle)
        layout.addWidget(self.btn)

        self.panel = QWidget()
        self.panel.hide()
        panel_layout = QVBoxLayout(self.panel)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(150)
        self._inner = QWidget()
        self.group_btns = []
        self.inner_l = QVBoxLayout(self._inner)
        self._rebuild()
        scroll.setWidget(self._inner)
        panel_layout.addWidget(scroll)
        layout.addWidget(self.panel)

    def set_filter_subject(self, subject: str):
        """Устанавливает фильтр — показывать только группы с этим предметом."""
        self._filter_subject = subject
        self._rebuild()

    def _rebuild(self):
        for w in self.group_btns:
            w.setParent(None)
        self.group_btns.clear()
        groups = secure_store.get_groups()
        for g in groups:
            # Фильтрация: если задан предмет — показывать только группы у которых он есть
            if self._filter_subject and self._filter_subject not in g.get("subjects", []):
                continue
            gb = QPushButton(g["name"])
            gb.clicked.connect(lambda _, name=g["name"]: self._select(name))
            self.group_btns.append(gb)
            self.inner_l.addWidget(gb)
        # Если нет подходящих групп — показываем подсказку
        if not self.group_btns and self._filter_subject:
            lbl = QPushButton("(нет групп с этим предметом)")
            lbl.setEnabled(False)
            lbl.setStyleSheet("color:#888;")
            self.group_btns.append(lbl)
            self.inner_l.addWidget(lbl)

    def _toggle(self):
        self._rebuild()
        self.panel.setVisible(not self.panel.isVisible())

    def _select(self, name: str):
        self._selected = name
        self.btn.setText(name)
        self.panel.hide()
        self.changed.emit(name)

    def get_selected(self) -> str:
        return self._selected


# ===================== СТРАНИЦЫ АДМИНИСТРАТОРА =====================

class ApiKeyPage(QWidget):
    def __init__(self, back_cb, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("API ключ OpenRouter")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        # Статус ключа
        current_key = secure_store.get_api_key()
        self.status_lbl = QLabel()
        self._update_status_label(current_key)
        self.status_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_lbl)

        layout.addWidget(QLabel("Введите новый API ключ:"))
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("sk-or-v1-...")
        self.key_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.key_input)

        btn_row = QHBoxLayout()
        save_btn = QPushButton("💾 Сохранить")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save)
        btn_row.addWidget(save_btn)

        delete_btn = QPushButton("🗑 Удалить ключ")
        delete_btn.setStyleSheet(STYLE_BTN_RED)
        delete_btn.clicked.connect(self._delete)
        btn_row.addWidget(delete_btn)
        layout.addLayout(btn_row)

        layout.addStretch()

    def _update_status_label(self, key: str):
        if key:
            masked = key[:8] + "..." + key[-4:] if len(key) > 12 else "***"
            self.status_lbl.setText(f"Текущий ключ: {masked}")
            self.status_lbl.setStyleSheet("color:#27ae60; font-size:12px;")
        else:
            self.status_lbl.setText("Ключ не установлен — ИИ работать не будет.")
            self.status_lbl.setStyleSheet("color:#e74c3c; font-size:12px;")

    def _save(self):
        key = self.key_input.text().strip()
        if not key:
            QMessageBox.warning(self, "Ошибка", "Введите API ключ.")
            return
        secure_store.set_api_key(key)
        self._update_status_label(key)
        self.key_input.clear()
        QMessageBox.information(self, "Сохранено", "API ключ успешно сохранён и зашифрован.")

    def _delete(self):
        confirm = QMessageBox.question(
            self, "Подтверждение",
            "Удалить API ключ?\nИИ-помощник перестанет работать до установки нового ключа.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            secure_store.delete("openrouter_api_key")
            self._update_status_label("")
            QMessageBox.information(self, "Удалено", "API ключ удалён.")


class TeacherDetailPage(QWidget):
    saved = Signal()

    def __init__(self, teacher_name: str, back_cb, parent=None):
        super().__init__(parent)
        self.teacher_name = teacher_name
        self._back_cb = back_cb
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        teachers = secure_store.get_teachers()
        data = teachers.get(teacher_name, {"password": "", "subjects": []})

        title = QLabel(f"Преподаватель: {teacher_name}")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Редактирование имени
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("Имя (ФИО):"))
        self.name_edit = QLineEdit(teacher_name)
        self.name_edit.setPlaceholderText("Фамилия Имя Отчество")
        name_row.addWidget(self.name_edit)
        layout.addLayout(name_row)

        layout.addWidget(QLabel("Предметы:"))
        self.subj_selector = SubjectSelector(data.get("subjects", []))
        layout.addWidget(self.subj_selector)

        # Назначение групп по предметам
        layout.addWidget(QLabel("Назначение групп по предметам:"))
        self.group_assignments = {}
        self.assign_widget = QWidget()
        self.assign_layout = QVBoxLayout(self.assign_widget)
        layout.addWidget(self.assign_widget)
        self.subj_selector.changed.connect(self._rebuild_assignments)
        self._rebuild_assignments(data.get("subjects", []))

        # Пароль
        pw_layout = QHBoxLayout()
        pw_layout.addWidget(QLabel("Пароль:"))
        self.pw_edit = QLineEdit(data.get("password", ""))
        self.pw_edit.setEchoMode(QLineEdit.Password)
        pw_layout.addWidget(self.pw_edit)
        show_btn = QPushButton("👁")
        show_btn.setFixedWidth(35)
        show_btn.clicked.connect(lambda: self.pw_edit.setEchoMode(
            QLineEdit.Normal if self.pw_edit.echoMode() == QLineEdit.Password else QLineEdit.Password))
        pw_layout.addWidget(show_btn)
        layout.addLayout(pw_layout)

        save_btn = QPushButton("Сохранить данные преподавателя")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save)
        layout.addWidget(save_btn)

        del_btn = QPushButton("🗑 Удалить преподавателя")
        del_btn.setStyleSheet(STYLE_BTN_RED)
        del_btn.clicked.connect(self._delete)
        layout.addWidget(del_btn)

    def _rebuild_assignments(self, subjects: list):
        for i in reversed(range(self.assign_layout.count())):
            w = self.assign_layout.itemAt(i).widget()
            if w:
                w.setParent(None)
        self.group_assignments = {}
        teachers = secure_store.get_teachers()
        data = teachers.get(self.teacher_name, {})
        existing = data.get("group_assignments", {})
        for subj in subjects:
            row = QHBoxLayout()
            lbl = QLabel(f"  {subj}:")
            lbl.setFixedWidth(300)
            # Передаём предмет как фильтр — показываем только группы у которых он есть
            combo = GroupSelector(existing.get(subj, ""), filter_subject=subj)
            row.addWidget(lbl)
            row.addWidget(combo)
            container = QWidget()
            container.setLayout(row)
            self.assign_layout.addWidget(container)
            self.group_assignments[subj] = combo

    def _save(self):
        teachers = secure_store.get_teachers()
        new_name = self.name_edit.text().strip()
        if not new_name:
            QMessageBox.warning(self, "Ошибка", "Имя преподавателя не может быть пустым.")
            return
        assignments = {subj: sel.get_selected() for subj, sel in self.group_assignments.items()}
        new_data = {
            "password": self.pw_edit.text(),
            "subjects": self.subj_selector.get_selected(),
            "group_assignments": assignments
        }
        # Если имя изменилось — переименовываем ключ
        if new_name != self.teacher_name:
            if new_name in teachers:
                QMessageBox.warning(self, "Ошибка",
                    f"Преподаватель «{new_name}» уже существует.")
                return
            teachers.pop(self.teacher_name, None)
            self.teacher_name = new_name
        teachers[new_name] = new_data
        secure_store.set_teachers(teachers)
        QMessageBox.information(self, "Сохранено",
            f"Данные преподавателя «{new_name}» сохранены.")
        self.saved.emit()

    def _delete(self):
        confirm = QMessageBox.question(
            self, "Удаление преподавателя",
            f"Удалить преподавателя «{self.teacher_name}»?\n\n"
            "Все его данные (предметы, назначения групп, пароль) будут удалены безвозвратно.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return
        teachers = secure_store.get_teachers()
        teachers.pop(self.teacher_name, None)
        secure_store.set_teachers(teachers)
        QMessageBox.information(self, "Удалено", f"Преподаватель «{self.teacher_name}» удалён.")
        self.saved.emit()


class TeachersPage(QWidget):
    def __init__(self, stack: QStackedWidget, back_cb, parent=None):
        super().__init__(parent)
        self.stack = stack
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("Преподаватели")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        # Фильтр по предмету
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Фильтр по предмету:"))
        self.subject_filter = QComboBox()
        self.subject_filter.addItem("Все")
        self.subject_filter.addItems(ALL_SUBJECTS)
        self.subject_filter.currentTextChanged.connect(self._refresh)
        filter_layout.addWidget(self.subject_filter)
        self.free_only = QCheckBox("Не занят")
        self.free_only.stateChanged.connect(self._refresh)
        filter_layout.addWidget(self.free_only)
        layout.addLayout(filter_layout)

        self.teacher_list = QListWidget()
        self.teacher_list.itemDoubleClicked.connect(self._open_teacher)
        layout.addWidget(self.teacher_list)

        io_row = QHBoxLayout()
        exp_btn = QPushButton("💾 Экспорт преподавателей (JSON)...")
        exp_btn.setStyleSheet(STYLE_BTN_GREEN)
        exp_btn.clicked.connect(self._export_teachers)
        io_row.addWidget(exp_btn)
        imp_btn = QPushButton("📂 Импорт преподавателей (JSON)...")
        imp_btn.setStyleSheet(STYLE_BTN)
        imp_btn.clicked.connect(self._import_teachers)
        io_row.addWidget(imp_btn)
        layout.addLayout(io_row)

        self._refresh()

    def _refresh(self):
        self.teacher_list.clear()
        teachers = secure_store.get_teachers()
        subj_filter = self.subject_filter.currentText()
        free_only = self.free_only.isChecked()

        for name, data in teachers.items():
            subjects = data.get("subjects", [])
            if subj_filter != "Все" and subj_filter not in subjects:
                continue
            if free_only and subj_filter != "Все":
                assignments = data.get("group_assignments", {})
                if assignments.get(subj_filter):
                    continue
            item = QListWidgetItem(name)
            self.teacher_list.addItem(item)

    def _export_teachers(self):
        from PySide6.QtWidgets import QFileDialog
        import json
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт преподавателей", "teachers.json",
            "JSON (*.json);;Все файлы (*)")
        if not path: return
        teachers = secure_store.get_teachers()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(teachers, f, ensure_ascii=False, indent=2)
        QMessageBox.information(self, "Экспорт выполнен",
            f"✅ Экспортировано {len(teachers)} преподавателей.")

    def _import_teachers(self):
        from PySide6.QtWidgets import QFileDialog
        import json
        path, _ = QFileDialog.getOpenFileName(
            self, "Импорт преподавателей", "", "JSON (*.json);;Все файлы (*)")
        if not path: return
        try:
            with open(path, encoding="utf-8") as f:
                new_t = json.load(f)
            if not isinstance(new_t, dict):
                QMessageBox.warning(self, "Ошибка", "Файл должен быть словарём преподавателей."); return
            cur   = secure_store.get_teachers()
            added = [n for n in new_t if n not in cur]
            dups  = [n for n in new_t if n in cur]
            diff  = {"added_students":[], "dup_students":[],
                     "added_teachers": added, "dup_teachers": dups,
                     "added_groups":[], "dup_groups":[],
                     "_new_teachers": new_t}
            dlg = ImportConfirmDialog("👨‍🏫 Импорт преподавателей", diff, self)
            dlg.confirmed.connect(lambda _: self._apply_teachers_import(added, new_t, dlg))
            dlg.cancelled.connect(dlg.close)
            dlg.show()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _apply_teachers_import(self, added, new_t, dlg):
        dlg.close()
        if not added:
            QMessageBox.information(self, "Импорт", "Нет новых преподавателей."); return
        cur = secure_store.get_teachers()
        for name in added:
            cur[name] = new_t[name]
        secure_store.set_teachers(cur)
        self._refresh()
        QMessageBox.information(self, "Импорт выполнен",
            f"✅ Добавлено новых преподавателей: {len(added)}")

    def _open_teacher(self, item):
        name = item.text()
        detail = TeacherDetailPage(name, lambda: self.stack.setCurrentWidget(self))
        detail.saved.connect(self._refresh)
        self.stack.addWidget(detail)
        self.stack.setCurrentWidget(detail)


class StudentDetailPage(QWidget):
    def __init__(self, student_data: dict, back_cb, parent=None):
        super().__init__(parent)
        self.student_data = student_data
        self._back_cb = back_cb
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel(f"Студент: {student_data.get('surname','')} {student_data.get('name','')}")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form.setSpacing(8)
        layout.addLayout(form)

        # Редактирование ФИО
        self.surname_edit = QLineEdit(student_data.get("surname", ""))
        self.surname_edit.setPlaceholderText("Фамилия")
        form.addRow("Фамилия:", self.surname_edit)

        self.name_edit = QLineEdit(student_data.get("name", ""))
        self.name_edit.setPlaceholderText("Имя / Инициалы")
        form.addRow("Имя:", self.name_edit)

        form.addRow("Группа:", QLabel())
        self.group_sel = GroupSelector(student_data.get("group", ""))
        form.addRow("", self.group_sel)

        self.group_sel.changed.connect(self._update_subjects)
        self.subjects_label = QLabel()
        layout.addWidget(QLabel("Предметы группы:"))
        layout.addWidget(self.subjects_label)
        self._update_subjects(student_data.get("group", ""))

        save_btn = QPushButton("💾 Сохранить изменения")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save)
        layout.addWidget(save_btn)

        del_btn = QPushButton("🗑 Удалить студента")
        del_btn.setStyleSheet(STYLE_BTN_RED)
        del_btn.clicked.connect(self._delete)
        layout.addWidget(del_btn)

    def _delete(self):
        name = f"{self.student_data.get('surname', '')} {self.student_data.get('name', '')}"
        confirm = QMessageBox.question(
            self, "Удаление студента",
            f"Удалить студента «{name}»?\n\n"
            "Студент будет удалён из списков. Оценки в журналах останутся в базе данных.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return
        students = secure_store.get_students()
        idx = self.student_data.get("_idx", -1)
        if 0 <= idx < len(students):
            students.pop(idx)
            secure_store.set_students(students)
        QMessageBox.information(self, "Удалено", f"Студент «{name}» удалён.")
        # Возвращаемся назад через back_cb
        self._back_cb()

    def _update_subjects(self, group_name: str):
        groups = secure_store.get_groups()
        for g in groups:
            if g["name"] == group_name:
                subjects = g.get("subjects", [])
                self.subjects_label.setText("\n".join(f"• {s}" for s in subjects) or "—")
                return
        self.subjects_label.setText("—")

    def _save(self):
        new_surname = self.surname_edit.text().strip()
        new_name    = self.name_edit.text().strip()
        new_group   = self.group_sel.get_selected()
        if not new_surname:
            QMessageBox.warning(self, "Ошибка", "Фамилия не может быть пустой.")
            return
        old_surname = self.student_data.get("surname", "")
        old_name    = self.student_data.get("name", "")
        old_group   = self.student_data.get("group", "")
        students = secure_store.get_students()
        idx = self.student_data.get("_idx", -1)
        new_data = {"name": new_name, "surname": new_surname, "group": new_group}
        if 0 <= idx < len(students):
            students[idx] = new_data
            secure_store.set_students(students)
        # Обновляем SQLite если изменились фамилия/имя/группа
        try:
            import sqlite3 as _sq
            conn = _sq.connect("vsgutu_grades.db")
            cur  = conn.cursor()
            if new_surname != old_surname or new_name != old_name:
                cur.execute(
                    "UPDATE students SET f=?, n=? WHERE f=? AND n=?",
                    (new_surname, new_name, old_surname, old_name)
                )
                cur.execute(
                    "UPDATE grades SET student_f=?, student_n=? WHERE student_f=? AND student_n=?",
                    (new_surname, new_name, old_surname, old_name)
                )
            if new_group != old_group:
                cur.execute(
                    "UPDATE students SET group_name=? WHERE f=? AND n=?",
                    (new_group, new_surname, new_name)
                )
            conn.commit(); conn.close()
        except Exception as e:
            pass  # SQLite может отсутствовать если журналов ещё нет
        # Обновляем локальный кэш
        self.student_data.update(new_data)
        QMessageBox.information(self, "Сохранено",
            f"Данные студента «{new_surname} {new_name}» сохранены.")


class StudentsPage(QWidget):
    def __init__(self, stack: QStackedWidget, back_cb, parent=None):
        super().__init__(parent)
        self.stack = stack
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("Студенты")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Фильтр по группе:"))
        self.group_filter = QComboBox()
        self.group_filter.addItem("Все")
        self.group_filter.addItem("Без группы")
        groups = secure_store.get_groups()
        for g in groups:
            self.group_filter.addItem(g["name"])
        self.group_filter.currentTextChanged.connect(self._refresh)
        filter_layout.addWidget(self.group_filter)
        layout.addLayout(filter_layout)

        self.student_list = QListWidget()
        self.student_list.itemDoubleClicked.connect(self._open_student)
        layout.addWidget(self.student_list)

        io_row = QHBoxLayout()
        exp_btn = QPushButton("💾 Экспорт студентов (JSON)...")
        exp_btn.setStyleSheet(STYLE_BTN_GREEN)
        exp_btn.clicked.connect(self._export_students)
        io_row.addWidget(exp_btn)
        imp_btn = QPushButton("📂 Импорт студентов (JSON)...")
        imp_btn.setStyleSheet(STYLE_BTN)
        imp_btn.clicked.connect(self._import_students)
        io_row.addWidget(imp_btn)
        layout.addLayout(io_row)

        self._refresh()

    def _refresh(self):
        self.student_list.clear()
        students = secure_store.get_students()
        group_filter = self.group_filter.currentText()
        for i, s in enumerate(students):
            group = s.get("group", "")
            if group_filter == "Без группы" and group:
                continue
            if group_filter not in ("Все", "Без группы") and group != group_filter:
                continue
            label = f"{s.get('surname', '')} {s.get('name', '')}  [{group or '—'}]"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, i)
            self.student_list.addItem(item)

    def _export_students(self):
        from PySide6.QtWidgets import QFileDialog
        import json
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт студентов", "students.json",
            "JSON (*.json);;Все файлы (*)")
        if not path: return
        students = secure_store.get_students()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(students, f, ensure_ascii=False, indent=2)
        QMessageBox.information(self, "Экспорт выполнен",
            f"✅ Экспортировано {len(students)} студентов.")

    def _import_students(self):
        from PySide6.QtWidgets import QFileDialog
        import json
        path, _ = QFileDialog.getOpenFileName(
            self, "Импорт студентов", "", "JSON (*.json);;Все файлы (*)")
        if not path: return
        try:
            with open(path, encoding="utf-8") as f:
                new_list = json.load(f)
            if not isinstance(new_list, list):
                QMessageBox.warning(self, "Ошибка", "Файл должен быть массивом студентов."); return

            def _norm(s):
                f_ = s.get("surname","").strip().lower()
                n_ = s.get("name","").strip().lower()
                return tuple(sorted([f_, n_]))

            cur   = secure_store.get_students()
            cur_k = {_norm(s) for s in cur}
            added = [s for s in new_list if _norm(s) not in cur_k]
            dups  = [s for s in new_list if _norm(s) in cur_k]

            diff = {"added_students": added, "dup_students": dups,
                    "added_teachers": [], "dup_teachers": [],
                    "added_groups": [], "dup_groups": [],
                    "_new_students": new_list}
            dlg = ImportConfirmDialog("🎓 Импорт студентов", diff, self)
            dlg.confirmed.connect(lambda _: self._apply_students_import(added, dlg))
            dlg.cancelled.connect(dlg.close)
            dlg.show()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _apply_students_import(self, added, dlg):
        dlg.close()
        if not added:
            QMessageBox.information(self, "Импорт", "Нет новых студентов."); return
        cur = secure_store.get_students()
        cur.extend(added)
        secure_store.set_students(cur)
        self._refresh()
        QMessageBox.information(self, "Импорт выполнен",
            f"✅ Добавлено новых студентов: {len(added)}")

    def _open_student(self, item):
        idx = item.data(Qt.UserRole)
        students = secure_store.get_students()
        data = dict(students[idx])
        data["_idx"] = idx
        detail = StudentDetailPage(data, lambda: self.stack.setCurrentWidget(self))
        self.stack.addWidget(detail)
        self.stack.setCurrentWidget(detail)


class SubjectSelectorWithSearch(QWidget):
    """Список предметов с поиском и чекбоксами."""
    changed = Signal(list)

    def __init__(self, selected: list = None, parent=None):
        super().__init__(parent)
        self._selected = list(selected or [])
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        # Поле поиска
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍 Поиск предмета...")
        self.search.textChanged.connect(self._filter)
        layout.addWidget(self.search)

        # Счётчик выбранных
        self.count_lbl = QLabel()
        self.count_lbl.setStyleSheet("color:#7eb8f7; font-size:11px;")
        layout.addWidget(self.count_lbl)

        # Список с прокруткой
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(220)
        self._inner = QWidget()
        self._inner_l = QVBoxLayout(self._inner)
        self._inner_l.setSpacing(2)
        self.checkboxes = []
        for subj in get_all_subjects():
            cb = QCheckBox(subj)
            cb.setChecked(subj in self._selected)
            cb.stateChanged.connect(self._update)
            self.checkboxes.append(cb)
            self._inner_l.addWidget(cb)
        scroll.setWidget(self._inner)
        layout.addWidget(scroll)
        self._update_count()

    def _filter(self, text: str):
        text = text.lower()
        for cb in self.checkboxes:
            cb.setVisible(text == "" or text in cb.text().lower())

    def _update(self):
        self._selected = [cb.text() for cb in self.checkboxes if cb.isChecked()]
        self._update_count()
        self.changed.emit(self._selected)

    def _update_count(self):
        self.count_lbl.setText(f"Выбрано предметов: {len(self._selected)}")

    def get_selected(self) -> list:
        return self._selected

    def set_selected(self, subjects: list):
        self._selected = list(subjects)
        for cb in self.checkboxes:
            cb.setChecked(cb.text() in self._selected)
        self._update_count()


class GroupDetailPage(QWidget):
    saved = Signal()

    def __init__(self, group_data: dict, back_cb, parent=None):
        super().__init__(parent)
        self.group_name = group_data.get("name", "")
        self._back_cb = back_cb
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel(f"Группа: {self.group_name}")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Переименование группы
        rename_row = QHBoxLayout()
        rename_row.addWidget(QLabel("Название группы:"))
        self.name_edit = QLineEdit(self.group_name)
        self.name_edit.setPlaceholderText("Например: к74/1")
        rename_row.addWidget(self.name_edit)
        layout.addLayout(rename_row)

        # Студенты группы
        students = secure_store.get_students()
        group_students = [s for s in students if s.get("group") == self.group_name]
        stud_lbl = QLabel(f"Студентов в группе: {len(group_students)}")
        stud_lbl.setStyleSheet("color:#7eb8f7; font-size:12px;")
        layout.addWidget(stud_lbl)

        stud_scroll = QScrollArea()
        stud_scroll.setFixedHeight(100)
        stud_scroll.setWidgetResizable(True)
        stud_inner = QWidget()
        stud_inner_l = QVBoxLayout(stud_inner)
        if group_students:
            for s in group_students:
                stud_inner_l.addWidget(QLabel(f"• {s.get('surname', '')} {s.get('name', '')}"))
        else:
            stud_inner_l.addWidget(QLabel("  (нет студентов)"))
        stud_scroll.setWidget(stud_inner)
        layout.addWidget(stud_scroll)

        # Предметы — с поиском и редактированием
        layout.addWidget(QLabel("Предметы группы (можно редактировать):"))
        self.subj_selector = SubjectSelectorWithSearch(group_data.get("subjects", []))
        layout.addWidget(self.subj_selector)

        save_btn = QPushButton("💾 Сохранить изменения группы")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save)
        layout.addWidget(save_btn)

        del_btn = QPushButton("🗑 Удалить группу")
        del_btn.setStyleSheet(STYLE_BTN_RED)
        del_btn.clicked.connect(self._delete)
        layout.addWidget(del_btn)

    def _delete(self):
        confirm = QMessageBox.question(
            self, "Удаление группы",
            f"Удалить группу «{self.group_name}»?\n\n"
            "Группа будет удалена из списков.\n"
            "Студенты этой группы останутся в базе, но без привязки к группе.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return
        # Удаляем группу
        groups = secure_store.get_groups()
        groups = [g for g in groups if g["name"] != self.group_name]
        secure_store.set_groups(groups)
        # Убираем привязку студентов к этой группе
        students = secure_store.get_students()
        for s in students:
            if s.get("group") == self.group_name:
                s["group"] = ""
        secure_store.set_students(students)
        QMessageBox.information(self, "Удалено", f"Группа «{self.group_name}» удалена.")
        self._back_cb()

    def _save(self):
        new_name = self.name_edit.text().strip()
        if not new_name:
            QMessageBox.warning(self, "Ошибка", "Название группы не может быть пустым.")
            return
        groups = secure_store.get_groups()
        # Проверка дублирования при переименовании
        if new_name != self.group_name:
            if any(g["name"] == new_name for g in groups):
                QMessageBox.warning(self, "Ошибка",
                    f"Группа «{new_name}» уже существует.")
                return
        for g in groups:
            if g["name"] == self.group_name:
                g["name"]     = new_name
                g["subjects"] = self.subj_selector.get_selected()
                break
        secure_store.set_groups(groups)
        # Обновляем студентов при переименовании
        if new_name != self.group_name:
            students = secure_store.get_students()
            for s in students:
                if s.get("group") == self.group_name:
                    s["group"] = new_name
            secure_store.set_students(students)
            # Обновляем SQLite
            try:
                import sqlite3 as _sq
                conn = _sq.connect("vsgutu_grades.db")
                cur  = conn.cursor()
                cur.execute("UPDATE students SET group_name=? WHERE group_name=?",
                            (new_name, self.group_name))
                cur.execute("UPDATE lessons SET group_name=? WHERE group_name=?",
                            (new_name, self.group_name))
                conn.commit(); conn.close()
            except Exception:
                pass
            self.group_name = new_name
        QMessageBox.information(self, "Сохранено",
            f"Группа «{new_name}» обновлена.")
        self.saved.emit()


class GroupsPage(QWidget):
    def __init__(self, stack: QStackedWidget, back_cb, parent=None):
        super().__init__(parent)
        self.stack = stack
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("Группы")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        layout.addWidget(QLabel("Фильтр по предметам:"))
        self.subject_checkboxes = {}
        scroll = QScrollArea()
        scroll.setFixedHeight(120)
        scroll.setWidgetResizable(True)
        inner = QWidget()
        inner_l = QVBoxLayout(inner)
        for subj in ALL_SUBJECTS:
            cb = QCheckBox(subj)
            cb.stateChanged.connect(self._refresh)
            self.subject_checkboxes[subj] = cb
            inner_l.addWidget(cb)
        scroll.setWidget(inner)
        layout.addWidget(scroll)

        self.group_list = QListWidget()
        self.group_list.itemDoubleClicked.connect(self._open_group)
        layout.addWidget(self.group_list)
        self._refresh()

    def _refresh(self):
        self.group_list.clear()
        selected_subjects = [subj for subj, cb in self.subject_checkboxes.items() if cb.isChecked()]
        groups = secure_store.get_groups()
        for g in groups:
            if selected_subjects:
                g_subjects = set(g.get("subjects", []))
                if not any(s in g_subjects for s in selected_subjects):
                    continue
            self.group_list.addItem(QListWidgetItem(g["name"]))

    def _open_group(self, item):
        name = item.text()
        groups = secure_store.get_groups()
        data = next((g for g in groups if g["name"] == name), {"name": name, "subjects": []})
        detail = GroupDetailPage(data, lambda: self.stack.setCurrentWidget(self))
        self.stack.addWidget(detail)
        self.stack.setCurrentWidget(detail)


class AddDataPage(QWidget):
    """Страница добавления: Преподаватель / Студент / Группа"""

    def __init__(self, back_cb, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("Добавить данные")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(title)

        tabs = QHBoxLayout()
        btn_t = QPushButton("Добавить преподавателя")
        btn_s = QPushButton("Добавить студента")
        btn_g = QPushButton("Добавить группу")
        for b in [btn_t, btn_s, btn_g]:
            b.setStyleSheet(STYLE_BTN)
            tabs.addWidget(b)
        layout.addLayout(tabs)

        self.inner_stack = QStackedWidget()
        layout.addWidget(self.inner_stack)

        btn_t.clicked.connect(lambda: self.inner_stack.setCurrentIndex(0))
        btn_s.clicked.connect(lambda: self.inner_stack.setCurrentIndex(1))
        btn_g.clicked.connect(lambda: self.inner_stack.setCurrentIndex(2))

        self.inner_stack.addWidget(self._build_teacher_form())
        self.inner_stack.addWidget(self._build_student_form())
        self.inner_stack.addWidget(self._build_group_form())

    def _build_teacher_form(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        form = QFormLayout()
        self.t_name = QLineEdit()
        self.t_surname = QLineEdit()
        self.t_password = QLineEdit()
        self.t_password.setEchoMode(QLineEdit.Password)
        form.addRow("Имя:", self.t_name)
        form.addRow("Фамилия:", self.t_surname)
        form.addRow("Пароль:", self.t_password)
        layout.addLayout(form)
        layout.addWidget(QLabel("Предметы:"))
        self.t_subj_sel = SubjectSelector()
        layout.addWidget(self.t_subj_sel)
        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save_teacher)
        layout.addWidget(save_btn)
        return w

    def _build_student_form(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        form = QFormLayout()
        self.s_name = QLineEdit()
        self.s_surname = QLineEdit()
        form.addRow("Имя:", self.s_name)
        form.addRow("Фамилия:", self.s_surname)
        layout.addLayout(form)
        layout.addWidget(QLabel("Группа:"))
        self.s_group_sel = GroupSelector()
        layout.addWidget(self.s_group_sel)
        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save_student)
        layout.addWidget(save_btn)

        sep = QLabel("── или импортировать из Excel ──")
        sep.setAlignment(Qt.AlignCenter)
        sep.setStyleSheet("color:#6080a0; font-size:11px; margin-top:10px;")
        layout.addWidget(sep)

        imp_btn = QPushButton("📥 Импортировать студентов из Excel...")
        imp_btn.setStyleSheet(STYLE_BTN)
        imp_btn.clicked.connect(self._import_students_excel)
        layout.addWidget(imp_btn)
        return w

    def _import_students_excel(self):
        """Импорт студентов из Excel-журнала колледжа. Группу выбираем вручную."""
        from PySide6.QtWidgets import QFileDialog, QInputDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть Excel с журналом", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if not path:
            return
        # Выбор группы вручную
        groups = [g["name"] for g in secure_store.get_groups()]
        if not groups:
            QMessageBox.warning(self, "Нет групп", "Сначала создайте группы в системе.")
            return
        group, ok = QInputDialog.getItem(
            self, "Выберите группу",
            "В какую группу добавить студентов?", groups, 0, False
        )
        if not ok:
            return
        try:
            from openpyxl import load_workbook
            wb   = load_workbook(path)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                QMessageBox.warning(self, "Ошибка", "Файл пустой.")
                return

            row0 = rows[0]
            row1 = rows[1]
            is_college = (
                row1 and any(str(v).strip() in ("1 час","2 час") for v in row1 if v)
            )

            existing = secure_store.get_students()
            existing_keys = {
                (s.get("surname","").lower(), s.get("group",""))
                for s in existing
            }
            added = 0

            if is_college:
                # Строки с 3-й: col 1 = ФИО
                data_rows = rows[2:]
                fio_col   = 1
            else:
                # Строки с 2-й: col 0 = фамилия, col 1 = имя
                data_rows = rows[1:]
                fio_col   = None

            for row in data_rows:
                if not row:
                    continue
                if is_college:
                    if not row[fio_col]:
                        continue
                    fio   = str(row[fio_col]).strip()
                    parts = fio.split()
                    if not parts:
                        continue
                    surname  = parts[0]
                    initials = " ".join(parts[1:]) if len(parts) > 1 else ""
                    name     = initials
                else:
                    if not row[0]:
                        continue
                    surname = str(row[0]).strip()
                    name    = str(row[1]).strip() if len(row) > 1 and row[1] else ""

                if not surname:
                    continue
                key = (surname.lower(), group)
                if key not in existing_keys:
                    existing.append({"name": name, "surname": surname, "group": group})
                    existing_keys.add(key)
                    added += 1

            secure_store.set_students(existing)
            QMessageBox.information(
                self, "Импорт завершён",
                f"Группа: {group}\nДобавлено новых студентов: {added}"
            )
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "Установите: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка импорта", str(e))

    def _build_group_form(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        form = QFormLayout()
        self.g_name = QLineEdit()
        form.addRow("Название группы:", self.g_name)
        layout.addLayout(form)
        layout.addWidget(QLabel("Предметы:"))
        self.g_subj_sel = SubjectSelector()
        layout.addWidget(self.g_subj_sel)
        save_btn = QPushButton("Сохранить")
        save_btn.setStyleSheet(STYLE_BTN_GREEN)
        save_btn.clicked.connect(self._save_group)
        layout.addWidget(save_btn)
        return w

    def _save_teacher(self):
        name = self.t_name.text().strip()
        surname = self.t_surname.text().strip()
        if not name or not surname:
            QMessageBox.warning(self, "Ошибка", "Введите имя и фамилию.")
            return
        full_name = f"{surname} {name}"
        teachers = secure_store.get_teachers()
        teachers[full_name] = {
            "password": self.t_password.text(),
            "subjects": self.t_subj_sel.get_selected(),
            "group_assignments": {}
        }
        secure_store.set_teachers(teachers)
        QMessageBox.information(self, "Добавлено", f"Преподаватель {full_name} добавлен.")
        self.t_name.clear(); self.t_surname.clear(); self.t_password.clear()

    def _save_student(self):
        name = self.s_name.text().strip()
        surname = self.s_surname.text().strip()
        if not name or not surname:
            QMessageBox.warning(self, "Ошибка", "Введите имя и фамилию.")
            return
        students = secure_store.get_students()
        students.append({"name": name, "surname": surname, "group": self.s_group_sel.get_selected()})
        secure_store.set_students(students)
        QMessageBox.information(self, "Добавлено", f"Студент {surname} {name} добавлен.")
        self.s_name.clear(); self.s_surname.clear()

    def _save_group(self):
        gname = self.g_name.text().strip()
        if not gname:
            QMessageBox.warning(self, "Ошибка", "Введите название группы.")
            return
        groups = secure_store.get_groups()
        if any(g["name"] == gname for g in groups):
            QMessageBox.warning(self, "Ошибка", "Группа с таким именем уже существует.")
            return
        groups.append({"name": gname, "subjects": self.g_subj_sel.get_selected()})
        secure_store.set_groups(groups)
        QMessageBox.information(self, "Добавлено", f"Группа {gname} добавлена.")
        self.g_name.clear()


class SubjectsEditorPage(QWidget):
    """
    Редактор глобального списка ALL_SUBJECTS.
    Позволяет добавлять, переименовывать и удалять предметы.
    Изменения сохраняются в secure_store и применяются ко всем выборщикам.
    """
    saved = Signal()

    def __init__(self, back_cb, parent=None):
        super().__init__(parent)
        self._back_cb = back_cb
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("📚 Редактор учебных предметов")
        title.setStyleSheet("font-size:16px; font-weight:bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)


        # ── Панель добавления нового предмета
        add_row = QHBoxLayout()
        self.new_subj_edit = QLineEdit()
        self.new_subj_edit.setPlaceholderText("Название нового предмета...")
        self.new_subj_edit.returnPressed.connect(self._add_subject)
        add_row.addWidget(self.new_subj_edit)
        add_btn = QPushButton("➕ Добавить")
        add_btn.setStyleSheet(STYLE_BTN_GREEN)
        add_btn.setFixedWidth(120)
        add_btn.clicked.connect(self._add_subject)
        add_row.addWidget(add_btn)
        layout.addLayout(add_row)

        # ── Поиск по списку
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Поиск в списке...")
        self.search_edit.textChanged.connect(self._filter)
        layout.addWidget(self.search_edit)

        lbl_count = QHBoxLayout()
        self.count_lbl = QLabel()
        self.count_lbl.setStyleSheet("color:#7eb8f7; font-size:11px;")
        lbl_count.addWidget(self.count_lbl)
        lbl_count.addStretch()
        layout.addLayout(lbl_count)

        # ── Список предметов
        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet(
            "QListWidget{background:#1a2035;border:1px solid #3d4460;border-radius:6px;}"
            "QListWidget::item{padding:6px 8px;border-bottom:1px solid #2d3450;}"
            "QListWidget::item:selected{background:#2d4a7a;}"
        )
        self.list_widget.setSelectionMode(QListWidget.SingleSelection)
        layout.addWidget(self.list_widget, 1)

        # ── Кнопки действий над выбранным предметом
        action_row = QHBoxLayout()
        rename_btn = QPushButton("✏ Переименовать")
        rename_btn.setStyleSheet(STYLE_BTN)
        rename_btn.clicked.connect(self._rename_subject)
        action_row.addWidget(rename_btn)

        del_btn = QPushButton("🗑 Удалить выбранный")
        del_btn.setStyleSheet(STYLE_BTN_RED)
        del_btn.clicked.connect(self._delete_subject)
        action_row.addWidget(del_btn)
        layout.addLayout(action_row)

        # ── Импорт / Экспорт ────────────────────────────────
        sep = QLabel("── Импорт / Экспорт ──")
        sep.setAlignment(Qt.AlignCenter)
        sep.setStyleSheet("color:#6080a0;font-size:11px;margin-top:4px;")
        layout.addWidget(sep)
        io_row = QHBoxLayout()
        exp_btn = QPushButton("💾 Экспорт в JSON...")
        exp_btn.setStyleSheet(STYLE_BTN_GREEN)
        exp_btn.clicked.connect(self._export_subjects)
        io_row.addWidget(exp_btn)
        imp_btn = QPushButton("📂 Импорт из JSON...")
        imp_btn.setStyleSheet(STYLE_BTN)
        imp_btn.clicked.connect(self._import_subjects)
        io_row.addWidget(imp_btn)
        layout.addLayout(io_row)

        self._load_subjects()

    def _get_subjects(self):
        """Возвращает актуальный список предметов из subjects.json."""
        return load_subjects()

    def _save_subjects(self, subjects: list):
        """Сохраняет список предметов в subjects.json."""
        save_subjects(subjects)

    def _export_subjects(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Экспорт предметов", "subjects.json",
            "JSON (*.json);;Все файлы (*)")
        if not path: return
        ok, msg = secure_store.export_subjects_json(path)
        if ok:
            QMessageBox.information(self, "Экспорт выполнен",
                f"✅ {msg}\nФайл можно редактировать в любом текстовом редакторе.")
        else:
            QMessageBox.critical(self, "Ошибка", msg)

    def _import_subjects(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "Импорт предметов", "", "JSON (*.json);;Все файлы (*)")
        if not path: return
        ok, msg, diff = secure_store.import_subjects_json(path)
        if not ok:
            QMessageBox.critical(self, "Ошибка", msg); return
        dlg = ImportConfirmDialog("📚 Импорт предметов", diff, self)
        dlg.confirmed.connect(lambda _: self._apply_subjects_import(diff, dlg))
        dlg.cancelled.connect(dlg.close)
        dlg.show()

    def _apply_subjects_import(self, diff, dlg):
        dlg.close()
        from subjects import save_subjects, load_subjects
        cur    = load_subjects()
        merged = sorted(list(set(cur) | set(diff["all"])))
        save_subjects(merged)
        self._load_subjects()
        QMessageBox.information(self, "Импорт выполнен",
            f"✅ Добавлено: {len(diff['added'])}  |  Уже было: {len(diff['already'])}")

    def _load_subjects(self):
        self._subjects = load_subjects()
        self._filter(self.search_edit.text())

    def _filter(self, query: str):
        self.list_widget.clear()
        q = query.strip().lower()
        shown = [s for s in self._subjects if q in s.lower()] if q else self._subjects
        for s in shown:
            self.list_widget.addItem(s)
        self.count_lbl.setText(
            f"Предметов: {len(self._subjects)} | показано: {len(shown)}"
        )

    def _add_subject(self):
        name = self.new_subj_edit.text().strip()
        if not name:
            return
        if not add_subject(name):
            QMessageBox.warning(self, "Дубликат", f"Предмет «{name}» уже существует.")
            return
        self._subjects = load_subjects()
        self.new_subj_edit.clear()
        self._filter(self.search_edit.text())
        QMessageBox.information(self, "Добавлено", f"Предмет «{name}» добавлен в subjects.json.")

    def _rename_subject(self):
        item = self.list_widget.currentItem()
        if not item:
            QMessageBox.warning(self, "Не выбрано", "Выберите предмет для переименования.")
            return
        old_name = item.text()
        new_name, ok = QInputDialog.getText(
            self, "Переименовать предмет",
            "Новое название для: " + old_name,
            text=old_name
        )
        if not ok or not new_name.strip():
            return
        new_name = new_name.strip()
        if new_name == old_name:
            return
        current = load_subjects()
        if new_name in current:
            QMessageBox.warning(self, "Дубликат", f"Предмет «{new_name}» уже существует.")
            return
        # 1. Переименовываем в subjects.json
        rename_subject(old_name, new_name)
        # 2. Обновляем в памяти
        self._subjects = load_subjects()
        # 3. Каскадно обновляем группы
        groups = secure_store.get_groups()
        for g in groups:
            subjs = g.get("subjects", [])
            if old_name in subjs:
                g["subjects"] = [new_name if s == old_name else s for s in subjs]
        secure_store.set_groups(groups)
        # 4. Каскадно обновляем преподавателей
        teachers = secure_store.get_teachers()
        for t_data in teachers.values():
            subjs = t_data.get("subjects", [])
            if old_name in subjs:
                t_data["subjects"] = [new_name if s == old_name else s for s in subjs]
            assignments = t_data.get("group_assignments", {})
            if old_name in assignments:
                assignments[new_name] = assignments.pop(old_name)
        secure_store.set_teachers(teachers)
        # 5. Обновляем SQLite — таблица lessons хранит subject текстом
        try:
            import sqlite3 as _sq
            conn = _sq.connect("vsgutu_grades.db")
            cur  = conn.cursor()
            cur.execute(
                "UPDATE lessons SET subject=? WHERE subject=?",
                (new_name, old_name)
            )
            rows_updated = cur.rowcount
            conn.commit()
            conn.close()
        except Exception as _e:
            rows_updated = 0
        # 6. Обновляем список на экране
        self._filter(self.search_edit.text())
        QMessageBox.information(self, "Переименовано",
            f"«{old_name}» → «{new_name}»\n"
            f"Обновлено в subjects.json, у групп, преподавателей "
            f"и в базе данных ({rows_updated} занятий).")

    def _delete_subject(self):
        item = self.list_widget.currentItem()
        if not item:
            QMessageBox.warning(self, "Не выбрано", "Выберите предмет для удаления.")
            return
        name = item.text()
        groups   = secure_store.get_groups()
        teachers = secure_store.get_teachers()
        used_in_groups   = [g["name"] for g in groups if name in g.get("subjects",[])]
        used_in_teachers = [t for t,d in teachers.items() if name in d.get("subjects",[])]
        warn = ""
        if used_in_groups:
            warn += f"\nГруппы: {', '.join(used_in_groups)}"
        if used_in_teachers:
            warn += f"\nПреподаватели: {', '.join(used_in_teachers)}"
        msg = f"Удалить предмет «{name}» из общего списка?"
        if warn:
            msg += f"\n\nПредмет используется:{warn}\n\nОн будет удалён из всех списков."
        confirm = QMessageBox.question(self, "Удаление предмета", msg,
            QMessageBox.Yes | QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        # 1. Удаляем из subjects.json
        delete_subject(name)
        # 2. Обновляем в памяти
        self._subjects = load_subjects()
        # 3. Каскадно удаляем из групп и преподавателей
        for g in groups:
            g["subjects"] = [s for s in g.get("subjects",[]) if s != name]
        secure_store.set_groups(groups)
        for t_data in teachers.values():
            t_data["subjects"] = [s for s in t_data.get("subjects",[]) if s != name]
        secure_store.set_teachers(teachers)
        self._filter(self.search_edit.text())
        QMessageBox.information(self, "Удалено", f"Предмет «{name}» удалён из subjects.json.")


class AdminLoginPage(QWidget):
    """Страница входа в аккаунт администратора."""
    login_success = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)

        title = QLabel("⚙ Администратор")
        title.setStyleSheet("font-size:20px; font-weight:bold;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        form = QFormLayout()
        self.login_input = QLineEdit()
        self.login_input.setPlaceholderText("Логин")
        self.pass_input = QLineEdit()
        self.pass_input.setEchoMode(QLineEdit.Password)
        self.pass_input.setPlaceholderText("Пароль")
        form.addRow("Логин:", self.login_input)
        form.addRow("Пароль:", self.pass_input)
        layout.addLayout(form)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: red;")
        layout.addWidget(self.error_label)

        enter_btn = QPushButton("Войти")
        enter_btn.setStyleSheet(STYLE_BTN)
        enter_btn.clicked.connect(self._try_login)
        self.pass_input.returnPressed.connect(self._try_login)
        layout.addWidget(enter_btn)

    def _try_login(self):
        login = self.login_input.text().strip()
        password = self.pass_input.text()
        admin_password = secure_store.get_admin_password()
        if login == "admin_vsgutu" and password == admin_password:
            self.error_label.setText("")
            self.login_success.emit()
        else:
            self.error_label.setText("Неверный логин или пароль.")


class ImportConfirmDialog(QWidget):
    """
    Диалог подтверждения импорта — показывает что будет добавлено/пропущено.
    Используется для ZIP, JSON предметов, студентов и т.д.
    """
    confirmed = Signal(str)  # 'merge' или 'replace'
    cancelled = Signal()

    def __init__(self, title: str, diff: dict, parent=None):
        super().__init__(parent, Qt.Dialog | Qt.WindowTitleHint | Qt.WindowCloseButtonHint)
        self.setWindowTitle(title)
        self.setMinimumWidth(520)
        self.setMinimumHeight(400)
        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        lbl = QLabel(title)
        lbl.setStyleSheet("font-size:14px;font-weight:bold;")
        lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        il = QVBoxLayout(inner)
        il.setSpacing(4)

        def section(heading, items, color):
            if not items:
                return
            h = QLabel(heading)
            h.setStyleSheet(f"font-weight:bold;color:{color};font-size:12px;")
            il.addWidget(h)
            for item in items[:30]:
                if isinstance(item, dict):
                    txt = f"  • {item.get('surname','')} {item.get('name','')}  [{item.get('group','')}]"
                else:
                    txt = f"  • {item}"
                il.addWidget(QLabel(txt))
            if len(items) > 30:
                il.addWidget(QLabel(f"  ... и ещё {len(items)-30}"))

        section(f"✅ Будет добавлено студентов: {len(diff.get('added_students',[]))}",
                diff.get("added_students", []), "#7dd87d")
        section(f"⚠ Уже существуют (пропустим): {len(diff.get('dup_students',[]))}",
                diff.get("dup_students", []), "#e8a87c")
        section(f"✅ Будет добавлено преподавателей: {len(diff.get('added_teachers',[]))}",
                diff.get("added_teachers", []), "#7dd87d")
        section(f"⚠ Преподаватели-дубли: {len(diff.get('dup_teachers',[]))}",
                diff.get("dup_teachers", []), "#e8a87c")
        section(f"✅ Новых групп: {len(diff.get('added_groups',[]))}",
                diff.get("added_groups", []), "#7dd87d")
        section(f"✅ Новых предметов: {len(diff.get('added_subjects',[]))}",
                diff.get("added_subjects", []), "#7dd87d")
        section(f"⚠ Предметы-дубли: {len(diff.get('dup_subjects',[]))}",
                diff.get("dup_subjects", []), "#e8a87c")

        if "db_lessons" in diff:
            il.addWidget(QLabel(
                f"📊 База оценок: {diff['db_lessons']} занятий, "
                f"{diff.get('db_grades',0)} оценок"
            ))

        if not any(diff.get(k) for k in
                   ["added_students","added_teachers","added_groups",
                    "added_subjects","db_lessons"]):
            il.addWidget(QLabel("Нет новых данных для добавления."))

        scroll.setWidget(inner)
        layout.addWidget(scroll, 1)

        btn_row = QHBoxLayout()

        merge_btn = QPushButton("✅ Добавить новые (рекомендуется)")
        merge_btn.setStyleSheet(STYLE_BTN_GREEN)
        merge_btn.clicked.connect(lambda: self.confirmed.emit("merge"))
        btn_row.addWidget(merge_btn)

        replace_btn = QPushButton("♻ Заменить всё")
        replace_btn.setStyleSheet(STYLE_BTN_RED)
        replace_btn.clicked.connect(lambda: self.confirmed.emit("replace"))
        btn_row.addWidget(replace_btn)

        cancel_btn = QPushButton("Отмена")
        cancel_btn.setStyleSheet(STYLE_BTN)
        cancel_btn.clicked.connect(self.cancelled.emit)
        btn_row.addWidget(cancel_btn)

        layout.addLayout(btn_row)



class TransferPage(QWidget):
    """Перенос данных — полный ZIP-бэкап или только настройки."""

    def __init__(self, back_cb, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 16, 24, 16)
        layout.setSpacing(10)
        layout.addWidget(make_back_btn(back_cb))

        title = QLabel("📦 Перенос данных")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size:16px;font-weight:bold;")
        layout.addWidget(title)

        info = QLabel(
            "Полный бэкап содержит всё: преподавателей, студентов, группы,\n"
            "оценки, посещаемость и предметы — в одном файле."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background:#1e2d1e;border:1px solid #2d5a2d;"
            "border-radius:6px;padding:10px;color:#7dd87d;font-size:12px;"
        )
        layout.addWidget(info)

        layout.addWidget(self._section(
            "⬆ Полный экспорт (все данные)",
            "Упаковывает в .zip: настройки, оценки, предметы.",
            "💾 Экспортировать всё...", STYLE_BTN_GREEN, self._export_full
        ))
        layout.addWidget(self._section(
            "⬇ Полный импорт (все данные)",
            "Восстанавливает из .zip. Показывает что будет добавлено.",
            "📂 Импортировать из .zip...", STYLE_BTN, self._import_full, warn=True
        ))

        sep = QLabel("── Только настройки (без оценок) ──")
        sep.setAlignment(Qt.AlignCenter)
        sep.setStyleSheet("color:#6080a0;font-size:11px;margin-top:4px;")
        layout.addWidget(sep)

        row = QHBoxLayout()
        for lbl, fn in [("💾 Экспорт настроек...", self._export_settings),
                         ("📂 Импорт настроек...", self._import_settings)]:
            b = QPushButton(lbl); b.setStyleSheet(STYLE_BTN)
            b.clicked.connect(fn); row.addWidget(b)
        layout.addLayout(row)
        layout.addStretch()

    def _section(self, title, desc, btn_text, btn_style, cb, warn=False):
        f = QFrame()
        f.setStyleSheet("QFrame{background:#252b40;border:1px solid #3d4460;border-radius:8px;}")
        fl = QVBoxLayout(f); fl.setContentsMargins(16,12,16,12); fl.setSpacing(6)
        t = QLabel(title); t.setStyleSheet("font-weight:bold;color:#c5d0f0;"); fl.addWidget(t)
        d = QLabel(desc); d.setStyleSheet(f"color:{'#e8a87c' if warn else '#a0aac0'};font-size:12px;")
        d.setWordWrap(True); fl.addWidget(d)
        b = QPushButton(btn_text); b.setStyleSheet(btn_style)
        b.setMinimumHeight(38); b.clicked.connect(cb); fl.addWidget(b)
        return f

    def _export_full(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить полный бэкап", "vsgutu_backup.zip",
            "ZIP-архив (*.zip);;Все файлы (*)")
        if not path: return
        ok, msg = secure_store.export_full(path)
        if ok:
            QMessageBox.information(self, "Экспорт выполнен", f"✅ {msg}")
        else:
            QMessageBox.critical(self, "Ошибка", msg)

    def _import_full(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть бэкап", "", "ZIP-архив (*.zip);;Все файлы (*)")
        if not path: return
        ok, msg, diff = secure_store.analyze_import(path)
        if not ok:
            QMessageBox.critical(self, "Ошибка", msg); return
        dlg = ImportConfirmDialog("📦 Подтверждение полного импорта", diff, self)
        dlg.confirmed.connect(lambda mode: self._do_import(diff, mode, dlg))
        dlg.cancelled.connect(dlg.close)
        dlg.show()

    def _do_import(self, diff, mode, dlg):
        dlg.close()
        ok, msg = secure_store.apply_import(diff, mode)
        if ok:
            QMessageBox.information(self, "Импорт выполнен",
                f"✅ {msg}\n\nПерезапустите программу чтобы данные применились.")
        else:
            QMessageBox.critical(self, "Ошибка", f"❌ {msg}")

    def _export_settings(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить настройки", "vsgutu_settings.enc",
            "Зашифрованные данные (*.enc);;Все файлы (*)")
        if not path: return
        ok, msg = secure_store.export_portable(path)
        if ok:
            QMessageBox.information(self, "Экспорт настроек", f"✅ {msg}")
        else:
            QMessageBox.critical(self, "Ошибка", msg)

    def _import_settings(self):
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть настройки", "",
            "Зашифрованные данные (*.enc);;Все файлы (*)")
        if not path: return
        # Анализируем enc как diff
        try:
            with open(path, "r", encoding="utf-8") as f:
                test_data = __import__("json").load(f)
            from security import MASTER_KEY_FIELD, DEFAULT_ADMIN_PW, _unlock_master_key
            stored_pw = test_data.get("__admin_pw_hint__", DEFAULT_ADMIN_PW)
            test_key  = _unlock_master_key(test_data[MASTER_KEY_FIELD], stored_pw)
            tmp_store = type(secure_store).__new__(type(secure_store))
            tmp_store._data = test_data; tmp_store._master_key = test_key
            tmp_store._admin_pw = stored_pw

            def _norm(s):
                return tuple(sorted([s.get("surname","").lower(), s.get("name","").lower()]))
            cur_k = {_norm(s) for s in secure_store.get_students()}
            diff = {
                "added_students": [s for s in tmp_store.get_students() if _norm(s) not in cur_k],
                "dup_students":   [s for s in tmp_store.get_students() if _norm(s) in cur_k],
                "added_teachers": [n for n in tmp_store.get_teachers() if n not in secure_store.get_teachers()],
                "dup_teachers":   [n for n in tmp_store.get_teachers() if n in secure_store.get_teachers()],
                "added_groups":   [g["name"] for g in tmp_store.get_groups()
                                   if g["name"] not in {gg["name"] for gg in secure_store.get_groups()}],
                "dup_groups":     [],
                "_test_data": test_data, "_test_key": test_key, "_stored_pw": stored_pw,
                "_tmp_db": None, "_tmp_subj": None,
            }
            dlg = ImportConfirmDialog("⚙ Импорт настроек", diff, self)
            dlg.confirmed.connect(lambda mode: self._do_import(diff, mode, dlg))
            dlg.cancelled.connect(dlg.close)
            dlg.show()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))


class AdminDashboard(QWidget):
    """Главная панель администратора."""

    def __init__(self, back_to_login_cb, parent=None):
        super().__init__(parent)
        self.back_cb = back_to_login_cb
        main_layout = QVBoxLayout(self)

        self.stack = QStackedWidget()
        main_layout.addWidget(self.stack)

        self.main_page = self._build_main_page()
        self.stack.addWidget(self.main_page)

    def _build_main_page(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        title = QLabel("⚙ Панель администратора ВСГУТУ")
        title.setStyleSheet("font-size:18px; font-weight:bold; margin-bottom:10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        buttons = [
            ("👨‍🏫 Преподаватели", self._open_teachers),
            ("🎓 Студенты", self._open_students),
            ("👥 Группы", self._open_groups),
            ("📚 Учебные предметы", self._open_subjects_editor),
            ("➕ Добавить данные", self._open_add),
            ("🔑 Ввести API ключ", self._open_api_key),
            ("📦 Перенос данных", self._open_transfer),
        ]
        for label, callback in buttons:
            btn = QPushButton(label)
            btn.setStyleSheet(STYLE_BTN)
            btn.setMinimumHeight(45)
            btn.clicked.connect(callback)
            layout.addWidget(btn)

        exit_btn = QPushButton("Выйти из панели")
        exit_btn.setStyleSheet(STYLE_BTN_RED)
        exit_btn.clicked.connect(self.back_cb)
        layout.addWidget(exit_btn)
        layout.addStretch()
        return w

    def _open_teachers(self):
        page = TeachersPage(self.stack, lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_students(self):
        page = StudentsPage(self.stack, lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_groups(self):
        page = GroupsPage(self.stack, lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_add(self):
        page = AddDataPage(lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_api_key(self):
        page = ApiKeyPage(lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_subjects_editor(self):
        page = SubjectsEditorPage(lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)

    def _open_transfer(self):
        page = TransferPage(lambda: self.stack.setCurrentWidget(self.main_page))
        self.stack.addWidget(page)
        self.stack.setCurrentWidget(page)
