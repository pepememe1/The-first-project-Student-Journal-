import sys
import os
import re
import requests
import json
import sqlite3
from PySide6.QtCore import Qt, QTimer, QThread, Signal as QSignal
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QInputDialog, QMessageBox,
    QStackedWidget, QLabel, QLineEdit, QComboBox, QFormLayout,
    QGridLayout, QMenu, QFileDialog, QTextEdit, QFrame, QScrollArea,
    QSizePolicy, QSpacerItem
)
from PySide6.QtGui import QColor, QBrush, QFont, QPalette
from core import GradeBook, Student, APP_VERSION
from datetime import datetime, timedelta
from security import secure_store
from admin_panel import AdminLoginPage, AdminDashboard, ALL_SUBJECTS

# ══════════════════════════════════════════════════════════════
#  ГЛОБАЛЬНЫЙ СТИЛЬ
# ══════════════════════════════════════════════════════════════
APP_STYLE = """
QMainWindow, QWidget {
    background-color: #1e2330;
    color: #e8eaf0;
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}
QLabel {
    color: #e8eaf0;
}
QLineEdit, QComboBox, QTextEdit {
    background-color: #2a3045;
    border: 1px solid #3d4460;
    border-radius: 6px;
    padding: 6px 10px;
    color: #e8eaf0;
    selection-background-color: #4a6fa5;
}
QLineEdit:focus, QComboBox:focus, QTextEdit:focus {
    border: 1px solid #5b8dd9;
}
QComboBox::drop-down {
    border: none;
    padding-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #2a3045;
    border: 1px solid #3d4460;
    color: #e8eaf0;
    selection-background-color: #4a6fa5;
}
QPushButton {
    background-color: #2c3e6b;
    color: #e8eaf0;
    border: none;
    border-radius: 6px;
    padding: 8px 16px;
    font-size: 13px;
}
QPushButton:hover {
    background-color: #3a5096;
}
QPushButton:pressed {
    background-color: #1e2d5a;
}
QPushButton:disabled {
    background-color: #2a2f40;
    color: #555e7a;
}
QTableWidget {
    background-color: #232840;
    border: 1px solid #3d4460;
    border-radius: 6px;
    gridline-color: #3d4460;
    color: #e8eaf0;
}
QTableWidget::item:selected {
    background-color: #3a5096;
}
QHeaderView::section {
    background-color: #2c3e6b;
    color: #c5d0f0;
    border: 1px solid #3d4460;
    padding: 6px;
    font-weight: bold;
}
QScrollBar:vertical {
    background: #1e2330;
    width: 8px;
    border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #3d4460;
    border-radius: 4px;
}
QScrollBar:horizontal {
    background: #1e2330;
    height: 8px;
    border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #3d4460;
    border-radius: 4px;
}
QScrollArea {
    border: none;
    background-color: transparent;
}
QFrame#card {
    background-color: #252b40;
    border: 1px solid #3d4460;
    border-radius: 10px;
}
QMessageBox {
    background-color: #1e2330;
    color: #e8eaf0;
}
"""

BTN_PRIMARY = "background-color:#3a5096;color:white;border-radius:6px;padding:8px 18px;font-weight:bold;"
BTN_SUCCESS = "background-color:#27ae60;color:white;border-radius:6px;padding:8px 18px;font-weight:bold;"
BTN_DANGER  = "background-color:#c0392b;color:white;border-radius:6px;padding:8px 18px;font-weight:bold;"
BTN_INFO    = "background-color:#2980b9;color:white;border-radius:6px;padding:8px 18px;font-weight:bold;"
BTN_PURPLE  = "background-color:#8e44ad;color:white;border-radius:6px;padding:8px 18px;font-weight:bold;"
BTN_BACK    = "background-color:#2a3045;color:#a0aac0;border:1px solid #3d4460;border-radius:6px;padding:7px 16px;"

# ══════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════

DEFAULT_GROUPS   = ["к74/1", "к74/2", "к74/3", "к75/0"]
DEFAULT_SUBJECTS = [
    "Компьютерные сети",
    "Основы алгоритмизации и программирования",
    "Основы проектирования баз данных",
    "Разработка программных модулей",
]


def get_groups():
    """Возвращает список групп из secure_store, либо дефолтные."""
    stored = secure_store.get_groups()
    return [g["name"] for g in stored] if stored else DEFAULT_GROUPS


def get_subjects_for_group(group_name: str):
    """Предметы группы из secure_store."""
    for g in secure_store.get_groups():
        if g["name"] == group_name:
            return g.get("subjects", DEFAULT_SUBJECTS)
    return DEFAULT_SUBJECTS


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def get_api_key() -> str:
    return secure_store.get_api_key()


def parse_logins():
    """Загружает преподавателей из secure_store или logins.txt."""
    teachers_data = secure_store.get_teachers()
    if teachers_data:
        return teachers_data, ""
    teachers = {}
    pw = ""
    logins_path = resource_path("logins.txt")
    if os.path.exists(logins_path):
        try:
            with open(logins_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if "password=" in line.lower():
                        pw = line.split("=")[1].strip()
                    elif "(" in line:
                        parts = line.split("(")
                        name = parts[0].strip()
                        subj = parts[1].replace(")", "").strip()
                        teachers[name] = {"password": pw, "subjects": [subj], "group_assignments": {}}
        except Exception as e:
            print(f"Ошибка чтения logins.txt: {e}")
    return teachers, pw


def clean_ai_text(text: str) -> str:
    """
    Убирает markdown-символы (#, *, **), иероглифы и слова на латинице
    из ответа ИИ, оставляя только русский текст и цифры.
    """
    # Убираем markdown заголовки и выделения
    text = re.sub(r"#{1,6}\s*", "", text)
    text = re.sub(r"\*{1,3}(.*?)\*{1,3}", r"\1", text)
    text = re.sub(r"_{1,2}(.*?)_{1,2}", r"\1", text)
    # Убираем code-блоки
    text = re.sub(r"```.*?```", "", text, flags=re.DOTALL)
    text = re.sub(r"`(.*?)`", r"\1", text)
    # Убираем иероглифы (китайские, японские, корейские и др.)
    text = re.sub(r"[\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]+", "", text)
    # Убираем английские слова (оставляем только кириллицу, цифры, знаки препинания)
    text = re.sub(r"\b[a-zA-Z]{2,}\b", "", text)
    # Убираем лишние пробелы и пустые строки
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def make_card(parent=None) -> QFrame:
    """Создаёт карточку с тёмным фоном и скруглёнными углами."""
    frame = QFrame(parent)
    frame.setObjectName("card")
    frame.setFrameShape(QFrame.StyledPanel)
    return frame


def make_title(text: str, size: int = 18) -> QLabel:
    lbl = QLabel(text)
    lbl.setAlignment(Qt.AlignCenter)
    lbl.setStyleSheet(f"font-size:{size}px;font-weight:bold;color:#c5d0f0;padding:6px 0;")
    return lbl


def make_btn(text: str, style: str = "") -> QPushButton:
    btn = QPushButton(text)
    if style:
        btn.setStyleSheet(style)
    btn.setMinimumHeight(36)
    return btn


# ══════════════════════════════════════════════════════════════
#  ИИ-ПОМОЩНИК
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
#  ФОНОВЫЙ ПОТОК ДЛЯ ЗАПРОСОВ К ИИ
# ══════════════════════════════════════════════════════════════

from PySide6.QtCore import QThread, Signal as QSignal


class AIRequestThread(QThread):
    """Выполняет HTTP-запрос к OpenRouter в отдельном потоке."""
    finished = QSignal(str)
    error    = QSignal(str)

    def __init__(self, api_key: str, system_prompt: str, question: str, history: list = None):
        super().__init__()
        self.api_key       = api_key
        self.system_prompt = system_prompt
        self.question      = question
        self.history       = history or []

    def run(self):
        try:
            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            }
            # Системный промпт + история последних 20 сообщений
            messages = [{"role": "system", "content": self.system_prompt}]
            messages.extend(self.history[-20:])
            payload = {
                "model": "stepfun/step-3.5-flash:free",
                "messages": messages,
            }
            resp   = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=60,
            )
            result = resp.json()
            if "choices" not in result:
                self.error.emit(f"Ошибка ответа: {result}")
            else:
                self.finished.emit(result["choices"][0]["message"]["content"])
        except Exception as e:
            self.error.emit(f"Ошибка соединения: {e}")


# ══════════════════════════════════════════════════════════════
#  ИИ-ПОМОЩНИК  (не блокирует UI)
# ══════════════════════════════════════════════════════════════

class AIWidget(QWidget):
    """
    Универсальный ИИ-чат.
    role:       'student' | 'teacher' | 'admin'
    context_fn: callable -> str   (системный промпт)
    back_fn:    callable -> QWidget
    stack_ref:  QStackedWidget
    """

    def __init__(self, role: str, context_fn, back_fn, stack_ref, parent=None):
        super().__init__(parent)
        self.role       = role
        self.context_fn = context_fn
        self.back_fn    = back_fn
        self.stack_ref  = stack_ref
        self._thread    = None   # текущий поток запроса
        self._history   = []     # история сообщений сессии

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        # ── Заголовок
        header = QHBoxLayout()
        self.back_btn = make_btn("← Назад", BTN_BACK)
        self.back_btn.setFixedWidth(100)
        self.back_btn.clicked.connect(self._go_back)
        header.addWidget(self.back_btn)
        header.addWidget(make_title("🤖  ИИ Помощник"), 1)
        header.addSpacing(100)
        layout.addLayout(header)

        # ── Область чата
        self.chat_area = QTextEdit()
        self.chat_area.setReadOnly(True)
        self.chat_area.setStyleSheet(
            "background-color:#1a2035;border:1px solid #3d4460;border-radius:8px;"
            "padding:10px;font-size:13px;color:#dde3f5;"
        )
        layout.addWidget(self.chat_area, 1)

        # Приветствие при открытии чата
        self._show_greeting()

        # ── Индикатор загрузки
        self.loading_lbl = QLabel("⏳ ИИ думает...")
        self.loading_lbl.setStyleSheet("color:#7eb8f7;font-size:12px;")
        self.loading_lbl.hide()
        layout.addWidget(self.loading_lbl)

        # ── Поле ввода
        inp_row = QHBoxLayout()
        self.chat_input = QLineEdit()
        self.chat_input.setPlaceholderText("Введите сообщение...")
        self.chat_input.setMinimumHeight(38)
        self.send_btn = make_btn("Отправить ➤", BTN_PRIMARY)
        self.send_btn.setFixedWidth(130)
        inp_row.addWidget(self.chat_input)
        inp_row.addWidget(self.send_btn)
        layout.addLayout(inp_row)

        self.send_btn.clicked.connect(self._send)
        self.chat_input.returnPressed.connect(self._send)

    def _show_greeting(self):
        """Персональное приветствие при открытии чата."""
        greetings = {
            "student": (
                "Привет! Я твой учебный помощник.\n"
                "Могу рассказать о твоих оценках, посещаемости и среднем балле.\n"
                "Просто спроси — постараюсь помочь!"
            ),
            "teacher": (
                "Добрый день! Рад помочь вам.\n"
                "Готов помочь с анализом успеваемости группы и ответить\n"
                "на вопросы по вашему журналу. Чем могу быть полезен?"
            ),
            "admin": (
                "Добрый день! Я ИИ-ассистент системы ВСГУТУ.\n"
                "Готов предоставить статистику по группам, студентам и преподавателям.\n"
                "Чем могу помочь?"
            ),
        }
        msg = greetings.get(self.role, "Здравствуйте! Чем могу помочь?")
        msg_html = msg.replace("\n", "<br>")
        self.chat_area.append(
            f'<span style="color:#7dd87d;font-weight:bold;">ИИ:</span><br>'
            f'<span style="color:#dde3f5;">{msg_html}</span><br>'
        )

    def _go_back(self):
        target = self.back_fn()
        if target:
            self.stack_ref.setCurrentWidget(target)

    def _set_busy(self, busy: bool):
        """Блокирует/разблокирует ввод во время запроса."""
        self.send_btn.setEnabled(not busy)
        self.chat_input.setEnabled(not busy)
        self.back_btn.setEnabled(not busy)
        self.loading_lbl.setVisible(busy)

    def _send(self):
        text = self.chat_input.text().strip()
        if not text:
            return
        if self._thread and self._thread.isRunning():
            return  # уже идёт запрос

        # Показываем сообщение пользователя
        self.chat_area.append(
            f'<span style="color:#7eb8f7;font-weight:bold;">Вы:</span>'
            f'<span style="color:#dde3f5;"> {text}</span><br>'
        )
        self.chat_input.clear()

        # Сохраняем в историю сессии
        self._history.append({"role": "user", "content": text})

        api_key = get_api_key()
        if not api_key:
            self.chat_area.append(
                '<span style="color:#e74c3c;font-weight:bold;">ИИ:</span>'
                '<span style="color:#dde3f5;"> Не указан API ключ. '
                'Обратитесь к администратору.</span><br>'
            )
            return

        # Запускаем запрос в фоновом потоке
        system_prompt = self.context_fn()
        self._thread  = AIRequestThread(api_key, system_prompt, text, list(self._history))
        self._thread.finished.connect(self._on_answer)
        self._thread.error.connect(self._on_error)
        self._thread.start()
        self._set_busy(True)

    def _on_answer(self, raw_answer: str):
        self._set_busy(False)
        answer = clean_ai_text(raw_answer)
        self._history.append({"role": "assistant", "content": raw_answer})
        # Каждый перенос строки → <br> для отображения в HTML
        answer_html = answer.replace("\n", "<br>")
        self.chat_area.append(
            f'<span style="color:#7dd87d;font-weight:bold;">ИИ:</span><br>'
            f'<span style="color:#dde3f5;">{answer_html}</span><br>'
        )
        sb = self.chat_area.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _on_error(self, err: str):
        self._set_busy(False)
        self.chat_area.append(
            f'<span style="color:#e74c3c;font-weight:bold;">Ошибка:</span>'
            f'<span style="color:#dde3f5;"> {err}</span><br>'
        )


# ══════════════════════════════════════════════════════════════
#  КОНТЕКСТЫ ДЛЯ ИИ  (отдельные функции для чистоты)
# ══════════════════════════════════════════════════════════════

def build_student_context(cur_stud: dict) -> str:
    """
    Собирает все оценки студента из SQLite и передаёт ИИ.
    Студент видит ТОЛЬКО свои данные.
    """
    group   = cur_stud["g"]
    surname = cur_stud["f"]
    name    = cur_stud["n"]

    subjects = get_subjects_for_group(group)
    lines = []
    for subj in subjects:
        book = GradeBook(group, subj)
        student = next((s for s in book.spisok_stud if s.f == surname), None)
        if student:
            records = []
            for lesson in book.lessons:
                val = student.records.get(lesson.id, "")
                if val:
                    records.append(
                        f"  {lesson.type} №{lesson.number} "
                        f"({lesson.date}, тема: {lesson.topic}) — оценка: {val}"
                    )
            if records:
                lines.append(f"Предмет: {subj}\n" + "\n".join(records))
            else:
                lines.append(f"Предмет: {subj}\n  (оценок пока нет)")

    data_block = "\n\n".join(lines) if lines else "Оценок в системе не найдено."

    return (
        f"Ты ИИ-ассистент электронного журнала ВСГУТУ.\n"
        f"Ты общаешься со студентом {surname} {name}, группа {group}.\n"
        "Обращайся к студенту на 'ты', дружелюбно и по-простому.\n\n"
        "ПРАВИЛА ОТВЕТА:\n"
        "1. Отвечай ТОЛЬКО на русском языке. Никаких #, *, **, _. Никакого markdown.\n"
        "2. КАЖДЫЙ факт или пункт — с новой строки. Не пиши всё в одну строку.\n"
        "3. Отвечай кратко и по делу.\n"
        "4. Ты видишь ТОЛЬКО данные этого студента.\n"
        "5. НЕ придумывай оценки. Если данных нет — скажи об этом.\n"
        "6. 'Н' = пропуск занятия (при расчёте среднего = 2).\n"
        "   'Б' = болел (на балл не влияет). 'О' = опоздал.\n"
        "7. В столбцах ЛЕКЦИЙ '✓' или пустая ячейка = присутствовал (НЕ оценка!).\n"
        "   '✓' в лекции = ПОСЕЩЕНИЕ, а не зачёт.\n"
        "8. Будь вежлив и поддерживай студента.\n\n"
        f"ДАННЫЕ СТУДЕНТА:\n{data_block}"
    )


def build_teacher_context(teacher_name: str, book) -> str:
    """
    Контекст для преподавателя: его предметы и список студентов с оценками.
    """
    teachers  = secure_store.get_teachers()
    data      = teachers.get(teacher_name, {})
    subjects  = data.get("subjects", [])

    lines = [
        f"Преподаватель: {teacher_name}",
        f"Ведёт предметы: {', '.join(subjects) or 'не указаны'}",
    ]

    if book:
        lines.append(f"\nТекущий журнал: группа {book.group}, предмет {book.subject}")
        lines.append(f"Студентов: {len(book.spisok_stud)}")
        for s in book.spisok_stud:
            avg = book.calculate_average(s)
            lines.append(f"  {s.f} {s.n} — средний балл: {avg:.2f}")

    return (
        f"Ты ИИ-ассистент электронного журнала ВСГУТУ.\n"
        f"Ты общаешься с преподавателем {teacher_name}. Обращайся на 'вы'.\n\n"
        "ПРАВИЛА ОТВЕТА:\n"
        "1. Отвечай ТОЛЬКО на русском языке. Никаких #, *, **, _. Никакого markdown.\n"
        "2. КАЖДЫЙ факт или пункт — с новой строки. Не пиши всё в одну строку.\n"
        "3. Отвечай кратко, структурированно и по делу.\n"
        "4. НЕ раскрывай пароли и личные данные других пользователей.\n"
        "5. Отвечай только в контексте журнала и учебного процесса.\n"
        "6. НЕ придумывай данные которых нет.\n"
        "7. Будь вежлив и профессионален.\n\n"
        + "\n".join(lines)
    )


def build_admin_context() -> str:
    """
    Контекст для администратора: полная статистика системы.
    """
    groups   = secure_store.get_groups()
    teachers = secure_store.get_teachers()
    students = secure_store.get_students()

    group_lines = []
    for g in groups:
        subjs = g.get("subjects", [])
        group_lines.append(f"  {g['name']}: {len(subjs)} предметов")

    teacher_lines = []
    for name, d in teachers.items():
        subjs = d.get("subjects", [])
        teacher_lines.append(f"  {name}: {', '.join(subjs[:3]) or '—'}")

    return (
        "Ты ИИ-ассистент электронного журнала ВСГУТУ. Ты помогаешь администратору.\n"
        "СТРОГИЕ ПРАВИЛА:\n"
        "1. Отвечай ТОЛЬКО на русском языке. Никаких символов #, *, **, _. Никакого markdown.\n"
        "2. НЕ раскрывай пароли, API ключи, секретные данные.\n"
        "3. Ты можешь отвечать на вопросы об общей статистике, группах, студентах, преподавателях.\n\n"
        f"СТАТИСТИКА СИСТЕМЫ:\n"
        f"Групп: {len(groups)}\n" + "\n".join(group_lines) + "\n\n"
        f"Преподавателей: {len(teachers)}\n" + "\n".join(teacher_lines) + "\n\n"
        f"Студентов в системе: {len(students)}"
    )


# ══════════════════════════════════════════════════════════════
#  ПЕРЕНОС ДАННЫХ
# ══════════════════════════════════════════════════════════════

class TransferWidget(QWidget):
    """
    Страница переноса данных между компьютерами.
    Доступна и преподавателю, и администратору.

    Схема работы:
      Экспорт — копирует secure_data.enc в выбранное место.
      Импорт  — заменяет текущий secure_data.enc файлом с другого ПК.
      Файл зашифрован паролем администратора: открыть вручную невозможно.
    """

    def __init__(self, back_fn, stack_ref, parent=None):
        super().__init__(parent)
        self.back_fn   = back_fn
        self.stack_ref = stack_ref

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(14)

        # Шапка
        header = QHBoxLayout()
        back_btn = make_btn("← Назад", BTN_BACK)
        back_btn.setFixedWidth(100)
        back_btn.clicked.connect(self._go_back)
        header.addWidget(back_btn)
        header.addWidget(make_title("📦  Перенос данных между компьютерами"), 1)
        layout.addLayout(header)

        # Пояснение
        info = QLabel(
            "Данные шифруются паролем администратора.\n"
            "Файл можно скопировать на любой ПК — программа его расшифрует.\n"
            "Открыть файл вручную невозможно — там только случайный набор символов."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            "background-color:#1e2d1e;border:1px solid #2d5a2d;"
            "border-radius:8px;padding:12px;color:#7dd87d;font-size:12px;"
        )
        layout.addWidget(info)

        # ── Экспорт
        exp_card = make_card()
        exp_lay  = QVBoxLayout(exp_card)
        exp_lay.setContentsMargins(16, 14, 16, 14)

        exp_title = QLabel("⬆  Экспорт — отправить данные на другой ПК")
        exp_title.setStyleSheet("font-weight:bold;color:#c5d0f0;font-size:13px;")
        exp_lay.addWidget(exp_title)

        exp_desc = QLabel(
            "Сохраняет зашифрованный файл со всеми данными.\n"
            "Передайте его на другой ПК (флешка, почта, мессенджер) и импортируйте там."
        )
        exp_desc.setWordWrap(True)
        exp_desc.setStyleSheet("color:#a0aac0;font-size:12px;")
        exp_lay.addWidget(exp_desc)

        export_btn = make_btn("💾  Экспортировать данные...", BTN_SUCCESS)
        export_btn.setMinimumHeight(40)
        export_btn.clicked.connect(self._export)
        exp_lay.addWidget(export_btn)
        layout.addWidget(exp_card)

        # ── Импорт
        imp_card = make_card()
        imp_lay  = QVBoxLayout(imp_card)
        imp_lay.setContentsMargins(16, 14, 16, 14)

        imp_title = QLabel("⬇  Импорт — принять данные с другого ПК")
        imp_title.setStyleSheet("font-weight:bold;color:#c5d0f0;font-size:13px;")
        imp_lay.addWidget(imp_title)

        imp_desc = QLabel(
            "Загружает файл экспорта и заменяет текущие данные.\n"
            "⚠ Текущие данные будут полностью перезаписаны!"
        )
        imp_desc.setWordWrap(True)
        imp_desc.setStyleSheet("color:#e8a87c;font-size:12px;")
        imp_lay.addWidget(imp_desc)

        import_btn = make_btn("📂  Импортировать данные...", BTN_INFO)
        import_btn.setMinimumHeight(40)
        import_btn.clicked.connect(self._import)
        imp_lay.addWidget(import_btn)
        layout.addWidget(imp_card)

        layout.addStretch()

    def _show_greeting(self):
        """Персональное приветствие при открытии чата."""
        greetings = {
            "student": (
                "Привет! Я твой учебный помощник.\n"
                "Могу рассказать о твоих оценках, посещаемости и среднем балле.\n"
                "Просто спроси — постараюсь помочь!"
            ),
            "teacher": (
                "Добрый день! Рад помочь вам.\n"
                "Готов помочь с анализом успеваемости группы и ответить\n"
                "на вопросы по вашему журналу. Чем могу быть полезен?"
            ),
            "admin": (
                "Добрый день! Я ИИ-ассистент системы ВСГУТУ.\n"
                "Готов предоставить статистику по группам, студентам и преподавателям.\n"
                "Чем могу помочь?"
            ),
        }
        msg = greetings.get(self.role, "Здравствуйте! Чем могу помочь?")
        msg_html = msg.replace("\n", "<br>")
        self.chat_area.append(
            f'<span style="color:#7dd87d;font-weight:bold;">ИИ:</span><br>'
            f'<span style="color:#dde3f5;">{msg_html}</span><br>'
        )

    def _go_back(self):
        target = self.back_fn()
        if target:
            self.stack_ref.setCurrentWidget(target)

    def _export(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить файл данных", "vsgutu_data.enc",
            "Зашифрованные данные (*.enc);;Все файлы (*)"
        )
        if not path:
            return
        ok, msg = secure_store.export_portable(path)
        if ok:
            QMessageBox.information(
                self, "Экспорт выполнен",
                f"✅ {msg}\n\n"
                "Передайте этот файл на другой ПК\n"
                "и импортируйте его там через кнопку «📦 Перенос данных»."
            )
        else:
            QMessageBox.critical(self, "Ошибка экспорта", msg)

    def _import(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть файл данных", "",
            "Зашифрованные данные (*.enc);;Все файлы (*)"
        )
        if not path:
            return
        confirm = QMessageBox.question(
            self, "Подтверждение импорта",
            "Текущие данные будут полностью заменены данными из файла.\n\n"
            "Продолжить?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return
        ok, msg = secure_store.import_portable(path)
        if ok:
            QMessageBox.information(
                self, "Импорт выполнен",
                f"✅ {msg}\n\nДанные успешно загружены. Перезайдите в систему."
            )
            # Возвращаемся на экран входа чтобы данные перечитались
            self.stack_ref.setCurrentIndex(0)
        else:
            QMessageBox.critical(self, "Ошибка импорта", msg)


# ══════════════════════════════════════════════════════════════
#  ГЛАВНОЕ ОКНО
# ══════════════════════════════════════════════════════════════

class MainAppWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"GradeBookAI — {APP_VERSION}")
        self.resize(1300, 750)
        self.setStyleSheet(APP_STYLE)

        self.teachers_db, self.teacher_pass = parse_logins()
        self.book             = None
        self.cur_stud         = None
        self.cur_teacher_name = None

        self._admin_click_count = 0
        self._admin_click_timer = QTimer()
        self._admin_click_timer.setSingleShot(True)
        self._admin_click_timer.timeout.connect(lambda: setattr(self, "_admin_click_count", 0))

        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        self._init_login_ui()
        self._init_teacher_ui()
        self._init_student_ui()
        self._init_admin_ui()

    # ──────────────────────────────────────────────────────────
    #  ВХОД (главная страница)
    # ──────────────────────────────────────────────────────────
    def _init_login_ui(self):
        self.login_widget = QWidget()
        outer = QVBoxLayout(self.login_widget)
        outer.setAlignment(Qt.AlignCenter)

        card = make_card()
        card.setFixedWidth(460)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(40, 36, 40, 36)
        card_layout.setSpacing(14)

        # Логотип / заголовок
        logo = QLabel("📚")
        logo.setAlignment(Qt.AlignCenter)
        logo.setStyleSheet("font-size:42px;")
        card_layout.addWidget(logo)

        title = make_title("СИСТЕМА УЧЁТА УСПЕВАЕМОСТИ", 17)
        card_layout.addWidget(title)

        subtitle = QLabel("ВСГУТУ")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color:#7eb8f7;font-size:13px;margin-bottom:10px;")
        card_layout.addWidget(subtitle)

        # Переключатель ученик/учитель
        toggle = QHBoxLayout()
        self._btn_mode_stud = make_btn("Ученик", BTN_PRIMARY)
        self._btn_mode_tchr = make_btn("Учитель")
        self._btn_mode_stud.setCheckable(True)
        self._btn_mode_tchr.setCheckable(True)
        self._btn_mode_stud.setChecked(True)
        toggle.addWidget(self._btn_mode_stud)
        toggle.addWidget(self._btn_mode_tchr)
        card_layout.addLayout(toggle)

        self.login_stack = QStackedWidget()
        card_layout.addWidget(self.login_stack)

        # ── Форма ученика
        sw = QWidget()
        sl = QVBoxLayout(sw)
        sl.setSpacing(8)

        self.sn = QLineEdit(); self.sn.setPlaceholderText("Имя")
        self.sf = QLineEdit(); self.sf.setPlaceholderText("Фамилия")

        form_s = QFormLayout()
        form_s.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form_s.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        form_s.setSpacing(8)
        form_s.addRow("Имя:", self.sn)
        form_s.addRow("Фамилия:", self.sf)
        sl.addLayout(form_s)

        hint = QLabel("Группа и предметы определяются автоматически по данным из базы.")
        hint.setStyleSheet("color:#7eb8f7;font-size:11px;")
        hint.setWordWrap(True)
        sl.addWidget(hint)

        btn_s = make_btn("Войти как ученик", BTN_SUCCESS)
        btn_s.clicked.connect(self._login_student)
        self.sn.returnPressed.connect(self._login_student)
        self.sf.returnPressed.connect(self._login_student)
        sl.addWidget(btn_s)
        self.login_stack.addWidget(sw)

        # ── Форма учителя
        tw = QWidget()
        tl = QVBoxLayout(tw)
        tl.setSpacing(8)

        self.tn = QLineEdit(); self.tn.setPlaceholderText("Фамилия Имя")
        self.tp = QLineEdit(); self.tp.setPlaceholderText("Пароль")
        self.tp.setEchoMode(QLineEdit.Password)

        form_t = QFormLayout()
        form_t.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form_t.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        form_t.setSpacing(8)
        form_t.addRow("Фамилия Имя:", self.tn)
        form_t.addRow("Пароль:", self.tp)
        tl.addLayout(form_t)

        btn_t = make_btn("Войти как учитель", BTN_SUCCESS)
        btn_t.clicked.connect(self._login_teacher)
        self.tp.returnPressed.connect(self._login_teacher)
        tl.addWidget(btn_t)
        self.login_stack.addWidget(tw)

        # Переключение вкладок
        def switch_student():
            self._btn_mode_stud.setStyleSheet(BTN_PRIMARY)
            self._btn_mode_tchr.setStyleSheet("")
            self.login_stack.setCurrentIndex(0)

        def switch_teacher():
            self._btn_mode_tchr.setStyleSheet(BTN_PRIMARY)
            self._btn_mode_stud.setStyleSheet("")
            self.login_stack.setCurrentIndex(1)

        self._btn_mode_stud.clicked.connect(switch_student)
        self._btn_mode_tchr.clicked.connect(switch_teacher)

        outer.addWidget(card)

        # Скрытая кнопка-квадрат (левый нижний угол)
        self._secret_btn = QPushButton(self.login_widget)
        self._secret_btn.setFixedSize(14, 14)
        self._secret_btn.setStyleSheet(
            "background-color:rgba(255,255,255,8);"
            "border:1px solid rgba(255,255,255,10);"
            "border-radius:2px;"
        )
        self._secret_btn.clicked.connect(self._on_secret_click)

        self.stack.addWidget(self.login_widget)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "_secret_btn"):
            self._secret_btn.move(6, self.login_widget.height() - 20)

    def showEvent(self, event):
        super().showEvent(event)
        if hasattr(self, "_secret_btn"):
            self._secret_btn.move(6, self.login_widget.height() - 20)

    def _on_secret_click(self):
        self._admin_click_count += 1
        self._admin_click_timer.start(3000)
        if self._admin_click_count >= 5:
            self._admin_click_count = 0
            self._admin_click_timer.stop()
            self.stack.setCurrentWidget(self.admin_login_widget)

    # ──────────────────────────────────────────────────────────
    #  АДМИНИСТРАТОР
    # ──────────────────────────────────────────────────────────
    def _init_admin_ui(self):
        self.admin_login_widget = AdminLoginPage()
        self.admin_login_widget.login_success.connect(
            lambda: self.stack.setCurrentWidget(self.admin_dashboard)
        )
        self.stack.addWidget(self.admin_login_widget)

        self.admin_dashboard = AdminDashboard(
            back_to_login_cb=lambda: self.stack.setCurrentWidget(self.login_widget)
        )
        self.stack.addWidget(self.admin_dashboard)

        # ИИ для администратора
        self.admin_ai = AIWidget(
            role="admin",
            context_fn=build_admin_context,
            back_fn=lambda: self.admin_dashboard,
            stack_ref=self.stack,
        )
        self.stack.addWidget(self.admin_ai)

    # ──────────────────────────────────────────────────────────
    #  УЧЕНИК
    # ──────────────────────────────────────────────────────────
    def _init_student_ui(self):
        self.student_widget = QWidget()
        self._student_outer = QVBoxLayout(self.student_widget)
        self._student_outer.setContentsMargins(0, 0, 0, 0)
        self.stack.addWidget(self.student_widget)

    def _draw_student_dash(self):
        # Очищаем
        while self._student_outer.count():
            item = self._student_outer.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        main = QWidget()
        layout = QVBoxLayout(main)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(14)

        # Заголовок
        header_row = QHBoxLayout()
        back_btn = make_btn("← Выход", BTN_BACK)
        back_btn.setFixedWidth(100)
        back_btn.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        header_row.addWidget(back_btn)
        name_lbl = make_title(
            f"👤  {self.cur_stud['f']} {self.cur_stud['n']}  •  Группа: {self.cur_stud['g']}", 16
        )
        header_row.addWidget(name_lbl, 1)

        ai_btn = make_btn("🤖 ИИ Помощник", BTN_INFO)
        ai_btn.setFixedWidth(150)
        ai_btn.clicked.connect(self._open_student_ai)
        header_row.addWidget(ai_btn)
        layout.addLayout(header_row)

        # Разделитель
        line = QFrame(); line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color:#3d4460;")
        layout.addWidget(line)

        # Предметы группы из secure_store
        subjects = get_subjects_for_group(self.cur_stud["g"])

        lbl = QLabel("Выберите предмет для просмотра журнала:")
        lbl.setStyleSheet("color:#a0aac0;font-size:12px;margin-bottom:4px;")
        layout.addWidget(lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border:none;background:transparent;")
        inner = QWidget()
        grid = QGridLayout(inner)
        grid.setSpacing(10)

        for i, subj in enumerate(subjects):
            btn = make_btn(subj)
            btn.setMinimumHeight(56)
            btn.setStyleSheet(
                "background-color:#252b40;border:1px solid #3d4460;border-radius:8px;"
                "color:#c5d0f0;font-size:12px;padding:6px;"
                "text-align:left;"
            )
            btn.clicked.connect(lambda _, s=subj: self._show_student_journal(s))
            grid.addWidget(btn, i // 2, i % 2)

        scroll.setWidget(inner)
        layout.addWidget(scroll, 1)

        self._student_outer.addWidget(main)

    def _open_student_ai(self):
        """Открывает ИИ с контекстом оценок студента."""
        stud = self.cur_stud
        ai = AIWidget(
            role="student",
            context_fn=lambda: build_student_context(stud),
            back_fn=lambda: self.student_widget,
            stack_ref=self.stack,
        )
        self.stack.addWidget(ai)
        self.stack.setCurrentWidget(ai)

    def _show_student_journal(self, subj: str):
        self.book = GradeBook(self.cur_stud["g"], subj)

        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(10)

        # Шапка
        header_row = QHBoxLayout()
        back_btn = make_btn("← Назад", BTN_BACK)
        back_btn.setFixedWidth(100)
        back_btn.clicked.connect(lambda: self.stack.setCurrentWidget(self.student_widget))
        header_row.addWidget(back_btn)
        header_row.addWidget(make_title(f"📋  {subj}  |  {self.cur_stud['g']}", 15), 1)
        layout.addLayout(header_row)

        table = QTableWidget()
        layout.addWidget(table)

        table.setRowCount(1)
        table.setColumnCount(len(self.book.lessons) + 2)
        headers = ["Фамилия", "Имя"] + [
            f"{l.type} {l.number}\n{l.date}\n{l.topic}" for l in self.book.lessons
        ]
        table.setHorizontalHeaderLabels(headers)

        s = next((x for x in self.book.spisok_stud if x.f == self.cur_stud["f"]), None)
        if not s:
            QMessageBox.warning(self, "Информация", "Вас нет в списке по этому предмету.")
            return

        table.setItem(0, 0, QTableWidgetItem(s.f))
        table.setItem(0, 1, QTableWidgetItem(s.n))

        for i, l in enumerate(self.book.lessons):
            combo = QComboBox()
            if l.type == "Практика":
                combo.addItems(["", "2", "3", "4", "5", "Н"])
            else:
                combo.addItems(["", "О", "Б", "✓", "Н"])
            combo.setCurrentText(s.records.get(l.id, ""))
            combo.setEnabled(False)  # студент только просматривает
            table.setCellWidget(0, 2 + i, combo)

        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

    # ──────────────────────────────────────────────────────────
    #  ПРЕПОДАВАТЕЛЬ
    # ──────────────────────────────────────────────────────────
    def _init_teacher_ui(self):
        self.teacher_widget = QWidget()
        layout = QVBoxLayout(self.teacher_widget)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        # Заголовок
        header_row = QHBoxLayout()
        back_btn = make_btn("← Меню", BTN_BACK)
        back_btn.setFixedWidth(90)
        back_btn.clicked.connect(lambda: self.stack.setCurrentIndex(0))
        header_row.addWidget(back_btn)
        self._teacher_title_lbl = make_title("Журнал преподавателя", 15)
        header_row.addWidget(self._teacher_title_lbl, 1)
        layout.addLayout(header_row)

        # Таблица
        self.t_table = QTableWidget()
        self.t_table.setStyleSheet(
            "QTableCornerButton::section{"
            "background-color:#252b40;border:1px solid #3d4460;}"
        )
        self.t_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.t_table.horizontalHeader().customContextMenuRequested.connect(self._header_context_menu)
        layout.addWidget(self.t_table, 1)

        # Панель кнопок
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        def add(lbl, style, fn):
            b = make_btn(lbl, style)
            b.clicked.connect(fn)
            btn_row.addWidget(b)

        add("Добавить студента",  "",        self._add_student)
        add("🗑 Удалить студента", BTN_DANGER, self._delete_student)
        add("Добавить занятие",   "",        self._add_lesson)
        add("Экзамен",            "",        self._add_exam)
        add("💾 Сохранить",       BTN_SUCCESS, self._save_data)
        add("📊 Excel экспорт",   BTN_SUCCESS, self._export_excel)
        add("📥 Excel импорт",    BTN_INFO,    self._import_excel)
        add("🗄 SQLite импорт",   BTN_PURPLE,  self._import_sqlite)
        add("📦 Перенос данных",  BTN_PURPLE,  self._open_transfer)
        add("🤖 ИИ Помощник",     BTN_INFO,    self._open_teacher_ai)

        layout.addLayout(btn_row)
        self.stack.addWidget(self.teacher_widget)

    def _open_teacher_ai(self):
        book = self.book
        name = self.cur_teacher_name or "Преподаватель"
        ai = AIWidget(
            role="teacher",
            context_fn=lambda: build_teacher_context(name, book),
            back_fn=lambda: self.teacher_widget,
            stack_ref=self.stack,
        )
        self.stack.addWidget(ai)
        self.stack.setCurrentWidget(ai)

    def _open_transfer(self):
        """Открывает страницу переноса данных для преподавателя."""
        w = TransferWidget(back_fn=lambda: self.teacher_widget, stack_ref=self.stack)
        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

    def _add_student(self):
        f, ok1 = QInputDialog.getText(self, "Фамилия", "Введите фамилию:")
        n, ok2 = QInputDialog.getText(self, "Имя",     "Введите имя:")
        if ok1 and ok2 and f.strip() and n.strip():
            f, n = f.strip(), n.strip()
            # Добавляем в журнал (SQLite)
            self.book.add_student(Student(n, f, self.book.group))
            # Добавляем в общую базу secure_store
            students = secure_store.get_students()
            already = any(
                s.get("surname", "").lower() == f.lower()
                and s.get("name", "").lower() == n.lower()
                for s in students
            )
            if not already:
                students.append({
                    "name": n,
                    "surname": f,
                    "group": self.book.group
                })
                secure_store.set_students(students)
            self._update_teacher_table()

    def _delete_student(self):
        if not self.book or not self.book.spisok_stud:
            QMessageBox.warning(self, "Ошибка", "Нет студентов для удаления.")
            return
        names = [f"{s.f} {s.n}" for s in self.book.spisok_stud]
        chosen, ok = QInputDialog.getItem(
            self, "Удалить студента", "Выберите студента:", names, 0, False
        )
        if not ok:
            return
        confirm = QMessageBox.question(
            self, "Подтверждение",
            f"Удалить студента «{chosen}» полностью?\n\n"
            "Будут удалены:\n"
            "• все его оценки во всех журналах\n"
            "• его запись в базе данных\n\n"
            "Это действие необратимо.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm != QMessageBox.Yes:
            return

        parts   = chosen.split(" ", 1)
        surname = parts[0].strip()
        name    = parts[1].strip() if len(parts) > 1 else ""

        # Удаляем из SQLite (grades + students)
        self.book.delete_student(surname, name)

        # Удаляем из secure_store чтобы не появлялся при входе
        store_students = secure_store.get_students()
        store_students = [
            s for s in store_students
            if not (s.get("surname", "").strip().lower() == surname.lower()
                    and s.get("name", "").strip().lower() == name.lower())
        ]
        secure_store.set_students(store_students)

        self._update_teacher_table()
        QMessageBox.information(self, "Готово", f"Студент «{chosen}» удалён полностью.")

    def _add_lesson(self):
        t, ok = QInputDialog.getItem(
            self, "Тип занятия", "Выберите тип:", ["Лекция", "Практика"], 0, False
        )
        if not ok:
            return
        date_str, ok_d = QInputDialog.getText(
            self, "Дата занятия", "Введите дату (дд.мм.гггг):",
            text=datetime.now().strftime("%d.%m.%Y")
        )
        if not ok_d:
            return
        topic, ok2 = QInputDialog.getText(self, "Тема", "Введите тему занятия:")
        if not ok2:
            return
        lesson = self.book.add_lesson(t, topic=topic, date=date_str)
        self.book.save_to_sqlite()
        self._update_teacher_table()
        if t == "Лекция":
            QMessageBox.information(self, "Занятие добавлено",
                f"Лекция №{lesson.number} ({date_str})\nСозданы 2 столбца: 1-й час и 2-й час.")

    def _add_exam(self):
        date_str, ok = QInputDialog.getText(
            self, "Дата экзамена", "Введите дату (дд.мм.гггг):",
            text=datetime.now().strftime("%d.%m.%Y")
        )
        if not ok:
            return
        topic, ok2 = QInputDialog.getText(self, "Тема экзамена", "Введите тему:")
        if not ok2:
            return
        lesson = self.book.add_lesson("Экзамен")
        lesson.date        = date_str
        lesson.topic       = topic
        lesson.retake_date = ""
        self.book.save_to_sqlite()
        self._update_teacher_table()

    def _ask_retake_date(self, lesson):
        """Назначить дату пересдачи. Автоматически +7 дней от даты экзамена."""
        try:
            exam_dt      = datetime.strptime(lesson.date, "%d.%m.%Y")
            default_date = (exam_dt + timedelta(days=7)).strftime("%d.%m.%Y")
        except Exception:
            default_date = (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y")

        date_str, ok = QInputDialog.getText(
            self, "Дата пересдачи",
            f"Экзамен «{lesson.topic}» — {lesson.date}\n"
            "Студент не сдал. Назначьте дату пересдачи\n"
            "(автоматически +7 дней, можно изменить):",
            text=default_date
        )
        if ok and date_str.strip():
            self.book.set_retake_date(lesson.id, date_str.strip())
            self._update_teacher_table()

    def _save_data(self):
        if self.book:
            self.book.save_to_sqlite()
            QMessageBox.information(self, "Успех", "Данные сохранены.")

    def _export_excel(self):
        if not self.book:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчёт", "", "Excel Files (*.xlsx)"
        )
        if path:
            try:
                self.book.export_to_excel(path)
                QMessageBox.information(self, "Excel", "Отчёт успешно выгружен!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", str(e))

    def _import_excel(self):
        """
        Импорт Excel. Поддерживает два формата:
        1. Формат колледжа: строка 1 = даты, строка 2 = '1 час'/'2 час',
           столбец 2 = 'ФИО обучающегося' (Фамилия И.О.)
        2. Старый формат: столбец 1 = фамилия, столбец 2 = имя
        Студенты автоматически добавляются в группу текущего журнала.
        """
        if not self.book:
            QMessageBox.warning(self, "Ошибка", "Сначала войдите в журнал.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb   = load_workbook(path)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                QMessageBox.warning(self, "Ошибка", "Файл пустой.")
                return

            row0 = rows[0]  # даты или заголовки
            row1 = rows[1]  # '1 час'/'2 час' или данные

            # Определяем формат: колледжа или старый
            is_college_format = (
                row1 and len(row1) > 2 and
                any(str(v).strip() in ("1 час","2 час") for v in row1 if v)
            )

            added_students  = 0
            updated_records = 0

            if is_college_format:
                # Строим карту колонок: col_index -> (date_str, hour)
                col_map = {}  # col_idx -> (date, hour)
                cur_date = None
                for ci, val in enumerate(row0):
                    if val and str(val).strip() and str(val).strip() != "ФИО обучающегося":
                        try:
                            # Может быть datetime объект или строка
                            if hasattr(val, 'strftime'):
                                cur_date = val.strftime("%d.%m.%Y")
                            else:
                                cur_date = str(val).strip()
                        except Exception:
                            cur_date = str(val).strip()
                    if cur_date and row1 and ci < len(row1):
                        hour_str = str(row1[ci]).strip() if row1[ci] else ""
                        if hour_str == "1 час":
                            col_map[ci] = (cur_date, 1)
                        elif hour_str == "2 час":
                            col_map[ci] = (cur_date, 2)

                # Строим или находим занятия для каждой даты+час
                lesson_map = {}  # (date, hour) -> lesson
                for l in self.book.lessons:
                    if l.type == "Лекция" and getattr(l,'hour',0):
                        lesson_map[(l.date, l.hour)] = l

                # Создаём отсутствующие занятия
                dates_seen = {}
                for ci, (date, hour) in col_map.items():
                    if (date, hour) not in lesson_map:
                        if date not in dates_seen:
                            dates_seen[date] = True
                            if hour == 1:
                                # Создаём оба часа
                                l_new = self.book.add_lesson("Лекция", topic="", date=date)
                                # add_lesson creates both hours
                                for l in self.book.lessons:
                                    if l.type=="Лекция" and l.date==date and getattr(l,'hour',0):
                                        lesson_map[(l.date, l.hour)] = l

                # Обрабатываем студентов (строки с 3-й по предпоследнюю)
                for row in rows[2:]:
                    if not row or not row[1]:
                        continue
                    fio = str(row[1]).strip()
                    if not fio or fio == "ФИО обучающегося":
                        continue
                    # Разделяем "Фамилия И.О." -> фамилия, инициалы
                    parts = fio.split()
                    if not parts:
                        continue
                    surname = parts[0]
                    initials = " ".join(parts[1:]) if len(parts) > 1 else ""

                    # Находим или создаём студента
                    student = next(
                        (s for s in self.book.spisok_stud
                         if s.f.lower() == surname.lower()), None
                    )
                    if not student:
                        student = Student(initials, surname, self.book.group)
                        self.book.spisok_stud.append(student)
                        added_students += 1
                        # Добавляем в secure_store
                        try:
                            ss_list = secure_store.get_students()
                            if not any(s.get("surname","").lower()==surname.lower()
                                       for s in ss_list):
                                ss_list.append({
                                    "name": initials,
                                    "surname": surname,
                                    "group": self.book.group
                                })
                                secure_store.set_students(ss_list)
                        except Exception:
                            pass

                    # Записываем посещаемость
                    for ci, (date, hour) in col_map.items():
                        if ci >= len(row):
                            continue
                        val = row[ci]
                        if val is None:
                            val = "✓"  # пустая ячейка = присутствовал
                        lesson = lesson_map.get((date, hour))
                        if lesson:
                            student.records[lesson.id] = str(val).strip()
                            updated_records += 1

            else:
                # Старый формат: столбец 0=фамилия, 1=имя, 2+=оценки
                header = row0
                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    surname = str(row[0]).strip()
                    name    = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    if not surname:
                        continue
                    student = next(
                        (s for s in self.book.spisok_stud
                         if s.f==surname and s.n==name), None
                    )
                    if not student:
                        student = Student(name, surname, self.book.group)
                        self.book.spisok_stud.append(student)
                        added_students += 1
                    for ci, val in enumerate(row[2:], start=2):
                        if ci - 2 >= len(self.book.lessons):
                            break
                        lesson = self.book.lessons[ci - 2]
                        if val is not None and str(val).strip():
                            student.records[lesson.id] = str(val).strip()
                            updated_records += 1

            self.book.save_to_sqlite()
            self._update_teacher_table()
            QMessageBox.information(
                self, "Импорт завершён",
                f"Добавлено студентов: {added_students}\n"
                f"Обновлено записей: {updated_records}\n"
                f"{'(Формат: журнал колледжа)' if is_college_format else '(Формат: стандартный)'}"
            )
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "Установите: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка импорта", str(e))

    def _import_sqlite(self):
        if not self.book:
            QMessageBox.warning(self, "Ошибка", "Сначала войдите в журнал.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Открыть SQLite", "", "SQLite (*.db *.sqlite *.sqlite3)"
        )
        if not path:
            return
        try:
            conn = sqlite3.connect(path)
            cur  = conn.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [r[0] for r in cur.fetchall()]
            if not tables:
                QMessageBox.warning(self, "Ошибка", "Нет таблиц в базе данных.")
                conn.close()
                return

            tbl, ok = QInputDialog.getItem(
                self, "Таблица", "Выберите таблицу студентов:", tables, 0, False
            )
            if not ok:
                conn.close()
                return

            cur.execute(f"PRAGMA table_info({tbl})")
            columns = [c[1] for c in cur.fetchall()]
            cur.execute(f"SELECT * FROM {tbl}")
            ext_rows = cur.fetchall()
            conn.close()

            if not ext_rows:
                QMessageBox.warning(self, "Ошибка", "Таблица пустая.")
                return

            # Показываем пользователю колонки и просим выбрать вручную
            # (автоопределение ненадёжно — слово "name" может быть в group_name)
            col_labels = [f"{i}: {c}" for i, c in enumerate(columns)]

            sc, ok1 = QInputDialog.getItem(
                self, "Колонка «Фамилия»",
                f"Колонки таблицы «{tbl}»:\nВыберите колонку с ФАМИЛИЕЙ:",
                col_labels, 0, False
            )
            if not ok1: return
            s_col = col_labels.index(sc)

            nc, ok2 = QInputDialog.getItem(
                self, "Колонка «Имя»",
                f"Колонки таблицы «{tbl}»:\nВыберите колонку с ИМЕНЕМ:",
                col_labels, 0, False
            )
            if not ok2: return
            n_col = col_labels.index(nc)

            # Колонка группы — необязательна
            col_labels_opt = ["(не использовать — взять текущую группу)"] + col_labels
            gc, ok3 = QInputDialog.getItem(
                self, "Колонка «Группа»",
                "Выберите колонку с ГРУППОЙ (или пропустите):",
                col_labels_opt, 0, False
            )
            if not ok3: return
            g_col = None
            if gc != col_labels_opt[0]:
                g_col = col_labels_opt.index(gc) - 1  # -1 т.к. первый элемент — «пропустить»

            # Импортируем студентов
            count        = 0
            skipped      = 0
            store_stud   = secure_store.get_students()

            for row in ext_rows:
                raw_surname = row[s_col]
                raw_name    = row[n_col]

                surname = str(raw_surname).strip() if raw_surname is not None else ""
                name    = str(raw_name).strip()    if raw_name    is not None else ""

                if not surname or not name:
                    skipped += 1
                    continue

                # Группа: из колонки или текущая группа журнала
                if g_col is not None and row[g_col]:
                    group = str(row[g_col]).strip()
                else:
                    group = self.book.group

                # Student(n=имя, f=фамилия, group=группа)
                already_in_book = any(
                    s.f.lower() == surname.lower() and s.n.lower() == name.lower()
                    for s in self.book.spisok_stud
                )
                if not already_in_book:
                    self.book.add_student(Student(name, surname, self.book.group))
                    count += 1

                # Добавляем в общую базу secure_store
                already_in_store = any(
                    s.get("surname", "").lower() == surname.lower()
                    and s.get("name", "").lower() == name.lower()
                    for s in store_stud
                )
                if not already_in_store:
                    store_stud.append({
                        "name":    name,
                        "surname": surname,
                        "group":   group
                    })

            secure_store.set_students(store_stud)
            self._update_teacher_table()
            msg = f"Импортировано в журнал: {count} студентов."
            if skipped:
                msg += f"\nПропущено пустых строк: {skipped}."
            QMessageBox.information(self, "Импорт SQLite", msg)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка импорта", str(e))

    def _update_teacher_table(self):
        students = self.book.spisok_stud
        lessons  = self.book.lessons

        # Колонки: для экзаменов с пересдачей — доп. колонка
        col_defs = []  # (lesson, is_retake_col)
        for l in lessons:
            col_defs.append((l, False))
            if l.type == "Экзамен" and getattr(l, 'retake_date', ''):
                col_defs.append((l, True))

        self.t_table.setRowCount(len(students))
        self.t_table.setColumnCount(2 + len(col_defs))

        headers = ["Фамилия", "Имя"]
        for l, is_retake in col_defs:
            if is_retake:
                headers.append(f"Пересдача\n{l.retake_date}")
            elif l.type == "Лекция":
                hour_label = f" ({l.hour}-й ч.)" if getattr(l,'hour',0) else ""
                headers.append(f"Лекция {l.number}{hour_label}\n{l.date}\n{l.topic}")
            elif l.type == "Экзамен":
                headers.append(f"Экзамен №{l.number}\n{l.date}\n{l.topic}")
            else:
                headers.append(f"Практика {l.number}\n{l.date}\n{l.topic}")
        self.t_table.setHorizontalHeaderLabels(headers)

        for r, s in enumerate(students):
            # ФИО с возможностью двойного клика для редактирования
            f_item = QTableWidgetItem(s.f)
            n_item = QTableWidgetItem(s.n)
            self.t_table.setItem(r, 0, f_item)
            self.t_table.setItem(r, 1, n_item)

            for ci, (l, is_retake) in enumerate(col_defs):
                col = 2 + ci

                if is_retake:
                    main_val = s.records.get(l.id, "")
                    if "Не зачтено" in main_val:
                        container = QWidget()
                        vl = QVBoxLayout(container)
                        vl.setContentsMargins(2, 2, 2, 2)
                        vl.setSpacing(1)
                        combo = QComboBox()
                        combo.addItems(["", "2", "3", "4", "5", "Н"])
                        combo.setStyleSheet("background-color:#3d2a10;")
                        saved_retake = s.records.get(l.id + "_retake", "")
                        raw_rt = saved_retake.split(" ")[0] if saved_retake else ""
                        combo.setCurrentText(raw_rt)
                        status_lbl = QLabel()
                        status_lbl.setAlignment(Qt.AlignCenter)
                        if "Зачтено" in saved_retake and "Не" not in saved_retake:
                            status_lbl.setText("✅ Зачтено")
                            status_lbl.setStyleSheet("font-size:10px;color:#7dd87d;")
                        elif "Не зачтено" in saved_retake:
                            status_lbl.setText("❌ Не зачтено")
                            status_lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                        vl.addWidget(combo)
                        vl.addWidget(status_lbl)
                        def _save_retake(val, st=s, le=l, lbl=status_lbl):
                            if val in ("4","5"):
                                st.records[le.id+"_retake"] = val+" (Зачтено)"
                                lbl.setText("✅ Зачтено"); lbl.setStyleSheet("font-size:10px;color:#7dd87d;")
                            elif val in ("2","3"):
                                st.records[le.id+"_retake"] = val+" (Не зачтено)"
                                lbl.setText("❌ Не зачтено"); lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                            elif val == "Н":
                                st.records[le.id+"_retake"] = "Н (Не зачтено)"
                                lbl.setText("❌ Не зачтено"); lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                            else:
                                st.records[le.id+"_retake"] = val; lbl.setText("")
                            self.book.save_to_sqlite()
                        combo.currentTextChanged.connect(_save_retake)
                        self.t_table.setCellWidget(r, col, container)
                    else:
                        item = QTableWidgetItem("—")
                        item.setForeground(QColor(80,80,80))
                        self.t_table.setItem(r, col, item)

                elif l.type == "Экзамен":
                    container = QWidget()
                    vl = QVBoxLayout(container)
                    vl.setContentsMargins(2,2,2,2); vl.setSpacing(1)
                    combo = QComboBox()
                    combo.addItems(["","2","3","4","5","Н"])
                    combo.setStyleSheet("background-color:#2a2a10;")
                    saved = s.records.get(l.id,"")
                    raw_grade = saved.split(" ")[0] if saved else ""
                    combo.setCurrentText(raw_grade)
                    status_lbl = QLabel()
                    status_lbl.setAlignment(Qt.AlignCenter)
                    if "Зачтено" in saved and "Не" not in saved:
                        status_lbl.setText("✅ Зачтено"); status_lbl.setStyleSheet("font-size:10px;color:#7dd87d;")
                    elif "Не зачтено" in saved:
                        status_lbl.setText("❌ Не зачтено"); status_lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                    vl.addWidget(combo); vl.addWidget(status_lbl)
                    def _save_exam(val, st=s, le=l, lbl=status_lbl):
                        if val in ("4","5"):
                            full=val+" (Зачтено)"; lbl.setText("✅ Зачтено"); lbl.setStyleSheet("font-size:10px;color:#7dd87d;")
                        elif val in ("2","3"):
                            full=val+" (Не зачтено)"; lbl.setText("❌ Не зачтено"); lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                            if not getattr(le,'retake_date',''):
                                self._ask_retake_date(le)
                        elif val=="Н":
                            full="Н (Не зачтено)"; lbl.setText("❌ Не зачтено"); lbl.setStyleSheet("font-size:10px;color:#e87d7d;")
                            if not getattr(le,'retake_date',''):
                                self._ask_retake_date(le)
                        else:
                            full=val; lbl.setText("")
                        st.records[le.id]=full; self.book.save_to_sqlite()
                    combo.currentTextChanged.connect(_save_exam)
                    self.t_table.setCellWidget(r, col, container)

                elif l.type == "Практика":
                    combo = QComboBox()
                    combo.addItems(["","2","3","4","5","Н"])
                    combo.setCurrentText(s.records.get(l.id,""))
                    combo.setStyleSheet("background-color:#1a2a1a;")
                    def _save_pr(val, st=s, le=l):
                        st.records[le.id]=val; self.book.save_to_sqlite()
                    combo.currentTextChanged.connect(_save_pr)
                    self.t_table.setCellWidget(r, col, combo)

                else:  # Лекция
                    combo = QComboBox()
                    combo.addItems(["","О","Б","✓","Н"])
                    combo.setCurrentText(s.records.get(l.id,""))
                    combo.setStyleSheet("background-color:#2a1a1a;")
                    def _save_lk(val, st=s, le=l):
                        st.records[le.id]=val; self.book.save_to_sqlite()
                    combo.currentTextChanged.connect(_save_lk)
                    self.t_table.setCellWidget(r, col, combo)

        self.t_table.resizeColumnsToContents()
        # Включаем редактирование ФИО (двойной клик на ячейку)
        self.t_table.itemChanged.connect(self._on_table_item_changed)
    def _header_context_menu(self, pos):
        col = self.t_table.horizontalHeader().logicalIndexAt(pos)
        if col < 2:
            return
        col_defs = []
        for l in self.book.lessons:
            col_defs.append((l, False))
            if l.type == "Экзамен" and getattr(l, 'retake_date', ''):
                col_defs.append((l, True))
        ci = col - 2
        if ci < 0 or ci >= len(col_defs):
            return
        lesson, is_retake = col_defs[ci]

        menu = QMenu()
        if is_retake:
            act_rd  = menu.addAction("📅 Изменить дату пересдачи")
            act_del = menu.addAction("🗑 Удалить столбец пересдачи")
            action  = menu.exec(self.t_table.mapToGlobal(pos))
            if action == act_rd:
                val, ok = QInputDialog.getText(
                    self, "Дата пересдачи", "Новая дата:", text=lesson.retake_date)
                if ok and val.strip():
                    self.book.set_retake_date(lesson.id, val.strip())
                    self._update_teacher_table()
            elif action == act_del:
                confirm = QMessageBox.question(self, "Удалить пересдачу",
                    "Удалить столбец пересдачи? Оценки за пересдачу будут удалены.",
                    QMessageBox.Yes | QMessageBox.No)
                if confirm == QMessageBox.Yes:
                    self.book.set_retake_date(lesson.id, "")
                    for s in self.book.spisok_stud:
                        s.records.pop(lesson.id + "_retake", None)
                    self.book.save_to_sqlite()
                    self._update_teacher_table()
        else:
            act_date   = menu.addAction("📅 Изменить дату")
            act_topic  = menu.addAction("✏ Изменить тему")
            act_retake = menu.addAction("🔄 Назначить/изменить пересдачу") if lesson.type == "Экзамен" else None
            act_del    = menu.addAction("🗑 Удалить это занятие")
            action = menu.exec(self.t_table.mapToGlobal(pos))
            if action == act_date:
                val, ok = QInputDialog.getText(
                    self, "Дата", "Новая дата:", text=lesson.date)
                if ok:
                    lesson.date = val
                    self.book.save_to_sqlite()
                    self._update_teacher_table()
            elif action == act_topic:
                val, ok = QInputDialog.getText(
                    self, "Тема", "Новая тема:", text=lesson.topic)
                if ok:
                    lesson.topic = val
                    self.book.save_to_sqlite()
                    self._update_teacher_table()
            elif act_retake and action == act_retake:
                self._ask_retake_date(lesson)
            elif action == act_del:
                # Для лекций — удаляем оба часа сразу
                to_delete = [lesson]
                if lesson.type == "Лекция" and getattr(lesson, 'hour', 0) in (1, 2):
                    partner = next((l for l in self.book.lessons
                                    if l.type=="Лекция" and l.number==lesson.number
                                    and l.id != lesson.id), None)
                    if partner:
                        to_delete.append(partner)
                names = ", ".join(
                    f"{l.type} №{l.number} ({l.date})" for l in to_delete
                )
                confirm = QMessageBox.question(self, "Удалить занятие",
                    f"Удалить: {names}?\nВсе оценки за это занятие будут удалены.",
                    QMessageBox.Yes | QMessageBox.No)
                if confirm == QMessageBox.Yes:
                    import sqlite3 as _sq
                    conn = _sq.connect(self.book.db_name)
                    cur  = conn.cursor()
                    for dl in to_delete:
                        cur.execute("DELETE FROM lessons WHERE id=?", (dl.id,))
                        cur.execute("DELETE FROM grades WHERE lesson_id=?", (dl.id,))
                        for s in self.book.spisok_stud:
                            s.records.pop(dl.id, None)
                    conn.commit(); conn.close()
                    self.book.lessons = [l for l in self.book.lessons
                                         if l not in to_delete]
                    self._update_teacher_table()

    def _on_table_item_changed(self, item):
        """Редактирование фамилии/имени студента двойным кликом."""
        row = item.row()
        col = item.column()
        if col not in (0, 1):
            return
        if row >= len(self.book.spisok_stud):
            return
        student = self.book.spisok_stud[row]
        new_val = item.text().strip()
        if not new_val:
            return
        import sqlite3 as _sq
        conn = _sq.connect(self.book.db_name)
        cur  = conn.cursor()
        if col == 0:  # Фамилия
            old_f = student.f
            cur.execute("UPDATE students SET f=? WHERE f=? AND n=?", (new_val, old_f, student.n))
            cur.execute("UPDATE grades SET student_f=? WHERE student_f=? AND student_n=?",
                        (new_val, old_f, student.n))
            student.f = new_val
        elif col == 1:  # Имя
            old_n = student.n
            cur.execute("UPDATE students SET n=? WHERE f=? AND n=?", (new_val, student.f, old_n))
            cur.execute("UPDATE grades SET student_n=? WHERE student_f=? AND student_n=?",
                        (new_val, student.f, old_n))
            student.n = new_val
        conn.commit(); conn.close()
        # Обновляем в secure_store тоже
        try:
            store_studs = secure_store.get_students()
            for ss in store_studs:
                if col == 1 and ss.get("surname","") == old_f and ss.get("name","") == student.n:
                    ss["surname"] = new_val
                elif col == 2 and ss.get("surname","") == student.f and ss.get("name","") == old_n:
                    ss["name"] = new_val
            secure_store.set_students(store_studs)
        except Exception:
            pass


    # ──────────────────────────────────────────────────────────
    #  ОБРАБОТЧИКИ ВХОДА
    # ──────────────────────────────────────────────────────────
    def _login_student(self):
        n = self.sn.text().strip()
        f = self.sf.text().strip()
        if not n or not f:
            QMessageBox.warning(self, "Ошибка", "Введите имя и фамилию.")
            return

        # Ищем студента в secure_store по имени и фамилии
        all_students = secure_store.get_students()
        matched = [s for s in all_students
                   if s.get("name", "").lower() == n.lower()
                   and s.get("surname", "").lower() == f.lower()]

        if matched:
            # Берём первое совпадение
            stud_data = matched[0]
            group = stud_data.get("group", "")
            if not group:
                QMessageBox.warning(self, "Группа не назначена",
                    "Вам ещё не назначена группа. Обратитесь к администратору.")
                return
            self.cur_stud = {"n": n, "f": f, "g": group}
        else:
            # Студента нет в базе — входим без группы (режим гостя)
            # Спрашиваем группу как запасной вариант
            groups = get_groups()
            if not groups:
                QMessageBox.warning(self, "Ошибка",
                    "Студент не найден в базе данных. Обратитесь к администратору.")
                return
            g, ok = QInputDialog.getItem(
                self, "Группа не найдена",
                f"Студент «{f} {n}» не найден в базе.\n"
                "Выберите группу вручную или обратитесь к администратору:",
                groups, 0, False
            )
            if not ok:
                return
            self.cur_stud = {"n": n, "f": f, "g": g}

        self._draw_student_dash()
        self.stack.setCurrentWidget(self.student_widget)

    def _login_teacher(self):
        name     = self.tn.text().strip()
        password = self.tp.text()

        # Новый формат (secure_store)
        teachers = secure_store.get_teachers()
        if name in teachers:
            if password != teachers[name].get("password", ""):
                QMessageBox.warning(self, "Ошибка", "Неверный пароль.")
                return
            subjects = teachers[name].get("subjects", [])
            if not subjects:
                QMessageBox.warning(self, "Ошибка", "У преподавателя нет назначенных предметов.")
                return
            subj, ok = QInputDialog.getItem(
                self, "Предмет", "Выберите предмет:", subjects, 0, False
            )
            if not ok:
                return
            # Только группы у которых этот предмет есть в программе
            groups_with_subj = [
                g["name"] for g in secure_store.get_groups()
                if subj in g.get("subjects", [])
            ]
            if not groups_with_subj:
                groups_with_subj = get_groups()  # fallback
            g, ok2 = QInputDialog.getItem(
                self, "Группа", f"Предмет: {subj}\nВыберите группу:", groups_with_subj, 0, False
            )
            if ok2:
                self.cur_teacher_name = name
                self.cur_teacher_subj = subj
                self.book = GradeBook(g, subj)
                self._teacher_title_lbl.setText(
                    f"Журнал: {name}  •  {subj}  •  {g}"
                )
                self._update_teacher_table()
                self.stack.setCurrentWidget(self.teacher_widget)
            return

        # Обратная совместимость (logins.txt)
        old = self.teachers_db
        if name in old and password == self.teacher_pass:
            data = old[name]
            subj = data.get("subjects", [""])[0] if isinstance(data, dict) else str(data)
            groups_with_subj = [
                g["name"] for g in secure_store.get_groups()
                if subj in g.get("subjects", [])
            ] or get_groups()
            g, ok = QInputDialog.getItem(
                self, "Группа", f"Предмет: {subj}\nВыберите группу:", groups_with_subj, 0, False
            )
            if ok:
                self.cur_teacher_name = name
                self.cur_teacher_subj = subj
                self.book = GradeBook(g, subj)
                self._teacher_title_lbl.setText(f"Журнал: {name}  •  {subj}  •  {g}")
                self._update_teacher_table()
                self.stack.setCurrentWidget(self.teacher_widget)
        else:
            QMessageBox.warning(self, "Ошибка", "Неверный логин или пароль.")
