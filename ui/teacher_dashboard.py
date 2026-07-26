"""
teacher_dashboard.py — TeacherDashboard
Часть рефакторинга GUI.py → модульная архитектура
"""

import ui_date
import log

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView, QCheckBox, QComboBox, QDialog, QDialogButtonBox, QFileDialog,
    QHeaderView, QHBoxLayout, QInputDialog, QMessageBox,
    QPushButton, QScrollArea, QStackedWidget,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QFrame
)

from styles import C, BTN
from widgets import (
    FlowLayout,
    lbl, title_lbl, section_lbl, btn, combo, card, stat_card, field_cell,
    vector_unavailable_widget
)
from ui_components import Sidebar
from utils import get_groups

from core import GradeBook, Student
from data_store import get_store


class TeacherDashboard(QWidget):
    def __init__(self, teacher_name: str, teacher_data: dict, parent=None):
        super().__init__(parent)
        self.teacher_name = teacher_name
        self.teacher_data = teacher_data
        self.book = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(0, 0, 0, 0)
        items = [
            ("__label__", "", "Журнал"),
            ("journal",  "clipboard", "Журнал"),
            ("students", "users",     "Студенты"),
        ]
        #Вкладка куратора появляется, ТОЛЬКО если админ назначил группы. Список едет
        #синком в users.curated_groups, поэтому куратор работает и офлайн — как весь
        #остальной десктоп (инвариант §1), без похода на сервер.
        if self._curated_groups():
            items.append(("curator", "check", "Курирование"))
        items += [
            ("schedule", "calendar",  "Расписание"),
            ("stats",    "chart",     "Статистика"),
            ("ai",       "bot",       "ИИ Помощник"),
            ("messages", "send",      "Сообщения"),
            ("__label__", "", "Личное"),
            ("profile",  "user",      "Профиль"),
            ("settings", "settings",  "Настройки"),
        ]
        self.sidebar = Sidebar(items)
        self.sidebar.tab_clicked.connect(self._switch)
        self.stack = QStackedWidget()
        self.pages = {}
        body = QHBoxLayout(); body.setContentsMargins(0, 0, 0, 0); body.setSpacing(0)
        body.addWidget(self.sidebar); body.addWidget(self.stack, 1)
        lay.addLayout(body)
        self._build_journal()
        self._build_students()
        if self._curated_groups():
            self._build_curator()
        self._build_schedule()
        self._build_stats()
        self._build_ai()
        self._build_messenger()
        self._build_profile()
        self._build_settings()
        self.sidebar.set_active("journal")
        self._init_selectors()

        #Вектор постоянно слева (⇄ — вправо, — свернуть/вернуть 🐯)
        try:
            from vector.widget import VectorPanel, VectorHost
            self._ensure_vector_session()
            self.vector_dock = VectorHost(
                body, VectorPanel(self.vector_session, docked=True))
            self.vector_dock.mount(side="left")
        except Exception as _e:
            log.get("teacher_dashboard").warning(f"[Vector] панель сбоку (препод): {_e}")

    def _ensure_vector_session(self):
        """Создаёт ОБЩУЮ сессию Вектора (одна история для шторки и вкладки). Движок
        собирается один раз; шторка и вкладка «ИИ Помощник» делят эту сессию."""
        if getattr(self, "vector_session", None) is not None:
            return self.vector_session
        from vector import VectorEngine, VectorScope, get_provider
        from vector.widget import VectorSession
        try:
            from data_store import get_store as _gs
            _cfg0 = _gs()._config()
        except Exception:
            _cfg0 = {}
        #Свои группы преподавателя — из назначений предмет→группа (а не весь колледж),
        #чтобы Вектор на вопрос «какие у меня группы» отвечал по этому списку.
        ga = self.teacher_data.get("group_assignments", {}) or {}
        my_groups = sorted({g for g in ga.values() if g})
        eng = VectorEngine(VectorScope(
            role="teacher", group=self._group_combo.currentText(),
            subject=self._subj_combo.currentText() or None,
            teacher_groups=my_groups), get_provider(_cfg0))
        self.vector_engine = eng
        self.vector_session = VectorSession(eng)
        return self.vector_session

    #Журнал

    def _build_journal(self):
        w = QWidget()
        lay = QVBoxLayout(w); lay.setContentsMargins(20, 18, 20, 18); lay.setSpacing(10)
        #Заголовок
        #Заголовок ОТДЕЛЬНОЙ строкой, а фильтры — в переносящемся ряду ниже. Раньше всё
        #лежало в одном QHBoxLayout, и в неполноэкранном окне выпадающие списки
        #сжимались до нечитаемых огрызков.
        lay.addWidget(title_lbl("Журнал преподавателя", 20))
        hdr = FlowLayout(h_spacing=8, v_spacing=6)
        self._subj_combo = combo(self.teacher_data.get("subjects", []))
        self._subj_combo.currentTextChanged.connect(self._on_subj_change)
        raw_groups = get_groups()
        group_names = [g["name"] if isinstance(g, dict) else str(g) for g in raw_groups]
        self._group_combo = combo(group_names)
        self._group_combo.currentTextChanged.connect(self._reload_journal)
        #Селектор учебного семестра (как на вебе): текущий по умолчанию, прошлые — архив.
        self._term_combo = combo([])
        self._term_combo.currentIndexChanged.connect(self._on_term_change)
        #Каждая пара «подпись + список» — единая ячейка одинаковой ширины (field_cell):
        #в переносящемся FlowLayout подпись не отрывается от поля, а поля при любом размере
        #окна стоят ровными параллельными колонками (см. §недочёт выравнивания).
        hdr.addWidget(field_cell("Предмет:", self._subj_combo))
        hdr.addWidget(field_cell("Группа:",  self._group_combo))
        hdr.addWidget(field_cell("Семестр:", self._term_combo))
        #Тумблер «Н = 2»: визуально пересчитывает столбец «Средний» (не трогая config).
        self._absence_chk = QCheckBox("Н = 2 в среднем")
        self._absence_chk.setChecked(True)
        self._absence_chk.setToolTip("Снимите галочку, чтобы пропуски «Н» не занижали средний балл")
        self._absence_chk.setStyleSheet(f"QCheckBox{{color:{C['text2']};font-size:12px;}}")
        self._absence_chk.stateChanged.connect(lambda: self._update_table())
        hdr.addWidget(self._absence_chk)
        lay.addLayout(hdr)
        #Кнопки. Изменяющие журнал кнопки блокируются на архивном (прошлом) семестре.
        #Девять кнопок в QHBoxLayout не помещались в неполноэкранное окно: layout их
        #СЖИМАЛ, и часть уезжала за край — до них было просто не добраться. FlowLayout
        #переносит на новую строку.
        btn_row = FlowLayout(h_spacing=6, v_spacing=6)
        self._edit_buttons = []
        _EDIT_LABELS = {"+ Лекция/Практика/ДЗ", "+ Экзамен", "+ Студент",
                        "💾 Сохранить", "🎓 Аттестация", "📥 Импорт"}
        for txt, style, cb in [
            ("+ Лекция/Практика/ДЗ", "green",  self._add_lesson),
            ("+ Экзамен",         "blue",   self._add_exam),
            ("+ Студент",         "ghost",  self._add_student),
            ("💾 Сохранить",      "ghost",  self._save),
            ("🎓 Аттестация",     "ghost",  self._open_attestation),
            ("📄 Ведомость",      "ghost",  self._export_vedomost),
            ("📘 Журнал",         "ghost",  self._export_journal),
            ("📥 Импорт",         "ghost",  self._import_excel),
        ]:
            b = btn(txt, style); b.clicked.connect(cb); btn_row.addWidget(b)
            if txt in _EDIT_LABELS:
                self._edit_buttons.append(b)
        #индикатор конфликтов синхронизации (виден, только если они есть)
        self._conflict_btn = btn("⚠ Конфликты", "blue")
        self._conflict_btn.clicked.connect(self._open_conflicts)
        self._conflict_btn.hide()
        btn_row.addWidget(self._conflict_btn)
        lay.addLayout(btn_row)
        self._refresh_conflicts_badge()
        #Плашка архива (виден только на прошлом семестре — журнал только для просмотра).
        self._archive_lbl = lbl("", 12, C['yellow'])
        self._archive_lbl.setWordWrap(True)
        self._archive_lbl.hide()
        lay.addWidget(self._archive_lbl)
        #«Пройдено X из Y ч» по открытому предмету. Скрыт, пока админ не задал план —
        #показывать «пройдено 24 из 0» хуже, чем не показывать ничего.
        self._hours_lbl = lbl("", 12, C['text3'])
        self._hours_lbl.hide()
        lay.addWidget(self._hours_lbl)
        #Таблица
        self.t_table = QTableWidget()
        self.t_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.t_table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
        self.t_table.horizontalHeader().customContextMenuRequested.connect(self._header_ctx)
        self.t_table.setWordWrap(True)
        #Журнал должен ЧИТАТЬСЯ: крупный текст, высокие строки, зебра и заметная шапка —
        #раньше таблица была мелкой и сливалась с фоном.
        self.t_table.setAlternatingRowColors(True)
        self.t_table.verticalHeader().setDefaultSectionSize(44)
        self.t_table.setStyleSheet(
            f"QTableWidget{{font-size:14px;alternate-background-color:{C['bg2']};"
            f"background:{C['card']};gridline-color:{C['border']};}}"
            f"QHeaderView::section{{font-size:12px;font-weight:700;color:{C['text']};"
            f"background:{C['bg2']};border:none;border-bottom:2px solid {C['green']};"
            "padding:6px 8px;}")
        self.t_table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.t_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        lay.addWidget(self.t_table, 1)
        self.pages["journal"] = w; self.stack.addWidget(w)

    def _init_selectors(self):
        subjects = self.teacher_data.get("subjects", [])
        self._subj_combo.clear(); self._subj_combo.addItems(subjects)
        if subjects:
            ga  = self.teacher_data.get("group_assignments", {})
            grp = ga.get(subjects[0], "")
            #get_groups() может вернуть строки или dict — нормализуем
            raw_groups = get_groups()
            group_names = [
                g["name"] if isinstance(g, dict) else str(g)
                for g in raw_groups
            ]
            self._group_combo.clear()
            self._group_combo.addItems(group_names)
            if grp:
                idx = self._group_combo.findText(grp)
                if idx >= 0: self._group_combo.setCurrentIndex(idx)
        self._reload_journal()

    def _on_subj_change(self):
        subj = self._subj_combo.currentText()
        ga   = self.teacher_data.get("group_assignments", {})
        grp  = ga.get(subj, "")
        if grp:
            idx = self._group_combo.findText(grp)
            if idx >= 0: self._group_combo.setCurrentIndex(idx)
        self._reload_journal()

    #Учебные семестры (как на вебе): текущий — редактируемый, прошлые — архив read-only.
    def _cur_term(self):
        import terms
        if not getattr(self, "_current_term", None):
            self._current_term = terms.current_term()
        return self._current_term

    def _term_label(self, t) -> str:
        import terms
        y, s = t
        return f"{y} · {terms.semester_label(s)} семестр"

    def _populate_terms(self):
        """Наполняет комбо семестров (локальные периоды + текущий), сохраняя выбор."""
        from core import DBManager
        cur = self._cur_term()
        prev = self._selected_term() if getattr(self, "_term_items", None) else cur
        items = DBManager.list_terms()
        if cur not in items:
            items = [cur] + items
        self._term_items = items
        self._term_combo.blockSignals(True)
        self._term_combo.clear()
        self._term_combo.addItems([self._term_label(t) for t in items])
        idx = items.index(prev) if prev in items else (items.index(cur) if cur in items else 0)
        self._term_combo.setCurrentIndex(idx)
        self._term_combo.blockSignals(False)

    def _selected_term(self):
        items = getattr(self, "_term_items", [])
        i = self._term_combo.currentIndex()
        return items[i] if 0 <= i < len(items) else self._cur_term()

    def _is_archive(self) -> bool:
        return self._selected_term() != self._cur_term()

    def _on_term_change(self):
        self._reload_journal()

    def _reload_journal(self):
        subj  = self._subj_combo.currentText()
        group = self._group_combo.currentText()
        if not subj or not group: return
        self._populate_terms()                       #подхватить новые периоды, сохранив выбор
        year, sem = self._selected_term()
        self.book = GradeBook(group, subj, year, sem)
        #Архив (прошлый семестр) — только просмотр: гасим кнопки правки и показываем плашку.
        archive = self._is_archive()
        for b in getattr(self, "_edit_buttons", []):
            b.setEnabled(not archive)
        if archive:
            self._archive_lbl.setText(
                "🔒 Прошлый семестр — архив (только просмотр). Правки доступны в текущем "
                f"периоде: {self._term_label(self._cur_term())}.")
            self._archive_lbl.show()
        else:
            self._archive_lbl.hide()
        self._update_hours_label()
        #Вектор отвечает по той группе/предмету, что открыты сейчас
        if getattr(self, "vector_engine", None):
            self.vector_engine.scope.group = group
            self.vector_engine.scope.subject = subj
        self._sync_students_from_store()
        self._update_table()

    def _update_hours_label(self):
        """Обновить строку «Пройдено X из Y ч» под шапкой журнала."""
        import study_hours
        text = ""
        if self.book:
            try:
                done, total = self.book.hours_progress()
                text = study_hours.format_progress(done, total)
            except Exception:
                text = ""      #часы — справочная строка, ронять из-за неё журнал нельзя
        self._hours_lbl.setText(text)
        self._hours_lbl.setVisible(bool(text))

    def _sync_students_from_store(self):
        """Подтянуть студентов из общего хранилища и добавить недостающих в журнал."""
        if not self.book: return
        store = get_store()
        if not store: return
        try:
            known_students = store.get_students()
            group = self.book.group
            existing = {s.f.lower() for s in self.book.spisok_stud}
            added = False
            for s in known_students:
                if s.get("group", "") != group: continue
                surname = s.get("surname", "")
                if surname.lower() not in existing:
                    st = Student(s.get("name", ""), surname, group)
                    self.book.add_student(st)
                    existing.add(surname.lower())
                    added = True
            if added:
                self.book.save_to_db()
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[sync students] {e}")

    def _avg_cfg(self) -> dict:
        """config с учётом тумблера «Н = 2» — визуальный пересчёт среднего, не сохраняем."""
        try:
            base = dict(get_store()._config() or {})
        except Exception:
            base = {}
        chk = getattr(self, "_absence_chk", None)
        base["avg_count_absence"] = chk.isChecked() if chk is not None else True
        return base

    @staticmethod
    def _avg_color(avg: float) -> str:
        if avg >= 4.5: return C['green']
        if avg >= 3.5: return C['blue']
        if avg >= 2.5: return C['yellow']
        if avg > 0:    return C['red']
        return C['text3']

    def _update_table(self):
        if not self.book: return
        #Часы пересчитываем здесь: сюда сходятся все пути правки журнала (добавили
        #занятие, удалили, переключили семестр), и счётчик не отстанет от таблицы.
        self._update_hours_label()
        students = self.book.spisok_stud
        col_defs = []
        for l in self.book.lessons:
            col_defs.append((l, 0))
            if l.type == "Экзамен":
                ri = 1
                while getattr(l, f'retake_date{"" if ri == 1 else "_" + str(ri)}', ''):
                    col_defs.append((l, ri)); ri += 1
        avg_col = 2 + len(col_defs)                     #«Средний»
        abs_col = avg_col + 1                            #«Пропусков (ч)» — правее среднего
        self.t_table.setRowCount(len(students))
        self.t_table.setColumnCount(abs_col + 1)
        #§12: счётчик пропущенных часов — та же логика, что у Вектора/сервера
        #(vector.intents._count_absences, webdata.absences): строка лекции с «Н/Б/О» —
        #1 час, у практики считается только «Н» (ДЗ не входит — это «не сдал», не «не был»).
        #Ленивый импорт — как и остальные обращения к vector.* в этом модуле (см. dashboards.py).
        from vector.intents import _count_absences as count_absences
        abs_lessons = [(l.id, l.type) for l in self.book.lessons]
        #Тему в шапке показываем коротко (чтобы столбец не разъезжался), но
        #полную тему кладём в подсказку — наведёшь мышь и прочитаешь целиком.
        def _short(text: str, limit: int = 22) -> str:
            text = (text or "").strip()
            if len(text) <= limit:
                return text
            cut = text[:limit].rstrip()
            sp = cut.rfind(" ")
            if sp >= limit - 8:          #обрезаем по последнему пробелу, не посреди слова
                cut = cut[:sp].rstrip()
            return cut + "…"
        headers = ["Фамилия", "Имя"]
        tooltips = ["", ""]
        for l, ri in col_defs:
            if ri > 0:
                rd = getattr(l, f'retake_date{"" if ri == 1 else "_" + str(ri)}', '')
                headers.append(f"Пересдача №{ri}\n{rd}"); tooltips.append("")
            elif l.type == "Лекция":
                headers.append(f"Лекция {l.number}{f' ({l.hour}-й ч.)' if l.hour else ''}\n{l.date}\n{_short(l.topic)}")
                tooltips.append(l.topic or "")
            elif l.type == "Экзамен":
                headers.append(f"Экзамен №{l.number}\n{l.date}\n{_short(l.topic)}")
                tooltips.append(l.topic or "")
            else:
                #Ветка «всё остальное» исторически подписывалась «Практика». С появлением
                #ДЗ (и семинаров/лабораторных с веба) это врало бы в заголовке — берём
                #настоящий тип занятия, а «Практика» остаётся дефолтом для пустого типа.
                headers.append(f"{l.type or 'Практика'} {l.number}\n{l.date}\n{_short(l.topic)}")
                tooltips.append(l.topic or "")
        headers.append("Средний"); tooltips.append("")
        headers.append("Пропусков (ч)")
        tooltips.append("Пропущено часов: «Н» неуважительно + «Б» болезнь + «О» уважительно")
        self.t_table.setHorizontalHeaderLabels(headers)
        #полная тема в подсказке заголовка
        for c, tip in enumerate(tooltips):
            it = self.t_table.horizontalHeaderItem(c)
            if it is not None and tip:
                it.setToolTip(tip)
        avg_cfg = self._avg_cfg()                       #считаем один раз на перерисовку
        self.t_table.blockSignals(True)
        for r, s in enumerate(students):
            fi = QTableWidgetItem(s.f); fi.setForeground(QColor(C['text']))
            ni = QTableWidgetItem(s.n); ni.setForeground(QColor(C['text3']))
            self.t_table.setItem(r, 0, fi); self.t_table.setItem(r, 1, ni)
            #Столбец «Средний» — с учётом тумблера «Н = 2» (визуально, не в config).
            avg = self.book.calculate_average(s, avg_cfg)
            ai = QTableWidgetItem(f"{avg:.2f}" if avg > 0 else "—")
            ai.setTextAlignment(Qt.AlignCenter); ai.setFlags(Qt.ItemIsEnabled)
            ai.setForeground(QColor(self._avg_color(avg)))
            af = ai.font(); af.setBold(True); af.setPointSize(af.pointSize() + 1); ai.setFont(af)
            self.t_table.setItem(r, avg_col, ai)
            #Столбец «Пропусков (ч)» — считаем по фактическим отметкам, не по col_defs
            #(там дублируются строки пересдач, которые пропусками не являются).
            absc = count_absences(abs_lessons, s.records)
            total = absc.get("всего", 0)
            bi = QTableWidgetItem(str(total) if total else "—")
            bi.setTextAlignment(Qt.AlignCenter); bi.setFlags(Qt.ItemIsEnabled)
            bi.setForeground(QColor(C['red'] if total else C['text3']))
            bi.setToolTip(f"Неуваж. (Н): {absc.get('Н', 0)} · болезнь (Б): {absc.get('Б', 0)} · "
                         f"уваж. (О): {absc.get('О', 0)}")
            self.t_table.setItem(r, abs_col, bi)
            for ci, (l, ri) in enumerate(col_defs):
                col = 2 + ci
                rk  = l.id + ("_retake" if ri == 1 else f"_retake_{ri}" if ri > 1 else "")
                val = s.records.get(rk if ri > 0 else l.id, "")
                if l.type == "Лекция" and ri == 0:
                    cb = QComboBox()
                    cb.addItems(["", "✓", "Н", "Б", "О"])
                    cb.setCurrentText(val)
                    cb.setStyleSheet(f"background:{C['card2']};border:1px solid {C['border']};border-radius:6px;color:{C['text']};font-size:14px;font-weight:600;padding:4px 6px;")
                    cb.currentTextChanged.connect(lambda v, st=s, k=l.id: self._set_val(st, k, v))
                    self.t_table.setCellWidget(r, col, cb)
                elif l.type == "Практика" and ri == 0:
                    cb = QComboBox()
                    cb.addItems(["", "2", "3", "4", "5", "Н"])
                    cb.setCurrentText(val)
                    cb.setStyleSheet(f"background:{C['card2']};border:1px solid {C['border']};border-radius:6px;color:{C['text']};font-size:14px;font-weight:600;padding:4px 6px;")
                    cb.currentTextChanged.connect(lambda v, st=s, k=l.id: self._set_val(st, k, v))
                    self.t_table.setCellWidget(r, col, cb)
                elif l.type == "Экзамен" or ri > 0:
                    rk_full = l.id + ("_retake" if ri == 1 else f"_retake_{ri}" if ri > 1 else "")
                    #ФИКС БАГА: пересдача назначается не всем.
                    #Ячейка пересдачи №ri активна только у студента, который
                    #ЗАВАЛИЛ предыдущую попытку (2 / Н / «Не зачтено»).
                    #Сдавшим — серый прочерк без редактирования.
                    if ri > 0 and not self._needs_retake(s, l, ri):
                        it = QTableWidgetItem("—")
                        it.setTextAlignment(Qt.AlignCenter)
                        it.setFlags(Qt.ItemIsEnabled)          #не редактируется
                        it.setForeground(QColor(C['text3']))
                        self.t_table.setItem(r, col, it)
                        continue
                    cb = QComboBox()
                    cb.addItems(["", "2", "3", "4", "5", "Н"])
                    raw = val.split()[0] if val else ""
                    cb.setCurrentText(raw)
                    cb.setStyleSheet(f"background:{C['card2']};border:1px solid {C['border']};border-radius:6px;color:{C['text']};font-size:14px;font-weight:600;padding:4px 6px;")
                    cb.currentTextChanged.connect(lambda v, st=s, le=l, rk=rk_full, ri_=ri: self._set_exam_val(st, le, rk, v, ri_))
                    self.t_table.setCellWidget(r, col, cb)
        self.t_table.blockSignals(False)
        #Архив (прошлый семестр) — только просмотр: гасим все редактируемые ячейки.
        if self._is_archive():
            for rr in range(self.t_table.rowCount()):
                for cc in range(2, self.t_table.columnCount()):
                    wcell = self.t_table.cellWidget(rr, cc)
                    if wcell is not None:
                        wcell.setEnabled(False)
        self.t_table.resizeColumnsToContents()   #ширины по содержимому (столбцы 0,1 — под ФИО)
        hh = self.t_table.horizontalHeader()
        hh.setMinimumSectionSize(110); hh.setDefaultAlignment(Qt.AlignCenter)
        #Фамилия/Имя: НЕ фиксируем жёстко (иначе длинные ФИО обрезались) — берём ширину по
        #содержимому, но не меньше комфортного минимума. Так все ФИО влезают целиком.
        self.t_table.setColumnWidth(0, max(self.t_table.columnWidth(0) + 16, 160))
        self.t_table.setColumnWidth(1, max(self.t_table.columnWidth(1) + 16, 120))
        for c in range(2, self.t_table.columnCount()):
            self.t_table.setColumnWidth(c, max(self.t_table.columnWidth(c), 112))

    def _set_val(self, student, key, val):
        student.records[key] = val
        self._persist_grade(student, key, val)

    def _persist_grade(self, student, key, val):
        """Пишем оценку в локальную БД СРАЗУ (offline-first: каждое действие → диск).

        Раньше оценка жила только в памяти до нажатия «Сохранить»; при переключении
        группы/предмета (журнал перезагружается из БД) или закрытии программы
        несохранённое терялось. Теперь правка не пропадёт. После записи будим
        фоновую синхронизацию, чтобы оценка ушла на сервер без ожидания интервала."""
        try:
            from core import DBManager
            conn = DBManager.get_conn()
            cur = conn.cursor()
            DBManager.upsert_grade(cur, (student.f, student.n, key, val))
            conn.commit()
            conn.close()
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[journal] не удалось сохранить оценку: {e}")
        try:
            from sync_runner import trigger
            trigger()
        except Exception:
            pass

    #конфликты синхронизации
    def _refresh_conflicts_badge(self):
        try:
            from core import DBManager
            n = len(DBManager.list_conflicts(unresolved_only=True))
            if hasattr(self, "_conflict_btn"):
                self._conflict_btn.setVisible(n > 0)
                if n > 0:
                    self._conflict_btn.setText(f"⚠ Конфликты ({n})")
        except Exception:
            pass

    def _open_conflicts(self):
        try:
            from conflict_dialog import ConflictDialog
            ConflictDialog(self).exec()
            self._refresh_conflicts_badge()
            self._update_table()
        except Exception as e:
            QMessageBox.warning(self, "Конфликты", f"Не удалось открыть: {e}")

    #хелперы пересдач — тонкие обёртки над ЕДИНЫМ источником grading (тот же код на
    #сервере и вебе, закреплён контрактом grade-cases.json).
    @staticmethod
    def _attempt_key(lesson, ri: int) -> str:
        import grading
        return grading.attempt_key(lesson.id, ri)

    @staticmethod
    def _is_failed(val: str) -> bool:
        import grading
        return grading.is_failed(val)

    def _needs_retake(self, student, lesson, ri: int) -> bool:
        import grading
        return grading.needs_retake(student.records, lesson.id, ri)

    def _set_exam_val(self, student, lesson, key, val, ri):
        if not val: return
        if val in ("4", "5"):
            full = val + " (Зачтено)"
        elif val == "3":
            dlg = QMessageBox(self)
            dlg.setWindowTitle("Оценка 3")
            dlg.setText(f"Оценка 3 у {student.f}. Засчитать или на пересдачу?")
            bp = dlg.addButton("✅ Зачёт", QMessageBox.AcceptRole)
            dlg.addButton("❌ Пересдача", QMessageBox.RejectRole)
            dlg.exec()
            if dlg.clickedButton() == bp:
                full = "3 (Зачтено)"
            else:
                full = "3 (Не зачтено)"
                rd = ui_date.ask_date(self, "Дата пересдачи", "Когда пересдача:",
                    default=ui_date.plus_days_str(7), min_today=True)
                if rd: self.book.set_retake_date(lesson.id, rd, ri + 1)
        elif val in ("2", "Н"):
            full = val + " (Не зачтено)"
            rd = ui_date.ask_date(self, "Дата пересдачи", "Когда пересдача:",
                default=ui_date.plus_days_str(7), min_today=True)
            if rd: self.book.set_retake_date(lesson.id, rd, ri + 1 if ri > 0 else 1)
        else:
            full = val
        student.records[key] = full
        self._persist_grade(student, key, full)   #сразу на диск (offline-first)
        self._update_table()

    def _header_ctx(self, pos):
        if self._is_archive():   #архивный семестр — правка/удаление занятий недоступны
            return
        col = self.t_table.horizontalHeader().logicalIndexAt(pos)
        if col < 2: return
        ci = col - 2
        col_defs = []
        for l in self.book.lessons:
            col_defs.append((l, 0))
            if l.type == "Экзамен":
                ri = 1
                while getattr(l, f'retake_date{"" if ri == 1 else "_" + str(ri)}', ''):
                    col_defs.append((l, ri)); ri += 1
        if ci >= len(col_defs): return
        l, ri = col_defs[ci]
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.setStyleSheet(
            f"QMenu{{background:{C['card2']};border:1px solid {C['border2']};border-radius:8px;"
            f"color:{C['text']};padding:4px;}}"
            f"QMenu::item{{padding:8px 16px;border-radius:6px;}}"
            f"QMenu::item:selected{{background:rgba(20,124,139,0.1);color:{C['green']};}}"
        )
        if ri == 0:
            e = menu.addAction("✎  Изменить тему / дату")
            e.triggered.connect(lambda: self._edit_lesson(l))
        if l.type == "Экзамен" and ri == 0:
            a = menu.addAction("📅 Назначить пересдачу")
            a.triggered.connect(lambda: self._ask_retake(l, 1))
        menu.addSeparator()
        d = menu.addAction("🗑  Удалить занятие")
        d.triggered.connect(lambda: self._delete_lesson(l.id))
        menu.exec(self.t_table.horizontalHeader().mapToGlobal(pos))

    def _edit_lesson(self, l):
        """Правка уже добавленного занятия через ПКМ: тема, дата, номер. Пишем в БД
        (offline-first) и будим синк — правка уедет на сервер и на другие ПК."""
        topic, ok = QInputDialog.getText(self, "Тема занятия",
                                          "Тема (что было на паре):", text=l.topic or "")
        if not ok:
            return
        date, ok = QInputDialog.getText(self, "Дата занятия",
                                        "Дата (дд.мм.гггг):", text=l.date or "")
        if not ok:
            return
        num, ok = QInputDialog.getInt(self, "Номер занятия", "№:", l.number, 1, 999)
        if not ok:
            return
        l.topic = topic.strip()
        l.date = date.strip()
        l.number = num
        try:
            self.book.save_to_db()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить: {e}")
            return
        try:
            from sync_runner import trigger
            trigger()
        except Exception:
            pass
        self._update_table()

    def _ask_retake(self, lesson, n):
        rd = ui_date.ask_date(self, "Дата пересдачи", "Когда пересдача:",
            default=ui_date.plus_days_str(7), min_today=True)
        if rd:
            self.book.set_retake_date(lesson.id, rd, n); self._update_table()

    def _delete_lesson(self, lesson_id):
        if QMessageBox.question(self, "Удалить занятие?", "Удалить этот столбец?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        #delete_lesson ставит надгробие в БД (а не просто убирает из списка) —
        #раньше столбец возвращался после перезагрузки и не удалялся на других ПК.
        self.book.delete_lesson(lesson_id); self._update_table()

    def _journal_ready(self) -> bool:
        """Журнал открыт (выбраны предмет и группа)? Если нет — self.book пуст, и
        добавление молча падало бы: в .exe без консоли исключение не видно. Поэтому
        проверяем заранее и понятно сообщаем пользователю."""
        if not self.book:
            QMessageBox.warning(self, "Журнал не открыт",
                "Сначала выберите предмет и группу на вкладке «Журнал».")
            return False
        return True

    def _add_lesson(self):
        if not self._journal_ready():
            return
        t, ok = QInputDialog.getItem(self, "Тип", "Выберите тип:",
                                     ["Лекция", "Практика", "ДЗ"], 0, False)
        if not ok: return
        d = ui_date.ask_date(self, "Дата занятия", "Когда занятие:",
                             default=ui_date.today_str())
        if not d: return
        #У ДЗ «тема» — это текст самого задания: он же уходит студентам в уведомление.
        if t == "ДЗ":
            tp, ok3 = QInputDialog.getText(self, "Домашнее задание",
                                           "Что задано (например: создать базу данных):")
        else:
            tp, ok3 = QInputDialog.getText(self, "Тема", "Тема занятия:")
        if not ok3: return
        try:
            self.book.add_lesson(t, topic=tp.strip(), date=d)
            self._update_table()
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Не удалось добавить занятие", str(e))

    def _add_exam(self):
        if not self._journal_ready():
            return
        d = ui_date.ask_date(self, "Дата экзамена", "Когда экзамен:",
                             default=ui_date.today_str())
        if not d: return
        tp, ok2 = QInputDialog.getText(self, "Тема", "Тема экзамена:")
        if not ok2: return
        try:
            l = self.book.add_lesson("Экзамен", topic=tp.strip(), date=d)
            l.retake_date = ""
            self.book.save_to_db(); self._update_table()
        except Exception as e:
            import traceback; traceback.print_exc()
            QMessageBox.critical(self, "Не удалось добавить экзамен", str(e))

    def _add_student(self):
        #Если журнал ещё не открыт — предупреждаем
        if not self.book:
            QMessageBox.warning(
                self, "Журнал не открыт",
                "Сначала откройте журнал:\nвыберите предмет и группу на вкладке «Журнал»."
            )
            return
        sn, ok = QInputDialog.getText(self, "Фамилия", "Введите фамилию:")
        if not ok or not sn.strip(): return
        nm, ok2 = QInputDialog.getText(self, "Имя", "Введите имя:")
        if not ok2: return
        sn, nm = sn.strip(), nm.strip()
        if any(s.f.lower() == sn.lower() for s in self.book.spisok_stud):
            QMessageBox.information(self, "Уже есть", f"{sn} уже в списке"); return
        st = Student(nm, sn, self.book.group)
        self.book.add_student(st)
        gh = get_store()
        if gh:
            try:
                ss = gh.get_students()
                if not any(s.get("surname", "").lower() == sn.lower() and
                           s.get("name", "").lower() == nm.lower() for s in ss):
                    ss.append({"surname": sn, "name": nm, "group": self.book.group})
                    gh.set_students(ss)
            except Exception as e:
                log.get("teacher_dashboard").warning(f"[GH add student] {e}")
        self._update_table()

    def _save(self):
        if self.book: self.book.save_to_db()
        QMessageBox.information(self, "Сохранено", "Данные сохранены.")

    #Экспорт / аттестация / ведомость ────────────────────────────────────────────
    def _ask_format(self) -> str:
        """Спрашивает формат выгрузки — как на вебе (Excel / Word). Возвращает
        'xlsx' | 'docx' | '' (отмена)."""
        dlg = QMessageBox(self)
        dlg.setWindowTitle("Формат выгрузки")
        dlg.setText("В каком формате сохранить?")
        bx = dlg.addButton("📊 Excel (.xlsx)", QMessageBox.AcceptRole)
        bw = dlg.addButton("📄 Word (.docx)", QMessageBox.AcceptRole)
        dlg.addButton("Отмена", QMessageBox.RejectRole)
        dlg.exec()
        clicked = dlg.clickedButton()
        if clicked is bx:
            return "xlsx"
        if clicked is bw:
            return "docx"
        return ""

    def _teacher_display_name(self) -> str:
        """ФИО преподавателя для подписи ведомости (логин как fallback)."""
        d = self.teacher_data or {}
        return (d.get("full_name") or d.get("name") or self.teacher_name or "").strip()

    def _journal_rows(self) -> list:
        """Строки журнала для экспорта: [{surname, name, records, average}]."""
        rows = []
        for s in self.book.spisok_stud:
            rows.append({"surname": s.f, "name": s.n, "records": s.records,
                         "average": round(self.book.calculate_average(s), 2)})
        return rows

    def _patronymic_map(self) -> dict:
        """Отчества студентов из общего хранилища по (фамилия|имя) — для ведомости."""
        out = {}
        try:
            store = get_store()
            for st in (store.get_students() if store else []):
                key = f"{st.get('surname', '')}|{st.get('name', '')}"
                if st.get("patronymic"):
                    out[key] = st.get("patronymic", "")
        except Exception:
            pass
        return out

    def _export_journal(self):
        """Экспорт журнала: выбор формата (Excel/Word), как на вебе."""
        if not self._journal_ready():
            return
        fmt = self._ask_format()
        if not fmt:
            return
        ext = "xlsx" if fmt == "xlsx" else "docx"
        flt = "Excel (*.xlsx)" if fmt == "xlsx" else "Word (*.docx)"
        default = f"Журнал_{self.book.group}_{self.book.subject}.{ext}".replace("/", "-")
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить журнал", default, flt)
        if not path:
            return
        try:
            if fmt == "xlsx":
                self.book.export_to_excel(path)
            else:
                import exports
                data = exports.build_journal_docx(
                    self.book.group, self.book.subject, self.book.lessons,
                    self._journal_rows())
                with open(path, "wb") as f:
                    f.write(data)
            QMessageBox.information(self, "Журнал", "Готово!")
        except ImportError:
            QMessageBox.critical(self, "Нет пакета",
                "Для Word нужен пакет python-docx:\n\npip install python-docx")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _vedomost_rows(self, term_grades: dict) -> list:
        """Строки ведомости: студенты группы + итоговая оценка + отчество."""
        patr = self._patronymic_map()
        rows = []
        for s in self.book.spisok_stud:
            key = f"{s.f}|{s.n}"
            rows.append({"surname": s.f, "name": s.n,
                         "patronymic": patr.get(key, ""),
                         "grade": (term_grades.get(key) or {}).get("grade", "")})
        return rows

    def _export_vedomost(self):
        """Экзаменационно-зачётная ведомость (аттестация): форма контроля → формат."""
        if not self._journal_ready():
            return
        if not self.book.spisok_stud:
            QMessageBox.information(self, "Нет студентов", "В группе нет студентов.")
            return
        form, ok = QInputDialog.getItem(self, "Форма контроля",
            "Форма промежуточной аттестации:", ["экзамен", "зачёт", "диффзачёт"], 0, False)
        if not ok:
            return
        fmt = self._ask_format()
        if not fmt:
            return
        from core import DBManager
        year, sem = self._selected_term()   #ведомость — за выбранный семестр (в т.ч. архив)
        tg = DBManager.get_term_grades(self.book.subject, year, sem)
        rows = self._vedomost_rows(tg)
        term = {"year": year, "semester": sem}
        ext = "xlsx" if fmt == "xlsx" else "docx"
        flt = "Excel (*.xlsx)" if fmt == "xlsx" else "Word (*.docx)"
        default = f"Ведомость_{self.book.group}_{self.book.subject}.{ext}".replace("/", "-")
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить ведомость", default, flt)
        if not path:
            return
        try:
            import exports
            teacher = self._teacher_display_name()
            if fmt == "xlsx":
                data = exports.build_vedomost_xlsx(self.book.group, self.book.subject,
                                                   term, form, rows, teacher=teacher)
            else:
                data = exports.build_vedomost_docx(self.book.group, self.book.subject,
                                                   term, form, rows, teacher=teacher)
            with open(path, "wb") as f:
                f.write(data)
            QMessageBox.information(self, "Ведомость", "Готово!")
        except ImportError:
            QMessageBox.critical(self, "Нет пакета",
                "Для Word нужен пакет python-docx:\n\npip install python-docx")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    def _open_attestation(self):
        """Диалог итоговых оценок за семестр (промежуточная аттестация). Пишет в
        term_grades (offline-first) и будит синк — данные общие с сайтом."""
        if not self._journal_ready():
            return
        if not self.book.spisok_stud:
            QMessageBox.information(self, "Нет студентов", "В группе нет студентов.")
            return
        import terms
        from core import DBManager
        year, sem = self._selected_term()   #аттестация — за выбранный (обычно текущий) семестр
        existing = DBManager.get_term_grades(self.book.subject, year, sem)
        dlg = AttestationDialog(self, self.book.group, self.book.subject,
                                terms.semester_label(sem), year,
                                self.book.spisok_stud, existing)
        if dlg.exec() != QDialog.Accepted:
            return
        try:
            wrote = False
            for (f, n), (grade, form) in dlg.result_grades().items():
                had = f"{f}|{n}" in existing
                #Пустую оценку пишем (надгробием) ТОЛЬКО если она была раньше — снятие.
                #Для тех, кому оценку так и не поставили, лишних надгробий не плодим.
                if not (grade or "").strip() and not had:
                    continue
                DBManager.set_term_grade(f, n, self.book.subject, year, sem, grade, form)
                wrote = True
            if wrote:
                from sync_runner import trigger
                trigger()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            return
        QMessageBox.information(self, "Аттестация", "Итоговые оценки сохранены.")

    def _import_excel(self):
        """Импорт журнала из Excel — АДАПТИВНЫЙ к разным форматам файла.

        Раньше формат был жёстким (столбец 0 = фамилия, 1 = имя, дальше занятия строго
        по порядку) — чужие таблицы и даже наш собственный экспорт (с титульной шапкой)
        не читались. Теперь:
          • строку заголовков ищем в первых 15 строках (по слову «фамилия»/«фио»);
          • имя студента понимаем и как две колонки («Фамилия»+«Имя»), и как слитное
            «ФИО» (первое слово — фамилия, остальное — имя);
          • колонки занятий сопоставляем ПО ЗАГОЛОВКУ («Практика №2», «лек 1», …), а не
            по позиции; колонки «Пересдача» пишутся в <id>_retake соответствующего
            экзамена; колонка «Средний балл» пропускается;
          • если по заголовкам ничего не распозналось — прежний позиционный порядок;
          • «—»/«-» считаются пустыми (наш экспорт так помечает ненужные пересдачи)."""
        if not self.book:
            QMessageBox.warning(self, "Ошибка", "Сначала откройте журнал."); return
        path, _ = QFileDialog.getOpenFileName(self, "Открыть Excel", "", "Excel (*.xlsx *.xls)")
        if not path: return
        try:
            import re as _re
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True); ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            #1) Строка заголовков: первые 15 строк, ячейка со словом «фамилия»/«фио».
            hdr_i = None
            for i, row in enumerate(rows[:15]):
                cells = [str(c).strip().lower() for c in (row or []) if c is not None]
                if any(("фамилия" in c or "фио" in c or c == "студент") for c in cells):
                    hdr_i = i; break
            if hdr_i is None or len(rows) <= hdr_i + 1:
                QMessageBox.warning(self, "Ошибка",
                    "Не нашёл строку заголовков (нужна колонка «Фамилия» или «ФИО»)."); return
            hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
            low = [h.lower() for h in hdr]

            #2) Колонки имени: отдельные «Фамилия»+«Имя» либо слитное «ФИО».
            fio_col = next((i for i, h in enumerate(low) if "фио" in h or h == "студент"), None)
            sn_col = next((i for i, h in enumerate(low) if "фамилия" in h), fio_col)
            nm_col = next((i for i, h in enumerate(low) if h.startswith("имя")), None)
            combined = nm_col is None                    #одна колонка с полным ФИО

            #3) Колонки занятий: по заголовку → конкретное занятие журнала.
            TYPE_ALIASES = {"лек": "Лекция", "пр": "Практика", "лаб": "Лабораторная",
                            "сем": "Семинар", "экз": "Экзамен", "зач": "Зачёт",
                            "конс": "Консультация"}
            def _match_lesson(header: str):
                h = header.lower()
                m = _re.search(r"№?\s*(\d+)", h)
                if not m:
                    return None
                num = int(m.group(1))
                ltype = next((full for pref, full in TYPE_ALIASES.items()
                              if h.startswith(pref) or f" {pref}" in h), None)
                for l in self.book.lessons:
                    if l.number == num and (ltype is None or l.type.lower().startswith(ltype.lower()[:3])):
                        return l
                return None

            col_map = {}                                  #индекс колонки → ключ records
            last_exam = None
            for ci, h in enumerate(hdr):
                if ci in (sn_col, nm_col) or not h or "средн" in low[ci]:
                    continue
                if "пересдач" in low[ci]:
                    if last_exam is not None:
                        col_map[ci] = last_exam.id + "_retake"
                    continue
                l = _match_lesson(h)
                if l is not None:
                    col_map[ci] = l.id
                    last_exam = l if l.type == "Экзамен" else last_exam
            if not col_map:
                #заголовки не распознались — прежний позиционный порядок
                start = (max(sn_col if sn_col is not None else 0,
                             nm_col if nm_col is not None else 0)) + 1
                for k, l in enumerate(self.book.lessons):
                    col_map[start + k] = l.id

            #4) Строки студентов.
            added = updated = 0
            for row in rows[hdr_i + 1:]:
                if not row or sn_col is None or sn_col >= len(row) or not row[sn_col]:
                    continue
                raw = str(row[sn_col]).strip()
                if raw.lower().startswith("средний"):     #итоговая строка нашего экспорта
                    continue
                if combined:
                    parts = raw.split()
                    sn, nm = parts[0], " ".join(parts[1:])
                else:
                    sn = raw
                    nm = str(row[nm_col]).strip() if nm_col < len(row) and row[nm_col] else ""
                s = next((x for x in self.book.spisok_stud
                          if x.f.lower() == sn.lower()
                          and (not nm or x.n.lower() == nm.lower())), None)
                if not s:
                    s = Student(nm, sn, self.book.group)
                    self.book.spisok_stud.append(s); added += 1
                for ci, key in col_map.items():
                    if ci >= len(row) or row[ci] is None:
                        continue
                    val = str(row[ci]).strip()
                    if not val or val in ("—", "-"):
                        continue
                    s.records[key] = val; updated += 1

            self.book.save_to_db(); self._update_table()
            QMessageBox.information(self, "Импорт",
                f"Добавлено студентов: {added}\nОбновлено записей: {updated}")
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))

    #Студенты

    def _build_students(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setContentsMargins(24, 20, 24, 20); lay.setSpacing(12)
        lay.addWidget(title_lbl("Студенты группы"))
        btn_row = QHBoxLayout()
        add_b = btn("+ Добавить", "green"); add_b.clicked.connect(self._add_student)
        btn_row.addWidget(add_b); btn_row.addStretch()
        lay.addLayout(btn_row)
        self._stud_table = QTableWidget()
        self._stud_table.setColumnCount(4)
        self._stud_table.setHorizontalHeaderLabels(["Фамилия", "Имя", "Средний балл", ""])
        self._stud_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._stud_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        lay.addWidget(self._stud_table, 1)
        self.pages["students"] = w; self.stack.addWidget(w)

    def _build_settings(self):
        """Вкладка «Настройки»: оформление, микрофон и озвучка Вектора (раньше в «Профиле»)."""
        from theme_ui import ThemeCustomizer
        import theme_service
        w = QWidget(); lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0); lay.setSpacing(0)

        identity = {"name": self.teacher_name}
        spec = theme_service.current_spec("teacher", identity)

        def _save(s):
            theme_service.save_user_theme(s, "teacher", identity)
            self._request_theme_rebuild()

        lay.addWidget(ThemeCustomizer(initial_spec=spec, on_save=_save))
        try:
            from vector.voice_ui import MicSelectorWidget, mic_available
            if mic_available():
                lay.addWidget(MicSelectorWidget())
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[settings] выбор микрофона недоступен: {e}")
        try:
            from vector.voice_ui import TtsSettingsWidget
            lay.addWidget(TtsSettingsWidget())
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[settings] настройки озвучки недоступны: {e}")
        self.pages["settings"] = w; self.stack.addWidget(w)

    #Курирование (ТОЛЬКО ЧТЕНИЕ)

    def _curated_groups(self) -> list:
        """Группы, которые преподаватель курирует. Пусто → он не куратор."""
        return [g for g in (self.teacher_data.get("curated_groups") or []) if g]

    def _build_curator(self):
        """Вкладка «Курирование»: группы → предметы → студенты с оценками.

        ⚠️ СТРОГО ТОЛЬКО ЧТЕНИЕ, как и на вебе (`/web/curator/*`). Куратор видит ВСЕ
        предметы своей группы, включая чужие, — но не ставит оценки: журнал ведёт
        преподаватель предмета. Поэтому здесь нет ни одной кнопки правки, а таблица
        нередактируема. Ослабить это нельзя: иначе куратор сможет менять чужие оценки
        мимо проверки прав (§6)."""
        w = QWidget(); lay = QVBoxLayout(w)
        lay.setContentsMargins(24, 20, 24, 20); lay.setSpacing(12)
        lay.addWidget(title_lbl("Курирование", 20))
        lay.addWidget(lbl("Успеваемость курируемых групп по всем предметам "
                          "(только просмотр)", 12, C['text3']))

        row = QHBoxLayout()
        self._cur_group_combo = combo(self._curated_groups())
        self._cur_subj_combo = combo([])
        self._cur_group_combo.currentTextChanged.connect(self._on_curator_group)
        self._cur_subj_combo.currentTextChanged.connect(self._refresh_curator)
        row.addWidget(lbl("Группа:", 12, C['text3'])); row.addWidget(self._cur_group_combo)
        row.addWidget(lbl("Предмет:", 12, C['text3'])); row.addWidget(self._cur_subj_combo)
        row.addStretch()
        lay.addLayout(row)

        self._cur_hint = lbl("", 12, C['text3'])
        lay.addWidget(self._cur_hint)

        self._cur_table = QTableWidget()
        self._cur_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._cur_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self._cur_table, 1)
        self.pages["curator"] = w; self.stack.addWidget(w)
        self._on_curator_group()

    def _curator_subjects(self, group: str) -> list:
        """Предметы, по которым у группы РЕАЛЬНО есть занятия.

        Берём из самих занятий, а не из справочника предметов: справочник и журнал
        расходятся (предмет могли переименовать или не завести вовсе), и тогда куратор
        не увидел бы существующие оценки — то есть решил бы, что их нет. Данные важнее
        справочника.

        Доступ к базе — через DBManager, как и везде на десктопе (инвариант §9)."""
        from core import DBManager
        #Соединение ОБЯЗАТЕЛЬНО закрываем: get_conn() открывает новое на каждый вызов, и
        #брошенное живёт до сборщика мусора. На каждое переключение группы это утечка,
        #а на Windows ещё и держит файл базы открытым.
        conn = None
        try:
            conn = DBManager.get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT DISTINCT subject FROM lessons "
                "WHERE group_name=? AND subject<>'' "
                "AND (deleted IS NULL OR deleted=0) ORDER BY subject",
                (group,))
            return [r[0] for r in cur.fetchall() if r[0]]
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[curator] предметы группы {group}: {e}")
            return []
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    def _on_curator_group(self):
        group = self._cur_group_combo.currentText()
        subjects = self._curator_subjects(group) if group else []
        self._cur_subj_combo.blockSignals(True)
        self._cur_subj_combo.clear(); self._cur_subj_combo.addItems(subjects)
        self._cur_subj_combo.blockSignals(False)
        self._refresh_curator()

    def _refresh_curator(self):
        group = self._cur_group_combo.currentText()
        subject = self._cur_subj_combo.currentText()
        table = self._cur_table
        if not (group and subject):
            table.setRowCount(0); table.setColumnCount(0)
            self._cur_hint.setText("Для этой группы пока нет занятий." if group else "")
            return
        try:
            book = GradeBook(group, subject)
        except Exception as e:
            log.get("teacher_dashboard").warning(f"[curator] книга {group}/{subject}: {e}")
            table.setRowCount(0); table.setColumnCount(0)
            self._cur_hint.setText("Не удалось прочитать журнал группы.")
            return

        lessons = book.lessons
        students = book.spisok_stud
        self._cur_hint.setText(f"Студентов: {len(students)} · занятий: {len(lessons)}")
        table.setColumnCount(3 + len(lessons))
        table.setHorizontalHeaderLabels(
            ["Фамилия", "Имя"] + [f"{l.type} №{l.number}" for l in lessons] + ["Средн."])
        table.setRowCount(len(students))

        #Средний балл — ТОЛЬКО через grading.py (инвариант §8): своя формула тут
        #разошлась бы с журналом преподавателя и с сайтом.
        import grading
        try:
            cfg = get_store()._config()
        except Exception:
            cfg = {}
        pairs = grading.pairs_from_objects(lessons)
        for r, st in enumerate(students):
            table.setItem(r, 0, QTableWidgetItem(st.f))
            table.setItem(r, 1, QTableWidgetItem(st.n))
            for c, l in enumerate(lessons, start=2):
                table.setItem(r, c, QTableWidgetItem(st.records.get(l.id, "")))
            avg = grading.practice_average(pairs, st.records, cfg)
            table.setItem(r, 2 + len(lessons),
                          QTableWidgetItem(f"{avg:.2f}" if avg else "—"))

    def _build_schedule(self):
        """Вкладка «Расписание» — пары преподавателя (инверсия групповых расписаний
        колледжа с портала ВСГУТУ). ФИО на сайте подбирается автоматически по имени
        аккаунта; при расхождении преподаватель выбирает себя из списка."""
        from schedule_view import ScheduleView
        w = ScheduleView(role="teacher", login=self.teacher_name,
                         app_hint=self.teacher_name)
        self.pages["schedule"] = w; self.stack.addWidget(w)

    def _refresh_students(self):
        if not self.book: return
        students = self.book.spisok_stud
        self._stud_table.setRowCount(len(students))
        for r, s in enumerate(students):
            gs = []
            for l in self.book.lessons:
                if l.type in ("Практика", "Экзамен"):
                    v = s.records.get(l.id, "")
                    try: gs.append(int(v.split()[0]))
                    except: pass
            avg = f"{sum(gs)/len(gs):.1f}" if gs else "—"
            self._stud_table.setItem(r, 0, QTableWidgetItem(s.f))
            self._stud_table.setItem(r, 1, QTableWidgetItem(s.n))
            self._stud_table.setItem(r, 2, QTableWidgetItem(avg))
            del_b = QPushButton("✕"); del_b.setStyleSheet(BTN["sm_red"])
            del_b.clicked.connect(lambda _, f=s.f, n=s.n: self._del_student(f, n))
            self._stud_table.setCellWidget(r, 3, del_b)

    def _del_student(self, f, n):
        if QMessageBox.question(self, "Удалить?", f"Удалить {f} {n}?",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        self.book.delete_student(f, n)
        gh = get_store()
        if gh:
            try:
                ss = [x for x in gh.get_students()
                      if not (x.get("surname", "") == f and x.get("name", "") == n)]
                gh.set_students(ss)
            except Exception as e:
                log.get("teacher_dashboard").warning(f"[GH del student] {e}")
        self._refresh_students(); self._update_table()

    #Статистика

    def _build_stats(self):
        w = QScrollArea(); w.setWidgetResizable(True); w.setStyleSheet("border:none;")
        inner = QWidget()
        lay = QVBoxLayout(inner); lay.setContentsMargins(24, 20, 24, 20); lay.setSpacing(14)
        lay.addWidget(title_lbl("Статистика группы"))
        self._tstat_content = QVBoxLayout(); lay.addLayout(self._tstat_content)
        lay.addStretch()
        w.setWidget(inner)
        self.pages["stats"] = w; self.stack.addWidget(w)

    def _refresh_stats(self):
        while self._tstat_content.count():
            it = self._tstat_content.takeAt(0)
            if it.widget(): it.widget().deleteLater()
        if not self.book: return
        students = self.book.spisok_stud
        dist = {2: 0, 3: 0, 4: 0, 5: 0}; att_rows = []
        for s in students:
            gs = []; lt = lp = 0
            for l in self.book.lessons:
                if l.type in ("Практика", "Экзамен"):
                    v = s.records.get(l.id, "")
                    try: g = int(v.split()[0]); gs.append(g); dist[g] += 1
                    except: pass
                elif l.type == "Лекция":
                    lt += 1
                    if s.records.get(l.id, "") != "Н": lp += 1
            avg = sum(gs) / len(gs) if gs else 0
            att = int(lp / lt * 100) if lt else 100
            att_rows.append((s.f, att, avg))
        all_g     = sum(dist.values())
        total_avg = f"{sum(k*v for k,v in dist.items())/all_g:.1f}" if all_g else "—"
        sr = QHBoxLayout(); sr.setSpacing(10)
        for l, v, c in [
            ("Студентов",  str(len(students)), "green"),
            ("Средний балл", total_avg, "blue"),
            ("Оценок",     str(all_g),  "text"),
            ("Занятий",    str(len(self.book.lessons)), "text"),
        ]:
            sr.addWidget(stat_card(l, v, c))
        self._tstat_content.addLayout(sr)
        #Распределение оценок
        dc = card(); dl = QVBoxLayout(dc); dl.setContentsMargins(18, 16, 18, 16)
        dl.addWidget(section_lbl("Распределение оценок"))
        cols = {5: C['green'], 4: C['blue'], 3: C['yellow'], 2: C['red']}
        for g in [5, 4, 3, 2]:
            cnt = dist[g]; pct = int(cnt / all_g * 100) if all_g else 0
            row = QHBoxLayout()
            row.addWidget(lbl(str(g), 14, cols[g], True))
            bar = QFrame(); bar.setFixedHeight(4)
            bar.setStyleSheet(f"background:{C['border']};border-radius:2px;")
            row.addWidget(bar, 1)
            row.addWidget(lbl(f"{cnt} ({pct}%)", 12, C['text3']))
            dl.addLayout(row)
        self._tstat_content.addWidget(dc)
        #Посещаемость
        ac = card(); al = QVBoxLayout(ac); al.setContentsMargins(18, 16, 18, 16)
        al.addWidget(section_lbl("Посещаемость по студентам"))
        for f, att, _avg_s in sorted(att_rows, key=lambda x: -x[1]):
            c = C['green'] if att >= 90 else C['blue'] if att >= 70 else C['red']
            row = QHBoxLayout()
            row.addWidget(lbl(f, 12, C['text'])); row.setStretch(0, 1)
            bar = QFrame(); bar.setFixedHeight(4)
            bar.setStyleSheet(f"background:{C['border']};border-radius:2px;")
            row.addWidget(bar, 2)
            row.addWidget(lbl(f"{att}%", 12, c, True))
            al.addLayout(row)
        self._tstat_content.addWidget(ac)

    #ИИ помощник

    def _build_ai(self):
        """Вкладка «ИИ Помощник» — Вектор (офлайн / GigaChat / Ollama).
        Делит ОБЩУЮ сессию со шторкой — переписка единая."""
        try:
            from vector.widget import VectorPanel
            self._ensure_vector_session()
            ai = VectorPanel(self.vector_session, docked=False)
        except Exception as _e:
            log.get("teacher_dashboard").warning(f"[Vector] вкладка не собралась (препод): {_e}")
            ai = vector_unavailable_widget()
        self.pages["ai"] = ai; self.stack.addWidget(ai)

    def _build_messenger(self):
        """Вкладка «Сообщения» — веб-мессенджер во встроенном веб-view (Фаза 0 «один UI»,
        см. ui/messenger_web.py). Онлайн-подсистема (§13), offline-first не нужен."""
        try:
            from messenger_web import MessengerWebPanel
            mv = MessengerWebPanel("/teacher/messages", "teacher")
        except Exception as _e:
            log.get("teacher_dashboard").warning(f"[messenger] вкладка не собралась: {_e}")
            from PySide6.QtWidgets import QLabel
            mv = QLabel("Сообщения недоступны")
        self.pages["messages"] = mv
        self.stack.addWidget(mv)

    def _build_profile(self):
        """Вкладка «Профиль»: сведения об аккаунте + уведомления."""
        from notifications_view import NotificationsView
        w = QWidget(); lay = QVBoxLayout(w)
        lay.setContentsMargins(24, 20, 24, 20); lay.setSpacing(12)
        lay.addWidget(title_lbl("Профиль", 20))

        acc = card()
        acc_lay = QVBoxLayout(acc); acc_lay.setContentsMargins(16, 14, 16, 14); acc_lay.setSpacing(2)
        acc_lay.addWidget(lbl(self.teacher_name or "Преподаватель", 16, C['text'], bold=True))
        acc_lay.addWidget(lbl("Преподаватель", 12, C['text3']))
        lay.addWidget(acc)

        from avatar_dialog import AvatarSection
        lay.addWidget(AvatarSection("teacher", {"name": self.teacher_name}))

        #Заголовок «Уведомления» уже внутри NotificationsView — без отдельного section_lbl,
        #иначе задваивается (правка по отчёту).
        #§12: свои группы — из назначений предмет→группа (тот же источник, что и у Вектора,
        #см. _ensure_vector_session), чтобы кнопка «Мероприятие» предлагала верный список.
        ga = self.teacher_data.get("group_assignments", {}) or {}
        my_groups = sorted({g for g in ga.values() if g})
        lay.addWidget(NotificationsView(role="teacher", groups=my_groups), 1)
        self.pages["profile"] = w; self.stack.addWidget(w)

    def _request_theme_rebuild(self):
        """Просим главное окно пересобрать дашборд — чтобы новая тема применилась
        целиком (инлайн-стили перекрашиваются только при пересборке)."""
        win = self.window()
        if hasattr(win, "reapply_current"):
            win.reapply_current()

    #Роутинг

    def _switch(self, key):
        if getattr(self, '_switching', False):
            return
        self._switching = True
        try:
            if key == "stats":    self._refresh_stats()
            if key == "students": self._refresh_students()
            #Шторка Вектора прячется на «ИИ» и «Сообщениях» (там Вектор лишний, а речь
            #должна замолчать — hush_if_vector_hidden ниже это доделает), на остальных — вернётся.
            dock = getattr(self, "vector_dock", None)
            if dock is not None:
                dock.suspend() if key in ("ai", "messages") else dock.resume()
            if key in self.pages:
                self.stack.setCurrentWidget(self.pages[key])
                self.sidebar.set_active(key)
            #Ушли туда, где Вектора не видно (шторка свёрнута) → глушим озвучку.
            from vector.widget import hush_if_vector_hidden
            hush_if_vector_hidden(dock, key)
        finally:
            self._switching = False


class AttestationDialog(QDialog):
    """Диалог выставления ИТОГОВЫХ оценок за семестр (промежуточная аттестация) —
    порт модалки веба (TeacherJournal.vue). Список студентов + форма контроля;
    сохранение делает вызывающий (пишет в term_grades и будит синк)."""

    GRADES = ["", "5", "4", "3", "2", "Зачтено", "Не зачтено"]
    FORMS = ["экзамен", "зачёт", "диффзачёт"]

    def __init__(self, parent, group, subject, sem_label, year, students, existing):
        super().__init__(parent)
        self.setWindowTitle("🎓 Итоговые оценки за семестр")
        self.setMinimumWidth(460)
        self._rows = []            #[(surname, name, QComboBox)]

        lay = QVBoxLayout(self); lay.setContentsMargins(20, 18, 20, 18); lay.setSpacing(10)
        lay.addWidget(title_lbl("Итоговые оценки за семестр", 16))
        lay.addWidget(lbl(f"{group} · {subject} · {year}, {sem_label} семестр", 12, C['text3']))

        #Форма контроля (общая для всей ведомости) — предзаполняем из уже сохранённого.
        form_row = QHBoxLayout()
        form_row.addWidget(lbl("Форма контроля:", 12, C['text3']))
        self._form_combo = combo(self.FORMS)
        saved_form = next((v.get("form") for v in existing.values() if v.get("form")), "")
        if saved_form:
            idx = self._form_combo.findText(saved_form)
            if idx >= 0:
                self._form_combo.setCurrentIndex(idx)
        form_row.addWidget(self._form_combo); form_row.addStretch()
        lay.addLayout(form_row)

        #Прокручиваемый список студентов с комбобоксом оценки у каждого.
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setStyleSheet("border:none;")
        inner = QWidget(); grid = QVBoxLayout(inner); grid.setContentsMargins(0, 0, 6, 0); grid.setSpacing(6)
        for s in students:
            r = QHBoxLayout()
            r.addWidget(lbl(f"{s.f} {s.n}", 13, C['text']), 1)
            cb = combo(self.GRADES)
            cur = (existing.get(f"{s.f}|{s.n}") or {}).get("grade", "")
            if cur:
                idx = cb.findText(cur)
                if idx >= 0:
                    cb.setCurrentIndex(idx)
            cb.setFixedWidth(140)
            r.addWidget(cb)
            grid.addLayout(r)
            self._rows.append((s.f, s.n, cb))
        grid.addStretch()
        scroll.setWidget(inner)
        lay.addWidget(scroll, 1)

        bb = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        bb.button(QDialogButtonBox.Save).setText("Сохранить")
        bb.button(QDialogButtonBox.Cancel).setText("Отмена")
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        lay.addWidget(bb)

    def result_grades(self) -> dict:
        """{(surname, name): (grade, form)} — что выставил преподаватель."""
        form = self._form_combo.currentText()
        return {(f, n): (cb.currentText(), form) for f, n, cb in self._rows}
